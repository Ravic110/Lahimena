"""
Excel file handling utilities
"""

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

import os
import re
import shutil
import zipfile
from datetime import datetime, time, timedelta
from time import monotonic

from config import (
    CLIENT_EXCEL_PATH,
    FINANCIAL_EXCEL_PATH,
    CLIENT_INFOS_SHEET_NAME,
    CLIENT_SHEET_NAME,
    COTATION_FRAIS_COL_SHEET_NAME,
    AVION_SHEET_NAME,
    AVION_SOURCE_SHEET_NAME,
    COTATION_H_SHEET_NAME,
    FRAIS_COLLECTIFS_SHEET_NAME,
    HOTEL_EXCEL_PATH,
    HOTEL_SHEET_NAME,
    VISITE_EXCURSION_SHEET_NAME,
    VISITE_EXCURSION_SOURCE_SHEET_NAME,
    TRANSPORT_SOURCE_SHEET_NAME,
    TRANSPORT_SHEET_NAME,
    KM_MADA_SHEET_NAME,
    PARAMETRAGE_SHEET_NAME,
    INVOICE_SHEET_NAME,
    FINANCIAL_STATE_SHEET_NAME,
)
from utils.cache import (
    cached_client_data,
    cached_hotel_data,
    invalidate_client_cache,
    invalidate_hotel_cache,
)
from utils.logger import logger


_KM_MADA_CACHE_TTL_SECONDS = 10.0
_KM_MADA_CACHE = {
    "path": None,
    "mtime": None,
    "loaded_at": 0.0,
    "rows": [],
    "lookup": {},
}
_THROTTLED_ERROR_STATE = {}
_THROTTLED_ERROR_WINDOW_SECONDS = 30.0


def _log_error_throttled(key, message, exc_info=True):
    """Log repeated errors at most once per throttling window."""
    now = monotonic()
    last = _THROTTLED_ERROR_STATE.get(key, 0.0)
    if now - last >= _THROTTLED_ERROR_WINDOW_SECONDS:
        _THROTTLED_ERROR_STATE[key] = now
        logger.error(message, exc_info=exc_info)


def _invalidate_km_mada_cache():
    _KM_MADA_CACHE["path"] = None
    _KM_MADA_CACHE["mtime"] = None
    _KM_MADA_CACHE["loaded_at"] = 0.0
    _KM_MADA_CACHE["rows"] = []
    _KM_MADA_CACHE["lookup"] = {}


def _parse_num(val):
    """Parse a cell value into int or float, stripping thousand separators and currency text.

    Returns 0 on failure or empty values.
    """
    if val is None or val == "":
        return 0
    if isinstance(val, (int, float)):
        return val
    try:
        s = str(val).strip()
        s = s.replace(",", "").replace(" ", "")
        m = re.search(r"-?\d+(?:\.\d+)?", s)
        if not m:
            return 0
        num_str = m.group(0)
        if "." in num_str:
            return float(num_str)
        return int(num_str)
    except Exception:
        return 0


def _parse_duration_hours(val):
    """Parse duration values into hours.

    Supports numeric hours, Excel time/timedelta values, and strings like
    "3h30", "03:30", "210 min", or "2.5".
    """
    if val is None or val == "":
        return 0.0

    if isinstance(val, timedelta):
        return max(0.0, float(val.total_seconds()) / 3600)

    if isinstance(val, time):
        return val.hour + (val.minute / 60.0) + (val.second / 3600.0)

    if isinstance(val, datetime):
        return val.hour + (val.minute / 60.0) + (val.second / 3600.0)

    if isinstance(val, (int, float)):
        hours = float(val)
        return hours if hours > 0 else 0.0

    raw = str(val).strip().lower().replace(",", ".")
    if not raw:
        return 0.0

    compact = re.sub(r"\s+", "", raw)

    match_h = re.fullmatch(r"(\d+(?:\.\d+)?)h(?:(\d+(?:\.\d+)?)(?:mn|min|m)?)?", compact)
    if match_h:
        hours = float(match_h.group(1))
        minutes = float(match_h.group(2)) if match_h.group(2) else 0.0
        return max(0.0, hours + (minutes / 60.0))

    match_clock = re.fullmatch(r"(\d{1,3}):(\d{1,2})(?::(\d{1,2}))?", compact)
    if match_clock:
        hh = int(match_clock.group(1))
        mm = int(match_clock.group(2))
        ss = int(match_clock.group(3) or 0)
        return max(0.0, hh + (mm / 60.0) + (ss / 3600.0))

    match_min = re.fullmatch(r"(\d+(?:\.\d+)?)(?:mn|min|m)", compact)
    if match_min:
        minutes = float(match_min.group(1))
        return max(0.0, minutes / 60.0)

    try:
        parsed = float(compact)
        return parsed if parsed > 0 else 0.0
    except Exception:
        return 0.0


def create_backup(filepath):
    """
    Create a backup of Excel file before modification

    Args:
        filepath (str): Path to the Excel file to backup

    Returns:
        str: Path to backup file, or None if failed
    """
    if not os.path.exists(filepath):
        logger.warning(f"File not found for backup: {filepath}")
        return None

    try:
        # Create backups directory
        backup_dir = os.path.join(os.path.dirname(filepath), "backups")
        os.makedirs(backup_dir, exist_ok=True)

        # Create backup filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = os.path.basename(filepath)
        backup_path = os.path.join(backup_dir, f"{filename}.{timestamp}.bak")

        # Copy file
        shutil.copy2(filepath, backup_path)
        logger.info(f"Backup created: {backup_path}")

        return backup_path
    except Exception as e:
        logger.error(f"Failed to create backup for {filepath}: {e}", exc_info=True)
        return None


def _ensure_headers(ws, headers, header_style=None):
    header_map = {}
    max_col = ws.max_column if ws.max_column and ws.max_column > 0 else 0
    for col in range(1, max_col + 1):
        value = ws.cell(row=1, column=col).value
        if value:
            header_map[str(value).strip()] = col

    if max_col == 1 and not header_map and ws.cell(row=1, column=1).value is None:
        max_col = 0

    next_col = max_col + 1 if max_col else 1
    for header in headers:
        if header not in header_map:
            cell = ws.cell(row=1, column=next_col, value=header)
            if header_style:
                if header_style.get("font"):
                    cell.font = header_style["font"]
                if header_style.get("fill"):
                    cell.fill = header_style["fill"]
                if header_style.get("alignment"):
                    cell.alignment = header_style["alignment"]
            header_map[header] = next_col
            next_col += 1

    return header_map


def _get_header_map(ws, header_row=1):
    header_map = {}
    max_col = ws.max_column if ws.max_column and ws.max_column > 0 else 0
    for col in range(1, max_col + 1):
        value = ws.cell(row=header_row, column=col).value
        if value is not None and str(value).strip() != "":
            header_map[str(value).strip()] = col
    return header_map


def _iter_grouped_columns(ws, group_row=1, header_row=2):
    columns = []
    last_group = ""
    max_col = ws.max_column if ws.max_column and ws.max_column > 0 else 0
    for col in range(1, max_col + 1):
        group_val = ws.cell(row=group_row, column=col).value
        if group_val is not None and str(group_val).strip() != "":
            last_group = str(group_val).strip()
        header_val = ws.cell(row=header_row, column=col).value
        if header_val is None or str(header_val).strip() == "":
            continue
        columns.append((last_group, str(header_val).strip(), col))
    return columns


def _first_available(data, keys, default=""):
    for key in keys:
        if key in data and data.get(key) not in (None, ""):
            return data.get(key)
    return default


def _normalize_header_key(value):
    """Normalize header labels for resilient matching."""
    if value is None:
        return ""
    normalized = str(value).strip().lower()
    replacements = str.maketrans(
        {
            "é": "e",
            "è": "e",
            "ê": "e",
            "ë": "e",
            "à": "a",
            "â": "a",
            "ä": "a",
            "î": "i",
            "ï": "i",
            "ô": "o",
            "ö": "o",
            "ù": "u",
            "û": "u",
            "ü": "u",
            "ç": "c",
            "'": " ",
            "_": " ",
            "-": " ",
        }
    )
    normalized = normalized.translate(replacements)
    normalized = re.sub(r"\s+", " ", normalized).strip()
    return normalized


def _find_header_column(header_map, *aliases):
    """Find a header column using exact or normalized aliases."""
    for alias in aliases:
        if alias in header_map:
            return header_map[alias]

    normalized_map = {_normalize_header_key(k): v for k, v in header_map.items()}
    for alias in aliases:
        col = normalized_map.get(_normalize_header_key(alias))
        if col:
            return col
    return None


def save_client_to_excel(client_data):
    """
    Save client data to Excel file

    Args:
        client_data (dict): Client data dictionary

    Returns:
        int: Row number where data was saved, or -1 if failed
    """
    if not OPENPYXL_AVAILABLE:
        logger.warning("openpyxl not available. Cannot save to Excel.")
        return -1

    # Create file if it doesn't exist
    client_headers = [
        "Date",
        "Réf. Client",
        "Type Client",
        "Prénom",
        "Nom",
        "Date Arrivée",
        "Date Départ",
        "Durée Séjour",
        "Nombre Participants",
        "Nombre Adultes",
        "Enfants 2-12",
        "Bébés 0-2",
        "Téléphone",
        "Téléphone WhatsApp",
        "Email",
        "Période",
        "Restauration",
        "Hébergement",
        "Chambre",
        "Enfant",
        "Âge Enfant",
        "Forfait",
        "Circuit",
        "Statut",
        "SGL",
        "DBL",
        "TWN",
        "TPL",
        "FML",
    ]
    client_header_style = {
        "font": Font(bold=True, color="FFFFFF"),
        "fill": PatternFill(
            start_color="27AE60", end_color="27AE60", fill_type="solid"
        ),
        "alignment": Alignment(horizontal="center"),
    }

    if not os.path.exists(CLIENT_EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = CLIENT_SHEET_NAME

        _ensure_headers(ws, client_headers, client_header_style)

        wb.save(CLIENT_EXCEL_PATH)

    # Open existing file
    wb = load_workbook(CLIENT_EXCEL_PATH)
    if CLIENT_SHEET_NAME not in wb.sheetnames:
        ws = wb.create_sheet(CLIENT_SHEET_NAME)
    else:
        ws = wb[CLIENT_SHEET_NAME]

    header_map = _ensure_headers(ws, client_headers, client_header_style)

    # Find last empty row (column A)
    last_row = 2
    while ws[f"A{last_row}"].value is not None:
        last_row += 1

    # Write data by headers for compatibility and extensibility
    field_map = {
        "Date": ["Timestamp", "Date", "date_jour"],
        "Réf. Client": ["Ref_Client", "ref_client"],
        "Type Client": ["Type_Client", "type_client"],
        "Prénom": ["Prénom", "prenom"],
        "Nom": ["Nom", "nom"],
        "Date Arrivée": ["Date_Arrivée", "date_arrivee"],
        "Date Départ": ["Date_Départ", "date_depart"],
        "Durée Séjour": ["Durée_Séjour", "duree_sejour"],
        "Nombre Participants": ["Nombre_Participants", "nombre_participants"],
        "Nombre Adultes": ["Nombre_Adultes", "nombre_adultes"],
        "Enfants 2-12": ["Enfants_2_12_ans", "nombre_enfants_2_12"],
        "Bébés 0-2": ["Bébés_0_2_ans", "nombre_bebes_0_2"],
        "Téléphone": ["Téléphone", "telephone"],
        "Téléphone WhatsApp": ["Téléphone_WhatsApp", "telephone_whatsapp"],
        "Email": ["Email", "email"],
        "Période": ["Période", "periode"],
        "Restauration": ["Restauration", "restauration"],
        "Hébergement": ["Hébergement", "hebergement"],
        "Chambre": ["Chambre", "chambre"],
        "Enfant": ["Enfant", "enfant"],
        "Âge Enfant": ["Âge_Enfant", "age_enfant"],
        "Forfait": ["Forfait", "forfait"],
        "Circuit": ["Circuit", "circuit"],
        "Statut": ["Statut", "statut"],
        "SGL": ["SGL_Count", "sgl_count"],
        "DBL": ["DBL_Count", "dbl_count"],
        "TWN": ["TWN_Count", "twn_count"],
        "TPL": ["TPL_Count", "tpl_count"],
        "FML": ["FML_Count", "fml_count"],
    }
    for header, keys in field_map.items():
        col = header_map.get(header)
        if col:
            ws.cell(
                row=last_row, column=col, value=_first_available(client_data, keys, "")
            )

    # Auto-adjust column widths
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for row_idx in range(1, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            value_len = len(str(value))
            if value_len > max_length:
                max_length = value_len
        ws.column_dimensions[column_letter].width = min(max_length + 2, 25)

    wb.save(CLIENT_EXCEL_PATH)

    # Invalidate cache after modification
    invalidate_client_cache()

    # Also store extended infos in dedicated sheet
    _save_client_infos_to_excel(client_data)

    return last_row


def _save_client_infos_to_excel(client_data):
    """Save extended client infos to the INFOS_CLIENTS sheet"""
    if not OPENPYXL_AVAILABLE:
        logger.warning("openpyxl not available. Cannot save client infos.")
        return False

    infos_headers = [
        "Date",
        "Réf. Client",
        "Numéro Dossier",
        "Type Client",
        "Prénom",
        "Nom",
        "Date Arrivée",
        "Date Départ",
        "Durée Séjour",
        "Nombre Participants",
        "Nombre Adultes",
        "Enfants 2-12",
        "Bébés 0-2",
        "Téléphone",
        "Téléphone WhatsApp",
        "Email",
        "Période",
        "Restauration",
        "Hébergement",
        "Chambre",
        "Statut",
        "Enfant",
        "Âge Enfant",
        "Heure Arrivée",
        "Heure Départ",
        "Compagnie",
        "Aéroport",
        "Réf. Externe",
        "Forfait",
        "Circuit",
        "Type Circuit",
        "ID Circuit",
        "Itinéraire Circuit",
        "Activité Circuit",
        "Durée Circuit",
        "Condition Physique Circuit",
        "Type Voiture Circuit",
        "Hôtels Défaut Circuit",
        "Prestations Incluses Circuit",
        "Transports Associés Circuit",
        "Ville Départ",
        "Ville Arrivée",
        "Type Hôtel Arrivée",
        "SGL",
        "DBL",
        "TWN",
        "TPL",
        "FML",
    ]
    infos_header_style = {
        "font": Font(bold=True, color="FFFFFF"),
        "fill": PatternFill(
            start_color="27AE60", end_color="27AE60", fill_type="solid"
        ),
        "alignment": Alignment(horizontal="center"),
    }

    if not os.path.exists(CLIENT_EXCEL_PATH):
        wb = Workbook()
        wb.save(CLIENT_EXCEL_PATH)

    wb = load_workbook(CLIENT_EXCEL_PATH)
    if CLIENT_INFOS_SHEET_NAME not in wb.sheetnames:
        ws = wb.create_sheet(CLIENT_INFOS_SHEET_NAME)
    else:
        ws = wb[CLIENT_INFOS_SHEET_NAME]

    header_map = _ensure_headers(ws, infos_headers, infos_header_style)

    # Find existing row by ref client
    ref_client = _first_available(client_data, ["Ref_Client", "ref_client"], "")
    target_row = None
    if ref_client:
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=header_map["Réf. Client"]).value == ref_client:
                target_row = row
                break

    if target_row is None:
        target_row = 2
        while ws[f"A{target_row}"].value is not None:
            target_row += 1

    field_map = {
        "Date": ["Timestamp", "Date", "date_jour"],
        "Réf. Client": ["Ref_Client", "ref_client"],
        "Numéro Dossier": ["Numero_Dossier", "numero_dossier"],
        "Type Client": ["Type_Client", "type_client"],
        "Prénom": ["Prénom", "prenom"],
        "Nom": ["Nom", "nom"],
        "Date Arrivée": ["Date_Arrivée", "date_arrivee"],
        "Date Départ": ["Date_Départ", "date_depart"],
        "Durée Séjour": ["Durée_Séjour", "duree_sejour"],
        "Nombre Participants": ["Nombre_Participants", "nombre_participants"],
        "Nombre Adultes": ["Nombre_Adultes", "nombre_adultes"],
        "Enfants 2-12": ["Enfants_2_12_ans", "nombre_enfants_2_12"],
        "Bébés 0-2": ["Bébés_0_2_ans", "nombre_bebes_0_2"],
        "Téléphone": ["Téléphone", "telephone"],
        "Téléphone WhatsApp": ["Téléphone_WhatsApp", "telephone_whatsapp"],
        "Email": ["Email", "email"],
        "Période": ["Période", "periode"],
        "Restauration": ["Restauration", "restauration"],
        "Hébergement": ["Hébergement", "hebergement"],
        "Chambre": ["Chambre", "chambre"],
        "Statut": ["Statut", "statut"],
        "Enfant": ["Enfant", "enfant"],
        "Âge Enfant": ["Âge_Enfant", "age_enfant"],
        "Heure Arrivée": ["Heure_Arrivee", "heure_arrivee"],
        "Heure Départ": ["Heure_Depart", "heure_depart"],
        "Compagnie": ["Compagnie", "compagnie"],
        "Aéroport": ["Aeroport", "aeroport"],
        "Réf. Externe": ["Ext_Ref", "ext_ref"],
        "Forfait": ["Forfait", "forfait"],
        "Circuit": ["Circuit", "circuit"],
        "Type Circuit": ["Type_Circuit", "type_circuit"],
        "ID Circuit": ["ID_Circuit", "id_circuit"],
        "Itinéraire Circuit": ["Itineraire_Circuit", "itineraire_circuit"],
        "Activité Circuit": ["Activite_Circuit", "activite_circuit"],
        "Durée Circuit": ["Duree_Circuit", "duree_circuit"],
        "Condition Physique Circuit": [
            "Condition_Physique_Circuit",
            "condition_physique_circuit",
        ],
        "Type Voiture Circuit": ["Type_Voiture_Circuit", "type_voiture_circuit"],
        "Hôtels Défaut Circuit": [
            "Hotels_Defaut_Villes_Circuit",
            "hotels_defaut_villes_circuit",
            "hotels_defaut_circuit",
        ],
        "Prestations Incluses Circuit": [
            "Prestations_Incluses_Circuit",
            "prestations_incluses_circuit",
        ],
        "Transports Associés Circuit": [
            "Transports_Associes_Circuit",
            "transports_associes_circuit",
        ],
        "Ville Départ": ["Ville_Depart", "ville_depart"],
        "Ville Arrivée": ["Ville_Arrivee", "ville_arrivee"],
        "Type Hôtel Arrivée": ["Type_Hotel_Arrivee", "type_hotel_arrivee"],
        "SGL": ["SGL_Count", "sgl_count"],
        "DBL": ["DBL_Count", "dbl_count"],
        "TWN": ["TWN_Count", "twn_count"],
        "TPL": ["TPL_Count", "tpl_count"],
        "FML": ["FML_Count", "fml_count"],
    }
    for header, keys in field_map.items():
        col = header_map.get(header)
        if col:
            ws.cell(
                row=target_row,
                column=col,
                value=_first_available(client_data, keys, ""),
            )

    wb.save(CLIENT_EXCEL_PATH)
    return True


@cached_client_data(ttl_seconds=3600)  # Cache for 1 hour
def load_all_clients():
    """
    Load all client data from Excel file
    Results are cached for 1 hour

    Returns:
        list: List of client dictionaries
    """
    if not OPENPYXL_AVAILABLE:
        logger.warning("openpyxl not available. Cannot load from Excel.")
        return []

    if not os.path.exists(CLIENT_EXCEL_PATH):
        return []

    wb = load_workbook(CLIENT_EXCEL_PATH)
    if CLIENT_SHEET_NAME not in wb.sheetnames:
        return []

    ws = wb[CLIENT_SHEET_NAME]
    header_map = _ensure_headers(ws, [])

    clients = []
    infos_map = _load_client_infos_map()
    # Start from row 2 (skip headers)
    for row in range(2, ws.max_row + 1):
        if ws[f"A{row}"].value is None:
            continue

        def _cell(header, fallback_cell=None):
            col = header_map.get(header)
            if col:
                return ws.cell(row=row, column=col).value
            if fallback_cell:
                return ws[fallback_cell].value
            return None

        client = {
            "row_number": row,
            "timestamp": _cell("Date", f"A{row}") or "",
            "ref_client": _cell("Réf. Client", f"B{row}") or "",
            "type_client": _cell("Type Client") or "",
            "prenom": _cell("Prénom") or "",
            "nom": _cell("Nom", f"C{row}") or "",
            "date_arrivee": _cell("Date Arrivée") or "",
            "date_depart": _cell("Date Départ") or "",
            "duree_sejour": _cell("Durée Séjour") or "",
            "nombre_participants": _cell("Nombre Participants") or "",
            "nombre_adultes": _cell("Nombre Adultes") or "",
            "nombre_enfants_2_12": _cell("Enfants 2-12") or "",
            "nombre_bebes_0_2": _cell("Bébés 0-2") or "",
            "telephone": _cell("Téléphone", f"D{row}") or "",
            "telephone_whatsapp": _cell("Téléphone WhatsApp") or "",
            "email": _cell("Email", f"E{row}") or "",
            "periode": _cell("Période", f"F{row}") or "",
            "restauration": _cell("Restauration", f"G{row}") or "",
            "hebergement": _cell("Hébergement", f"H{row}") or "",
            "chambre": _cell("Chambre", f"I{row}") or "",
            "enfant": _cell("Enfant", f"J{row}") or "",
            "age_enfant": _cell("Âge Enfant", f"K{row}") or "",
            "forfait": _cell("Forfait", f"L{row}") or "",
            "circuit": _cell("Circuit", f"M{row}") or "",
            "statut": _cell("Statut") or "En cours",
            "sgl_count": _cell("SGL") or "",
            "dbl_count": _cell("DBL") or "",
            "twn_count": _cell("TWN") or "",
            "tpl_count": _cell("TPL") or "",
            "fml_count": _cell("FML") or "",
        }
        ref_key = client.get("ref_client")
        if ref_key and ref_key in infos_map:
            main_status = client.get("statut") or "En cours"
            client.update(infos_map[ref_key])
            client["statut"] = client.get("statut") or main_status
        clients.append(client)

    return clients


def update_client_in_excel(row_number, client_data):
    """
    Update client data in Excel file

    Args:
        row_number (int): Row number to update
        client_data (dict): Updated client data dictionary

    Returns:
        bool: True if successful, False otherwise
    """
    if not OPENPYXL_AVAILABLE:
        logger.warning("openpyxl not available. Cannot update Excel.")
        return False

    # Create backup before modifying Excel file
    create_backup(CLIENT_EXCEL_PATH)

    if not os.path.exists(CLIENT_EXCEL_PATH):
        return False

    wb = load_workbook(CLIENT_EXCEL_PATH)
    if CLIENT_SHEET_NAME not in wb.sheetnames:
        return False

    ws = wb[CLIENT_SHEET_NAME]
    client_headers = [
        "Date",
        "Réf. Client",
        "Type Client",
        "Prénom",
        "Nom",
        "Date Arrivée",
        "Date Départ",
        "Durée Séjour",
        "Nombre Participants",
        "Nombre Adultes",
        "Enfants 2-12",
        "Bébés 0-2",
        "Téléphone",
        "Téléphone WhatsApp",
        "Email",
        "Période",
        "Restauration",
        "Hébergement",
        "Chambre",
        "Enfant",
        "Âge Enfant",
        "Forfait",
        "Circuit",
        "Statut",
        "SGL",
        "DBL",
        "TWN",
        "TPL",
        "FML",
    ]
    client_header_style = {
        "font": Font(bold=True, color="FFFFFF"),
        "fill": PatternFill(
            start_color="27AE60", end_color="27AE60", fill_type="solid"
        ),
        "alignment": Alignment(horizontal="center"),
    }
    header_map = _ensure_headers(ws, client_headers, client_header_style)

    # Update data
    field_map = {
        "Date": ["Timestamp", "Date", "date_jour"],
        "Réf. Client": ["Ref_Client", "ref_client"],
        "Type Client": ["Type_Client", "type_client"],
        "Prénom": ["Prénom", "prenom"],
        "Nom": ["Nom", "nom"],
        "Date Arrivée": ["Date_Arrivée", "date_arrivee"],
        "Date Départ": ["Date_Départ", "date_depart"],
        "Durée Séjour": ["Durée_Séjour", "duree_sejour"],
        "Nombre Participants": ["Nombre_Participants", "nombre_participants"],
        "Nombre Adultes": ["Nombre_Adultes", "nombre_adultes"],
        "Enfants 2-12": ["Enfants_2_12_ans", "nombre_enfants_2_12"],
        "Bébés 0-2": ["Bébés_0_2_ans", "nombre_bebes_0_2"],
        "Téléphone": ["Téléphone", "telephone"],
        "Téléphone WhatsApp": ["Téléphone_WhatsApp", "telephone_whatsapp"],
        "Email": ["Email", "email"],
        "Période": ["Période", "periode"],
        "Restauration": ["Restauration", "restauration"],
        "Hébergement": ["Hébergement", "hebergement"],
        "Chambre": ["Chambre", "chambre"],
        "Enfant": ["Enfant", "enfant"],
        "Âge Enfant": ["Âge_Enfant", "age_enfant"],
        "Forfait": ["Forfait", "forfait"],
        "Circuit": ["Circuit", "circuit"],
        "Statut": ["Statut", "statut"],
        "SGL": ["SGL_Count", "sgl_count"],
        "DBL": ["DBL_Count", "dbl_count"],
        "TWN": ["TWN_Count", "twn_count"],
        "TPL": ["TPL_Count", "tpl_count"],
        "FML": ["FML_Count", "fml_count"],
    }
    for header, keys in field_map.items():
        col = header_map.get(header)
        if col:
            ws.cell(
                row=row_number,
                column=col,
                value=_first_available(client_data, keys, ""),
            )

    wb.save(CLIENT_EXCEL_PATH)
    invalidate_client_cache()
    _save_client_infos_to_excel(client_data)
    return True


def update_client_statut(row_number, new_statut):
    """
    Update only the Statut cell of a client row.

    Args:
        row_number (int): Excel row number of the client
        new_statut (str): New statut value

    Returns:
        bool: True if successful
    """
    if not OPENPYXL_AVAILABLE:
        return False
    if not os.path.exists(CLIENT_EXCEL_PATH):
        return False

    wb = load_workbook(CLIENT_EXCEL_PATH)
    try:
        if CLIENT_SHEET_NAME not in wb.sheetnames:
            return False

        ws = wb[CLIENT_SHEET_NAME]
        header_map = _ensure_headers(ws, [])
        status_col = header_map.get("Statut")
        ref_col = header_map.get("Réf. Client")
        if not status_col:
            return False

        ws.cell(row=row_number, column=status_col, value=new_statut)

        ref_client = ""
        if ref_col:
            ref_value = ws.cell(row=row_number, column=ref_col).value
            ref_client = "" if ref_value is None else str(ref_value).strip()

        if CLIENT_INFOS_SHEET_NAME in wb.sheetnames and ref_client:
            info_ws = wb[CLIENT_INFOS_SHEET_NAME]
            info_header_map = _ensure_headers(
                info_ws,
                ["Réf. Client", "Statut"],
            )
            info_ref_col = info_header_map.get("Réf. Client")
            info_status_col = info_header_map.get("Statut")
            if info_ref_col and info_status_col:
                for info_row in range(2, info_ws.max_row + 1):
                    info_ref_value = info_ws.cell(row=info_row, column=info_ref_col).value
                    if info_ref_value is None:
                        continue
                    if str(info_ref_value).strip() == ref_client:
                        info_ws.cell(row=info_row, column=info_status_col, value=new_statut)
                        break

        wb.save(CLIENT_EXCEL_PATH)
        invalidate_client_cache()
        return True
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _load_client_infos_map():
    """Load extended client infos from INFOS_CLIENTS sheet by ref client"""
    if not OPENPYXL_AVAILABLE:
        return {}
    if not os.path.exists(CLIENT_EXCEL_PATH):
        return {}

    wb = load_workbook(CLIENT_EXCEL_PATH, data_only=True)
    if CLIENT_INFOS_SHEET_NAME not in wb.sheetnames:
        return {}

    ws = wb[CLIENT_INFOS_SHEET_NAME]
    header_map = _ensure_headers(ws, [])
    infos_map = {}

    def _cell(row, header, fallback_cell=None):
        col = header_map.get(header)
        if col:
            return ws.cell(row=row, column=col).value
        if fallback_cell:
            return ws[fallback_cell].value
        return None

    for row in range(2, ws.max_row + 1):
        if ws[f"A{row}"].value is None:
            continue
        ref_client = _cell(row, "Réf. Client", f"B{row}") or ""
        if not ref_client:
            continue
        infos_map[ref_client] = {
            "numero_dossier": _cell(row, "Numéro Dossier") or "",
            "statut": _cell(row, "Statut") or "",
            "heure_arrivee": _cell(row, "Heure Arrivée") or "",
            "heure_depart": _cell(row, "Heure Départ") or "",
            "compagnie": _cell(row, "Compagnie") or "",
            "aeroport": _cell(row, "Aéroport") or _cell(row, "Aeroport") or "",
            "ext_ref": _cell(row, "Réf. Externe") or _cell(row, "Ref. Externe") or "",
            "type_circuit": _cell(row, "Type Circuit") or "",
            "id_circuit": _cell(row, "ID Circuit") or "",
            "itineraire_circuit": _cell(row, "Itinéraire Circuit") or "",
            "activite_circuit": _cell(row, "Activité Circuit") or "",
            "duree_circuit": _cell(row, "Durée Circuit") or "",
            "condition_physique_circuit": _cell(row, "Condition Physique Circuit")
            or "",
            "type_voiture_circuit": _cell(row, "Type Voiture Circuit") or "",
            "hotels_defaut_villes_circuit": _cell(row, "Hôtels Défaut Circuit")
            or _cell(row, "Hotels Defaut Circuit")
            or "",
            "prestations_incluses_circuit": _cell(
                row, "Prestations Incluses Circuit"
            )
            or "",
            "transports_associes_circuit": _cell(row, "Transports Associés Circuit")
            or _cell(row, "Transports Associes Circuit")
            or "",
            "ville_depart": _cell(row, "Ville Départ") or "",
            "ville_arrivee": _cell(row, "Ville Arrivée") or "",
            "type_hotel_arrivee": _cell(row, "Type Hôtel Arrivée") or "",
        }

    return infos_map


def delete_client_from_excel(row_number):
    """
    Delete client data from Excel file

    Args:
        row_number (int): Row number to delete

    Returns:
        bool: True if successful, False otherwise
    """
    if not OPENPYXL_AVAILABLE:
        logger.warning("openpyxl not available. Cannot delete from Excel.")
        return False

    if not os.path.exists(CLIENT_EXCEL_PATH):
        return False

    wb = load_workbook(CLIENT_EXCEL_PATH)
    if CLIENT_SHEET_NAME not in wb.sheetnames:
        return False

    ws = wb[CLIENT_SHEET_NAME]

    # Delete row
    ws.delete_rows(row_number)

    wb.save(CLIENT_EXCEL_PATH)
    invalidate_client_cache()
    return True


@cached_hotel_data(ttl_seconds=86400)  # Cache for 24 hours
def load_all_hotels(client_type=None):
    """
    Load all hotel data from Excel file
    Results are cached for 24 hours

    Args:
        client_type (str): Filter by client type ('TO' or 'PBC'), if None load all

    Returns:
        list: List of hotel dictionaries
    """
    if not OPENPYXL_AVAILABLE:
        logger.warning("openpyxl not available. Cannot load from Excel.")
        return []

    if not os.path.exists(HOTEL_EXCEL_PATH):
        return []

    wb = load_workbook(HOTEL_EXCEL_PATH, data_only=True)
    if HOTEL_SHEET_NAME not in wb.sheetnames:
        return []

    ws = wb[HOTEL_SHEET_NAME]
    header_map_row1 = _get_header_map(ws, 1)
    header_map_row2 = _get_header_map(ws, 2)
    use_grouped_format = (
        "Ville" in header_map_row2
        and "HTL" in header_map_row2
        and "Ville" not in header_map_row1
    )

    hotels = []
    if use_grouped_format:
        grouped_columns = _iter_grouped_columns(ws, 1, 2)
        group_key_map = {
            "HOTEL": "hotel",
            "STANDARD": "standard",
            "BUNGALOWS": "bungalows",
            "DE LUXE": "deluxe",
            "SUITE": "suite",
            "OPTIONS": "options",
            "TAXE": "taxes",
            "REPAS": "meals",
            "AUTRES INFORMATIONS ET REMARQUES": "extras",
        }
        room_key_map = {
            "SPL": "single",
            "DBL": "double",
            "TWINS": "twin",
            "FML": "familiale",
            "TRIPLE": "triple",
            "CHAMBRE CHAUFFEUR": "chauffeur",
            "DORTOIR": "dortoir",
            "SUPP": "supp",
            "STUDIOS": "studios",
            "VIP": "vip",
        }
        options_key_map = {"CH. EVASION": "ch_evasion", "MASSAGE": "massage"}
        taxes_key_map = {
            "VIGNETTE": "vignette",
            "TAXE DE SEJOUR": "taxe_sejour",
            "TAXE DE SÉJOUR": "taxe_sejour",
        }
        meals_key_map = {
            "PDJ": "petit_dejeuner",
            "DJ": "dejeuner",
            "DR": "diner",
            "REPAS GUIDE": "repas_guide",
            "REPAS CHAUFFEUR": "repas_chauffeur",
        }
        extras_key_map = {"INCLUE": "inclue", "SPA": "spa", "REMARQUES": "remarques"}

        for row in range(3, ws.max_row + 1):
            if ws[f"A{row}"].value is None:
                continue

            hotel = {
                "row_number": row,
                "id": "",
                "nom": "",
                "lieu": "",
                "categorie": "",
                "type_client": "",
                "unite": "",
                "type_hebergement": "Hôtel",
                "room_rates": {
                    "standard": {},
                    "bungalows": {},
                    "deluxe": {},
                    "suite": {},
                },
                "options": {},
                "taxes": {},
                "meals": {},
                "extras": {},
                "contact": "",
                "email": "",
            }

            raw_category = ""

            for group, header, col in grouped_columns:
                group_key = group_key_map.get(str(group).strip().upper())
                if not group_key:
                    continue
                value = ws.cell(row=row, column=col).value
                if value is None or value == "":
                    continue

                header_norm = str(header).strip()
                header_norm_upper = header_norm.upper()

                if group_key == "hotel":
                    if header_norm_upper == "VILLE":
                        hotel["lieu"] = str(value).strip()
                    elif header_norm_upper == "HTL":
                        hotel["nom"] = str(value).strip()
                    elif (
                        header_norm_upper == "CATÉGORIE"
                        or header_norm_upper == "CATEGORIE"
                    ):
                        raw_category = str(value).strip()
                    elif header_norm_upper == "UNITÉ" or header_norm_upper == "UNITE":
                        hotel["unite"] = str(value).strip()
                elif group_key in ("standard", "bungalows", "deluxe", "suite"):
                    room_key = room_key_map.get(header_norm_upper)
                    if room_key:
                        hotel["room_rates"][group_key][room_key] = _parse_num(value)
                elif group_key == "options":
                    opt_key = options_key_map.get(header_norm_upper)
                    if opt_key:
                        hotel["options"][opt_key] = _parse_num(value)
                elif group_key == "taxes":
                    tax_key = taxes_key_map.get(header_norm_upper)
                    if tax_key:
                        hotel["taxes"][tax_key] = _parse_num(value)
                elif group_key == "meals":
                    meal_key = meals_key_map.get(header_norm_upper)
                    if meal_key:
                        hotel["meals"][meal_key] = _parse_num(value)
                elif group_key == "extras":
                    extra_key = extras_key_map.get(header_norm_upper)
                    if extra_key:
                        hotel["extras"][extra_key] = value

            raw_category_norm = raw_category.strip().upper()
            if raw_category_norm in ("TO", "PBC", "TCO", "PCB", "DU"):
                hotel["type_client"] = raw_category_norm
                hotel["categorie"] = raw_category_norm
            else:
                hotel["categorie"] = raw_category
                hotel["type_client"] = ""

            if not hotel["type_client"]:
                hotel["type_client"] = "TO"

            hotel["id"] = f"{hotel['lieu']}_{hotel['nom']}_{hotel['type_client']}"

            standard_rates = hotel["room_rates"].get("standard", {})
            hotel["chambre_single"] = standard_rates.get("single", 0)
            hotel["chambre_double"] = standard_rates.get(
                "double", 0
            ) or standard_rates.get("twin", 0)
            hotel["chambre_twin"] = standard_rates.get("twin", 0)
            hotel["chambre_familiale"] = standard_rates.get("familiale", 0)
            hotel["chambre_triple"] = standard_rates.get("triple", 0)
            hotel["chambre_chauffeur"] = standard_rates.get("chauffeur", 0)
            hotel["dortoir"] = standard_rates.get("dortoir", 0)
            hotel["lit_supp"] = standard_rates.get("supp", 0)
            villa_rates = hotel["room_rates"].get("villa", {})
            hotel["villa_single"] = villa_rates.get("single", 0)
            hotel["villa_double"] = villa_rates.get("double", 0)
            hotel["villa_twin"] = villa_rates.get("twin", 0)
            hotel["villa_familiale"] = villa_rates.get("familiale", 0)
            hotel["villa_triple"] = villa_rates.get("triple", 0)
            hotel["villa_studios"] = villa_rates.get("studios", 0)
            hotel["villa_vip"] = villa_rates.get("vip", 0)
            hotel["villa_supp"] = villa_rates.get("supp", 0)
            hotel["petit_dejeuner"] = hotel["meals"].get("petit_dejeuner", 0)
            hotel["dejeuner"] = hotel["meals"].get("dejeuner", 0)
            hotel["diner"] = hotel["meals"].get("diner", 0)
            hotel["inclus"] = str(hotel["extras"].get("inclue", "")).strip()

            if (
                client_type
                and hotel["type_client"]
                and hotel["type_client"] != client_type
            ):
                continue

            hotels.append(hotel)
    else:
        header_map = _ensure_headers(ws, [])
        # Start from row 2 (skip headers)
        for row in range(2, ws.max_row + 1):
            if ws[f"A{row}"].value is None:
                continue

            def _cell(header, fallback_cell=None):
                col = header_map.get(header)
                if col:
                    return ws.cell(row=row, column=col).value
                if fallback_cell:
                    return ws[fallback_cell].value
                return None

            hotel = {
                "row_number": row,
                "id": _cell("ID")
                or f"{_cell('Ville', f'A{row}')}_{_cell('HTL', f'B{row}')}",
                "nom": _cell("HTL", f"B{row}") or "",
                "lieu": _cell("Ville", f"A{row}") or "",
                "type_hebergement": _cell("TYPE_HEBERGEMENT") or "Hôtel",
                "categorie": _cell("CATÉGORIE", f"C{row}") or "",
                "type_client": _cell("TYPE_CLIENT", f"N{row}") or "TO",
                "chambre_single": _parse_num(_cell("SPL", f"E{row}")),
                "chambre_double": _parse_num(_cell("DBL", f"F{row}")),
                "chambre_familiale": _parse_num(_cell("FML", f"H{row}")),
                "lit_supp": _parse_num(_cell("SUPP", f"I{row}")),
                "day_use": (
                    _parse_num(_cell("DAY_USE")) if _cell("DAY_USE") is not None else 0
                ),
                "vignette": (
                    _parse_num(_cell("VIGNETTE"))
                    if _cell("VIGNETTE") is not None
                    else 0
                ),
                "taxe_sejour": (
                    _parse_num(_cell("TAXE_SEJOUR"))
                    if _cell("TAXE_SEJOUR") is not None
                    else 0
                ),
                "petit_dejeuner": _parse_num(_cell("PDJ", f"K{row}")),
                "dejeuner": _parse_num(_cell("DJ", f"L{row}")),
                "diner": _parse_num(_cell("DR", f"M{row}")),
                "description": _cell("DESCRIPTION")
                or f"Unité: {_cell('UNITÉ', f'D{row}') or ''}, Suite: {_cell('SUITE', f'J{row}') or ''}",
                "contact": _cell("CONTACT") or "",
                "email": _cell("EMAIL") or "",
                "room_rates": {
                    "standard": {
                        "single": _parse_num(_cell("SPL", f"E{row}")),
                        "double": _parse_num(_cell("DBL", f"F{row}")),
                        "twin": _parse_num(_cell("DBL", f"F{row}")),
                        "familiale": _parse_num(_cell("FML", f"H{row}")),
                        "supp": _parse_num(_cell("SUPP", f"I{row}")),
                    }
                },
                "meals": {
                    "petit_dejeuner": _parse_num(_cell("PDJ", f"K{row}")),
                    "dejeuner": _parse_num(_cell("DJ", f"L{row}")),
                    "diner": _parse_num(_cell("DR", f"M{row}")),
                },
                "options": {},
                "taxes": {},
                "extras": {},
            }

            # Filter by client type if specified
            if client_type and hotel["type_client"] != client_type:
                continue

            hotels.append(hotel)

    return hotels


def load_circuit_catalog():
    """
    Load circuit catalog from the Circuits sheet in data-hotel.xlsx.

    Returns:
        list[dict]: List of normalized circuit dictionaries
    """
    if not OPENPYXL_AVAILABLE:
        logger.warning("openpyxl not available. Cannot load circuits from Excel.")
        return []

    if not os.path.exists(HOTEL_EXCEL_PATH):
        return []

    wb = load_workbook(HOTEL_EXCEL_PATH, data_only=True)
    if "Circuits" not in wb.sheetnames:
        return []

    ws = wb["Circuits"]
    header_map = _get_header_map(ws, 1)

    id_col = _find_header_column(header_map, "ID circuit", "ID", "Id")
    name_col = _find_header_column(
        header_map, "Nom du circuit", "Nom", "Circuit", "Type Circuit"
    )
    itinerary_col = _find_header_column(
        header_map, "itinéraire", "itineraire", "Itinéraire"
    )
    cities_col = _find_header_column(
        header_map,
        "Villes parcourues",
        "Villes du circuit",
        "Villes",
    )
    activity_col = _find_header_column(header_map, "Activité", "Activite")
    duration_col = _find_header_column(header_map, "Durée", "Duree")
    fitness_col = _find_header_column(
        header_map, "condition physique", "Condition Physique"
    )
    vehicle_col = _find_header_column(header_map, "Type de voiture", "Voiture")
    default_hotels_col = _find_header_column(
        header_map,
        "Hôtels défaut par ville",
        "Hotels defaut par ville",
        "Hôtels par défaut par ville",
        "Hotels par defaut par ville",
        "Hôtels par ville",
        "Hotels par ville",
        "Hôtels défaut",
        "Hotels defaut",
        "Hôtels défaut circuit",
        "Hotels defaut circuit",
    )
    included_services_col = _find_header_column(
        header_map,
        "Prestations incluses",
        "Prestations incluses circuit",
        "Prestations circuit",
    )
    linked_transports_col = _find_header_column(
        header_map,
        "Transports associés",
        "Transports associes",
        "Transports associés circuit",
        "Transports associes circuit",
        "Transport associé",
        "Transport associe",
    )

    if not name_col:
        return []

    circuits = []
    seen_names = set()
    for row in range(2, ws.max_row + 1):
        raw_name = ws.cell(row=row, column=name_col).value
        if raw_name is None or str(raw_name).strip() == "":
            continue
        name = str(raw_name).strip()
        if name in seen_names:
            continue
        seen_names.add(name)

        itinerary = (
            str(ws.cell(row=row, column=itinerary_col).value).strip()
            if itinerary_col and ws.cell(row=row, column=itinerary_col).value is not None
            else ""
        )
        cities = (
            str(ws.cell(row=row, column=cities_col).value).strip()
            if cities_col and ws.cell(row=row, column=cities_col).value is not None
            else ""
        )
        if not itinerary and cities:
            itinerary = cities
        activity = (
            str(ws.cell(row=row, column=activity_col).value).strip()
            if activity_col and ws.cell(row=row, column=activity_col).value is not None
            else ""
        )
        duration = (
            str(ws.cell(row=row, column=duration_col).value).strip()
            if duration_col and ws.cell(row=row, column=duration_col).value is not None
            else ""
        )
        fitness = (
            str(ws.cell(row=row, column=fitness_col).value).strip()
            if fitness_col and ws.cell(row=row, column=fitness_col).value is not None
            else ""
        )
        vehicle = (
            str(ws.cell(row=row, column=vehicle_col).value).strip()
            if vehicle_col and ws.cell(row=row, column=vehicle_col).value is not None
            else ""
        )
        circuit_id = (
            str(ws.cell(row=row, column=id_col).value).strip()
            if id_col and ws.cell(row=row, column=id_col).value is not None
            else ""
        )
        default_hotels = (
            str(ws.cell(row=row, column=default_hotels_col).value).strip()
            if default_hotels_col
            and ws.cell(row=row, column=default_hotels_col).value is not None
            else ""
        )
        included_services = (
            str(ws.cell(row=row, column=included_services_col).value).strip()
            if included_services_col
            and ws.cell(row=row, column=included_services_col).value is not None
            else ""
        )
        linked_transports = (
            str(ws.cell(row=row, column=linked_transports_col).value).strip()
            if linked_transports_col
            and ws.cell(row=row, column=linked_transports_col).value is not None
            else ""
        )

        circuits.append(
            {
                "id_circuit": circuit_id,
                "nom": name,
                "itineraire": itinerary,
                "villes_parcourues": cities,
                "activite": activity,
                "duree": duration,
                "condition_physique": fitness,
                "type_voiture": vehicle,
                "hotels_defaut_villes": default_hotels,
                "prestations_incluses": included_services,
                "transports_associes": linked_transports,
            }
        )

    return circuits


def load_all_circuits():
    """
    Load circuit names from the Circuits sheet in data-hotel.xlsx.

    Returns:
        list: List of circuit names
    """
    return [circuit["nom"] for circuit in load_circuit_catalog() if circuit.get("nom")]


def get_circuit_db_headers():
    """Load header list from data-hotel.xlsx / Circuits."""
    if not OPENPYXL_AVAILABLE:
        return []

    if not os.path.exists(HOTEL_EXCEL_PATH):
        return []

    wb = None
    try:
        wb = load_workbook(HOTEL_EXCEL_PATH)
        if "Circuits" not in wb.sheetnames:
            return []

        ws = wb["Circuits"]
        default_headers = [
            "ID circuit",
            "Nom du circuit",
            "itinéraire",
            "Villes parcourues",
            "Activité",
            "Durée",
            "condition physique",
            "Type de voiture",
            "Hôtels défaut par ville",
            "Prestations incluses",
            "Transports associés",
        ]
        header_map = _ensure_headers(ws, default_headers)
        wb.save(HOTEL_EXCEL_PATH)
        return list(header_map.keys())
    except Exception as e:
        logger.error(f"Failed to load circuit DB headers: {e}", exc_info=True)
        return []
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def load_circuit_db_rows():
    """Load raw DB rows from data-hotel.xlsx / Circuits."""
    if not OPENPYXL_AVAILABLE:
        return []

    if not os.path.exists(HOTEL_EXCEL_PATH):
        return []

    wb = None
    try:
        wb = load_workbook(HOTEL_EXCEL_PATH)
        if "Circuits" not in wb.sheetnames:
            return []

        ws = wb["Circuits"]
        header_map = _get_header_map(ws, 1)
        if not header_map:
            return []

        headers = list(header_map.keys())
        rows = []
        for row_idx in range(2, ws.max_row + 1):
            row_dict = {"row_number": row_idx}
            has_values = False
            for header in headers:
                col = header_map[header]
                value = ws.cell(row=row_idx, column=col).value
                if value not in (None, ""):
                    has_values = True
                row_dict[header] = "" if value is None else value
            if has_values:
                rows.append(row_dict)
        return rows
    except Exception as e:
        logger.error(f"Failed to load circuit DB rows: {e}", exc_info=True)
        return []
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def save_circuit_db_row(row_data):
    """Save one row into data-hotel.xlsx / Circuits."""
    if not OPENPYXL_AVAILABLE:
        return -1

    wb = None
    try:
        if not os.path.exists(HOTEL_EXCEL_PATH):
            wb = Workbook()
            ws = wb.active
            ws.title = "Circuits"
        else:
            wb = load_workbook(HOTEL_EXCEL_PATH)
            if "Circuits" not in wb.sheetnames:
                ws = wb.create_sheet("Circuits")
            else:
                ws = wb["Circuits"]

        default_headers = [
            "ID circuit",
            "Nom du circuit",
            "itinéraire",
            "Villes parcourues",
            "Activité",
            "Durée",
            "condition physique",
            "Type de voiture",
            "Hôtels défaut par ville",
            "Prestations incluses",
            "Transports associés",
        ]
        header_map = _ensure_headers(ws, default_headers)
        if row_data:
            header_map = _ensure_headers(ws, list(row_data.keys()))
        if not header_map:
            return -1

        next_row = 2
        while True:
            has_data = False
            for header, col in header_map.items():
                if ws.cell(row=next_row, column=col).value not in (None, ""):
                    has_data = True
                    break
            if not has_data:
                break
            next_row += 1

        for header, col in header_map.items():
            ws.cell(row=next_row, column=col, value=row_data.get(header, ""))

        wb.save(HOTEL_EXCEL_PATH)
        return next_row
    except PermissionError:
        return -2
    except Exception as e:
        logger.error(f"Failed to save circuit DB row: {e}", exc_info=True)
        return -1
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def update_circuit_db_row(row_number, row_data):
    """Update one row in data-hotel.xlsx / Circuits."""
    if not OPENPYXL_AVAILABLE:
        return -1

    if not os.path.exists(HOTEL_EXCEL_PATH):
        return -1

    wb = None
    try:
        wb = load_workbook(HOTEL_EXCEL_PATH)
        if "Circuits" not in wb.sheetnames:
            return -1

        ws = wb["Circuits"]
        header_map = _get_header_map(ws, 1)
        if not header_map:
            return -1

        for header, col in header_map.items():
            ws.cell(row=row_number, column=col, value=row_data.get(header, ""))

        wb.save(HOTEL_EXCEL_PATH)
        return 0
    except PermissionError:
        return -2
    except Exception as e:
        logger.error(f"Failed to update circuit DB row {row_number}: {e}", exc_info=True)
        return -1
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def delete_circuit_db_row(row_number):
    """Delete one row from data-hotel.xlsx / Circuits."""
    if not OPENPYXL_AVAILABLE:
        return False

    if not os.path.exists(HOTEL_EXCEL_PATH):
        return False

    wb = None
    try:
        wb = load_workbook(HOTEL_EXCEL_PATH)
        if "Circuits" not in wb.sheetnames:
            return False

        ws = wb["Circuits"]
        ws.delete_rows(row_number)
        wb.save(HOTEL_EXCEL_PATH)
        return True
    except Exception as e:
        logger.error(f"Failed to delete circuit DB row {row_number}: {e}", exc_info=True)
        return False
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def save_hotel_to_excel(hotel_data):
    """
    Save hotel data to Excel file

    Args:
        hotel_data (dict): Hotel data dictionary

    Returns:
        int: Row number where data was saved, or -1 if failed
    """
    if not OPENPYXL_AVAILABLE:
        logger.warning("openpyxl not available. Cannot save to Excel.")
        return -1

    # Create backup before modifying Excel file
    create_backup(HOTEL_EXCEL_PATH)
    # Legacy flat headers, kept for backward compatibility with old sheets.
    hotel_headers = [
        "Ville",
        "HTL",
        "CATÉGORIE",
        "UNITÉ",
        "SPL",
        "DBL",
        "TWINS",
        "FML",
        "SUPP",
        "SUITE",
        "PDJ",
        "DJ",
        "DR",
        "ID",
        "TYPE_HEBERGEMENT",
        "TYPE_CLIENT",
        "CONTACT",
        "EMAIL",
        "DESCRIPTION",
        "DAY_USE",
        "VIGNETTE",
        "TAXE_SEJOUR",
    ]
    hotel_header_style = {
        "font": Font(bold=True, color="FFFFFF"),
        "fill": PatternFill(
            start_color="27AE60", end_color="27AE60", fill_type="solid"
        ),
        "alignment": Alignment(horizontal="center"),
    }

    if not os.path.exists(HOTEL_EXCEL_PATH):
        wb = Workbook()
        wb.save(HOTEL_EXCEL_PATH)

    # Open existing file
    wb = load_workbook(HOTEL_EXCEL_PATH)

    if HOTEL_SHEET_NAME not in wb.sheetnames:
        ws = wb.create_sheet(HOTEL_SHEET_NAME)
        _ensure_headers(ws, hotel_headers, hotel_header_style)
        wb.save(HOTEL_EXCEL_PATH)

    # Open existing file again
    wb = load_workbook(HOTEL_EXCEL_PATH)
    ws = wb[HOTEL_SHEET_NAME]

    header_map_row1 = _get_header_map(ws, 1)
    header_map_row2 = _get_header_map(ws, 2)
    use_grouped_format = (
        "Ville" in header_map_row2 and "HTL" in header_map_row2 and "Ville" not in header_map_row1
    )

    last_row = ws.max_row + 1
    if use_grouped_format:
        last_row = max(3, last_row)
        grouped_map = {
            (str(group).strip(), str(header).strip()): col
            for group, header, col in _iter_grouped_columns(ws, 1, 2)
        }

        def _set(group, header, keys, default=""):
            col = grouped_map.get((group, header))
            if not col:
                return
            value = _first_available(hotel_data, keys, default)
            ws.cell(row=last_row, column=col, value=value)

        # HOTEL block
        _set("Hotel", "Ville", ["Lieu", "lieu"])
        _set("Hotel", "HTL", ["Nom", "nom"])
        _set("Hotel", "NB ETOILE", ["Type_Client", "type_client"], "")
        _set("Hotel", "CATÉGORIE", ["Catégorie", "categorie"], "")
        _set("Hotel", "UNITÉ", ["Unité", "unite"], "MGA")

        # STANDARD block
        _set("STANDARD", "SPL", ["Chambre_Single", "chambre_single"])
        _set("STANDARD", "DBL", ["Chambre_Double", "chambre_double"])
        _set("STANDARD", "TWINS", ["Chambre_Twin", "chambre_twin", "Chambre_Double", "chambre_double"])
        _set("STANDARD", "FML", ["Chambre_Familiale", "chambre_familiale"])
        _set("STANDARD", "triple", ["Chambre_Triple", "chambre_triple"])
        _set("STANDARD", "Chambre chauffeur", ["Chambre_Chauffeur", "chambre_chauffeur"])
        _set("STANDARD", "dortoir", ["Dortoir", "dortoir"])
        _set("STANDARD", "SUPP", ["Lit_Supp", "lit_supp"])

        # BUNGALOWS
        _set("BUNGALOWS", "SPL", ["Bungalow_Single", "bungalow_single"])
        _set("BUNGALOWS", "DBL", ["Bungalow_Double", "bungalow_double"])
        _set("BUNGALOWS", "TWINS", ["Bungalow_Twin", "bungalow_twin"])
        _set("BUNGALOWS", "FML", ["Bungalow_Familiale", "bungalow_familiale"])
        _set("BUNGALOWS", "Triple", ["Bungalow_Triple", "bungalow_triple"])
        _set("BUNGALOWS", "SUPP", ["Bungalow_Supp", "bungalow_supp"])

        # DE LUXE
        _set("DE LUXE", "SPL", ["Deluxe_Single", "deluxe_single"])
        _set("DE LUXE", "DBL", ["Deluxe_Double", "deluxe_double"])
        _set("DE LUXE", "TWINS", ["Deluxe_Twin", "deluxe_twin"])
        _set("DE LUXE", "FML", ["Deluxe_Familiale", "deluxe_familiale"])
        _set("DE LUXE", "triple", ["Deluxe_Triple", "deluxe_triple"])
        _set("DE LUXE", "SUPP", ["Deluxe_Supp", "deluxe_supp"])

        # SUITE
        _set("SUITE", "SPL", ["Suite_Single", "suite_single"])
        _set("SUITE", "DBL", ["Suite_Double", "suite_double"])
        _set("SUITE", "TWINS", ["Suite_Twin", "suite_twin"])
        _set("SUITE", "FML", ["Suite_Familiale", "suite_familiale"])
        _set("SUITE", "triple", ["Suite_Triple", "suite_triple"])
        _set("SUITE", "studios", ["Suite_Studios", "suite_studios"])
        _set("SUITE", "VIP", ["Suite_VIP", "suite_vip"])
        _set("SUITE", "SUPP", ["Suite_Supp", "suite_supp"])

        # VILLA
        _set("Villa", "SPL", ["Villa_Single", "villa_single"])
        _set("Villa", "DBL", ["Villa_Double", "villa_double"])
        _set("Villa", "TWINS", ["Villa_Twin", "villa_twin"])
        _set("Villa", "FML", ["Villa_Familiale", "villa_familiale"])
        _set("Villa", "triple", ["Villa_Triple", "villa_triple"])
        _set("Villa", "studios", ["Villa_Studios", "villa_studios"])
        _set("Villa", "VIP", ["Villa_VIP", "villa_vip"])
        _set("Villa", "SUPP", ["Villa_Supp", "villa_supp"])

        # TAXE / REPAS / AUTRES
        _set("TAXE", "Vignette", ["Vignette", "vignette"])
        _set("TAXE", "Taxe de séjour", ["Taxe_Séjour", "taxe_sejour"])
        _set("REPAS", "PDJ", ["Petit_Déjeuner", "petit_dejeuner"])
        _set("REPAS", "DJ", ["Déjeuner", "dejeuner"])
        _set("REPAS", "DR", ["Dîner", "diner"])
        _set("Autres informations et remarques", "Inclus", ["Inclus", "inclus", "Description", "description"])
        _set("Autres informations et remarques", "Inclus ", ["Inclus", "inclus", "Description", "description"])
    else:
        header_map = _ensure_headers(ws, hotel_headers, hotel_header_style)
        field_map = {
            "Ville": ["Lieu", "lieu"],
            "HTL": ["Nom", "nom"],
            "CATÉGORIE": ["Catégorie", "categorie"],
            "UNITÉ": ["Unité", "unite"],
            "SPL": ["Chambre_Single", "chambre_single"],
            "DBL": ["Chambre_Double", "chambre_double"],
            "TWINS": ["Chambre_Double", "chambre_double"],
            "FML": ["Chambre_Familiale", "chambre_familiale"],
            "SUPP": ["Lit_Supp", "lit_supp"],
            "SUITE": ["Suite", "suite"],
            "PDJ": ["Petit_Déjeuner", "petit_dejeuner"],
            "DJ": ["Déjeuner", "dejeuner"],
            "DR": ["Dîner", "diner"],
            "ID": ["ID", "id"],
            "TYPE_HEBERGEMENT": ["Type_Hébergement", "type_hebergement"],
            "TYPE_CLIENT": ["Type_Client", "type_client"],
            "CONTACT": ["Contact", "contact"],
            "EMAIL": ["Email", "email"],
            "DESCRIPTION": ["Description", "description"],
            "DAY_USE": ["Day_Use", "day_use"],
            "VIGNETTE": ["Vignette", "vignette"],
            "TAXE_SEJOUR": ["Taxe_Séjour", "taxe_sejour"],
        }
        for header, keys in field_map.items():
            col = header_map.get(header)
            if col:
                value = _first_available(hotel_data, keys, "")
                if header == "UNITÉ" and value == "":
                    value = "MGA"
                ws.cell(row=last_row, column=col, value=value)

    # Auto-adjust column widths
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for row_idx in range(1, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            value_len = len(str(value))
            if value_len > max_length:
                max_length = value_len
        ws.column_dimensions[column_letter].width = min(max_length + 2, 25)

    wb.save(HOTEL_EXCEL_PATH)
    invalidate_hotel_cache()
    return last_row


def update_hotel_in_excel(row_number, hotel_data):
    """
    Update hotel data in Excel file

    Args:
        row_number (int): Row number to update
        hotel_data (dict): Updated hotel data dictionary

    Returns:
        bool: True if successful, False otherwise
    """
    if not OPENPYXL_AVAILABLE:
        logger.warning("openpyxl not available. Cannot update Excel.")
        return False

    # Create backup before modifying Excel file
    create_backup(HOTEL_EXCEL_PATH)

    if not os.path.exists(HOTEL_EXCEL_PATH):
        return False

    wb = load_workbook(HOTEL_EXCEL_PATH)
    if HOTEL_SHEET_NAME not in wb.sheetnames:
        return False

    ws = wb[HOTEL_SHEET_NAME]
    header_map_row1 = _get_header_map(ws, 1)
    header_map_row2 = _get_header_map(ws, 2)
    use_grouped_format = (
        "Ville" in header_map_row2 and "HTL" in header_map_row2 and "Ville" not in header_map_row1
    )
    hotel_headers = [
        "Ville",
        "HTL",
        "CATÉGORIE",
        "UNITÉ",
        "SPL",
        "DBL",
        "TWINS",
        "FML",
        "SUPP",
        "SUITE",
        "PDJ",
        "DJ",
        "DR",
        "ID",
        "TYPE_HEBERGEMENT",
        "TYPE_CLIENT",
        "CONTACT",
        "EMAIL",
        "DESCRIPTION",
        "DAY_USE",
        "VIGNETTE",
        "TAXE_SEJOUR",
    ]
    hotel_header_style = {
        "font": Font(bold=True, color="FFFFFF"),
        "fill": PatternFill(
            start_color="27AE60", end_color="27AE60", fill_type="solid"
        ),
        "alignment": Alignment(horizontal="center"),
    }
    if use_grouped_format:
        grouped_map = {
            (str(group).strip(), str(header).strip()): col
            for group, header, col in _iter_grouped_columns(ws, 1, 2)
        }

        def _set(group, header, keys, default=""):
            col = grouped_map.get((group, header))
            if not col:
                return
            value = _first_available(hotel_data, keys, default)
            ws.cell(row=row_number, column=col, value=value)

        _set("Hotel", "Ville", ["Lieu", "lieu"])
        _set("Hotel", "HTL", ["Nom", "nom"])
        _set("Hotel", "NB ETOILE", ["Type_Client", "type_client"], "")
        _set("Hotel", "CATÉGORIE", ["Catégorie", "categorie"], "")
        _set("Hotel", "UNITÉ", ["Unité", "unite"], "MGA")
        _set("STANDARD", "SPL", ["Chambre_Single", "chambre_single"])
        _set("STANDARD", "DBL", ["Chambre_Double", "chambre_double"])
        _set("STANDARD", "TWINS", ["Chambre_Twin", "chambre_twin", "Chambre_Double", "chambre_double"])
        _set("STANDARD", "FML", ["Chambre_Familiale", "chambre_familiale"])
        _set("STANDARD", "triple", ["Chambre_Triple", "chambre_triple"])
        _set("STANDARD", "Chambre chauffeur", ["Chambre_Chauffeur", "chambre_chauffeur"])
        _set("STANDARD", "dortoir", ["Dortoir", "dortoir"])
        _set("STANDARD", "SUPP", ["Lit_Supp", "lit_supp"])
        _set("BUNGALOWS", "SPL", ["Bungalow_Single", "bungalow_single"])
        _set("BUNGALOWS", "DBL", ["Bungalow_Double", "bungalow_double"])
        _set("BUNGALOWS", "TWINS", ["Bungalow_Twin", "bungalow_twin"])
        _set("BUNGALOWS", "FML", ["Bungalow_Familiale", "bungalow_familiale"])
        _set("BUNGALOWS", "Triple", ["Bungalow_Triple", "bungalow_triple"])
        _set("BUNGALOWS", "SUPP", ["Bungalow_Supp", "bungalow_supp"])
        _set("DE LUXE", "SPL", ["Deluxe_Single", "deluxe_single"])
        _set("DE LUXE", "DBL", ["Deluxe_Double", "deluxe_double"])
        _set("DE LUXE", "TWINS", ["Deluxe_Twin", "deluxe_twin"])
        _set("DE LUXE", "FML", ["Deluxe_Familiale", "deluxe_familiale"])
        _set("DE LUXE", "triple", ["Deluxe_Triple", "deluxe_triple"])
        _set("DE LUXE", "SUPP", ["Deluxe_Supp", "deluxe_supp"])
        _set("SUITE", "SPL", ["Suite_Single", "suite_single"])
        _set("SUITE", "DBL", ["Suite_Double", "suite_double"])
        _set("SUITE", "TWINS", ["Suite_Twin", "suite_twin"])
        _set("SUITE", "FML", ["Suite_Familiale", "suite_familiale"])
        _set("SUITE", "triple", ["Suite_Triple", "suite_triple"])
        _set("SUITE", "studios", ["Suite_Studios", "suite_studios"])
        _set("SUITE", "VIP", ["Suite_VIP", "suite_vip"])
        _set("SUITE", "SUPP", ["Suite_Supp", "suite_supp"])
        _set("Villa", "SPL", ["Villa_Single", "villa_single"])
        _set("Villa", "DBL", ["Villa_Double", "villa_double"])
        _set("Villa", "TWINS", ["Villa_Twin", "villa_twin"])
        _set("Villa", "FML", ["Villa_Familiale", "villa_familiale"])
        _set("Villa", "triple", ["Villa_Triple", "villa_triple"])
        _set("Villa", "studios", ["Villa_Studios", "villa_studios"])
        _set("Villa", "VIP", ["Villa_VIP", "villa_vip"])
        _set("Villa", "SUPP", ["Villa_Supp", "villa_supp"])
        _set("TAXE", "Vignette", ["Vignette", "vignette"])
        _set("TAXE", "Taxe de séjour", ["Taxe_Séjour", "taxe_sejour"])
        _set("REPAS", "PDJ", ["Petit_Déjeuner", "petit_dejeuner"])
        _set("REPAS", "DJ", ["Déjeuner", "dejeuner"])
        _set("REPAS", "DR", ["Dîner", "diner"])
        _set("Autres informations et remarques", "Inclus", ["Inclus", "inclus", "Description", "description"])
        _set("Autres informations et remarques", "Inclus ", ["Inclus", "inclus", "Description", "description"])
    else:
        header_map = _ensure_headers(ws, hotel_headers, hotel_header_style)
        field_map = {
            "Ville": ["Lieu", "lieu"],
            "HTL": ["Nom", "nom"],
            "CATÉGORIE": ["Catégorie", "categorie"],
            "UNITÉ": ["Unité", "unite"],
            "SPL": ["Chambre_Single", "chambre_single"],
            "DBL": ["Chambre_Double", "chambre_double"],
            "TWINS": ["Chambre_Double", "chambre_double"],
            "FML": ["Chambre_Familiale", "chambre_familiale"],
            "SUPP": ["Lit_Supp", "lit_supp"],
            "SUITE": ["Suite", "suite"],
            "PDJ": ["Petit_Déjeuner", "petit_dejeuner"],
            "DJ": ["Déjeuner", "dejeuner"],
            "DR": ["Dîner", "diner"],
            "ID": ["ID", "id"],
            "TYPE_HEBERGEMENT": ["Type_Hébergement", "type_hebergement"],
            "TYPE_CLIENT": ["Type_Client", "type_client"],
            "CONTACT": ["Contact", "contact"],
            "EMAIL": ["Email", "email"],
            "DESCRIPTION": ["Description", "description"],
            "DAY_USE": ["Day_Use", "day_use"],
            "VIGNETTE": ["Vignette", "vignette"],
            "TAXE_SEJOUR": ["Taxe_Séjour", "taxe_sejour"],
        }
        for header, keys in field_map.items():
            col = header_map.get(header)
            if col:
                value = _first_available(hotel_data, keys, "")
                if header == "UNITÉ" and value == "":
                    value = "MGA"
                ws.cell(row=row_number, column=col, value=value)

    wb.save(HOTEL_EXCEL_PATH)
    invalidate_hotel_cache()
    return True


def delete_hotel_from_excel(row_number):
    """
    Delete hotel data from Excel file

    Args:
        row_number (int): Row number to delete

    Returns:
        bool: True if successful, False otherwise
    """
    if not OPENPYXL_AVAILABLE:
        logger.warning("openpyxl not available. Cannot delete from Excel.")
        return False

    if not os.path.exists(HOTEL_EXCEL_PATH):
        return False

    wb = load_workbook(HOTEL_EXCEL_PATH)
    if HOTEL_SHEET_NAME not in wb.sheetnames:
        return False

    ws = wb[HOTEL_SHEET_NAME]

    # Delete row
    ws.delete_rows(row_number)

    wb.save(HOTEL_EXCEL_PATH)
    invalidate_hotel_cache()
    return True


def get_collective_expense_headers():
    """
    Load header list from COTATION_FRAIS_COL sheet.

    Returns:
        list: Header names found in row 1.
    """
    if not OPENPYXL_AVAILABLE:
        logger.warning("openpyxl not available. Cannot load collective headers.")
        return []

    if not os.path.exists(CLIENT_EXCEL_PATH):
        return []

    wb = None
    try:
        wb = load_workbook(CLIENT_EXCEL_PATH)
        if COTATION_FRAIS_COL_SHEET_NAME not in wb.sheetnames:
            return []

        ws = wb[COTATION_FRAIS_COL_SHEET_NAME]
        headers = []
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=1, column=col).value
            if value is None:
                continue
            label = str(value).strip()
            if label:
                headers.append(label)
        return headers
    except Exception as e:
        logger.error(f"Failed to load collective expense headers: {e}", exc_info=True)
        return []
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def save_collective_expense_quotation_to_excel(form_data):
    """
    Save a collective expense quotation row into COTATION_FRAIS_COL sheet.

    Args:
        form_data (dict): Key/value mapping where keys are sheet header labels.

    Returns:
        int: Saved row number, or -1 on failure.
    """
    if not OPENPYXL_AVAILABLE:
        logger.warning(
            "openpyxl not available. Cannot save collective expense quotation to Excel."
        )
        return -1

    wb = None
    try:
        if not os.path.exists(CLIENT_EXCEL_PATH):
            wb = Workbook()
            ws = wb.active
            ws.title = COTATION_FRAIS_COL_SHEET_NAME
        else:
            wb = load_workbook(CLIENT_EXCEL_PATH)
            if COTATION_FRAIS_COL_SHEET_NAME not in wb.sheetnames:
                ws = wb.create_sheet(COTATION_FRAIS_COL_SHEET_NAME)
            else:
                ws = wb[COTATION_FRAIS_COL_SHEET_NAME]

        headers = []
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=1, column=col).value
            if value is None:
                continue
            label = str(value).strip()
            if label:
                headers.append(label)

        if not headers:
            headers = list(form_data.keys())
            if "Date" not in headers:
                headers.insert(0, "Date")
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    start_color="4472C4", end_color="4472C4", fill_type="solid"
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")

        header_map = {header: index for index, header in enumerate(headers, start=1)}

        next_row = ws.max_row + 1

        if "Date" in header_map and not form_data.get("Date"):
            form_data = {
                "Date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                **form_data,
            }

        for header, col in header_map.items():
            value = form_data.get(header, "")
            ws.cell(row=next_row, column=col, value=value)

        for col in range(1, len(headers) + 1):
            letter = ws.cell(row=1, column=col).column_letter
            ws.column_dimensions[letter].width = max(
                14,
                min(40, len(str(ws.cell(row=1, column=col).value or "")) + 4),
            )

        wb.save(CLIENT_EXCEL_PATH)
        logger.info(
            f"Collective expense quotation saved to row {next_row} in {COTATION_FRAIS_COL_SHEET_NAME}"
        )
        return next_row
    except PermissionError as e:
        logger.error(
            f"Excel file is locked. Close data.xlsx and retry: {e}", exc_info=True
        )
        return -2
    except Exception as e:
        logger.error(
            f"Failed to save collective expense quotation to Excel: {e}", exc_info=True
        )
        return -1
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def load_all_collective_expense_quotations():
    """
    Load all collective expense quotations from COTATION_FRAIS_COL.

    Returns:
        list: List of row dictionaries (keys = sheet headers).
    """
    if not OPENPYXL_AVAILABLE:
        logger.warning(
            "openpyxl not available. Cannot load collective expense quotations."
        )
        return []

    if not os.path.exists(CLIENT_EXCEL_PATH):
        return []

    wb = None
    try:
        wb = load_workbook(CLIENT_EXCEL_PATH)
        if COTATION_FRAIS_COL_SHEET_NAME not in wb.sheetnames:
            return []

        ws = wb[COTATION_FRAIS_COL_SHEET_NAME]
        headers = get_collective_expense_headers()
        if not headers:
            return []

        rows = []
        for row_index in range(2, ws.max_row + 1):
            row_dict = {"row_number": row_index}
            has_values = False

            for col_index, header in enumerate(headers, start=1):
                value = ws.cell(row=row_index, column=col_index).value
                if value not in (None, ""):
                    has_values = True
                row_dict[header] = "" if value is None else value

            if has_values:
                rows.append(row_dict)

        return rows
    except Exception as e:
        logger.error(
            f"Failed to load collective expense quotations: {e}", exc_info=True
        )
        return []
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def save_hotel_quotation_to_excel(quotation_data):
    """
    Save hotel quotation to COTATION_H sheet in data.xlsx

    Args:
        quotation_data (dict): Quotation data with keys:
            - client_id (str): ID/Ref of the client
            - client_name (str): Name of the client
            - hotel_name (str): Name of the hotel
            - city (str): City where the hotel is located
            - total_price (float): Total price of the quotation
            - currency (str): Currency of the price
            - nights (int): Number of nights
            - adults (int): Number of adults
            - children (int): Number of children
            - room_type (str): Type of room
            - meal_plan (str): Meal plan selected
            - period (str): Period/season
            - quote_date (str): Date of the quotation

    Returns:
        int: Row number where data was saved, or -1 if failed
    """
    if not OPENPYXL_AVAILABLE:
        logger.warning("openpyxl not available. Cannot save quotation to Excel.")
        return -1

    try:
        # Create or load the file
        if not os.path.exists(CLIENT_EXCEL_PATH):
            wb = Workbook()
            ws = wb.active
            ws.title = COTATION_H_SHEET_NAME
        else:
            wb = load_workbook(CLIENT_EXCEL_PATH)
            if COTATION_H_SHEET_NAME not in wb.sheetnames:
                ws = wb.create_sheet(COTATION_H_SHEET_NAME)
            else:
                ws = wb[COTATION_H_SHEET_NAME]

        headers = [
            "Date",
            "ID_Client",
            "Numero_Dossier",
            "Nom_Client",
            "Prénom_Client",
            "Hôtel",
            "Ville",
            "Nuits",
            "Type_Chambre",
            "Adultes",
            "Enfants",
            "Plan_Repas",
            "Période",
            "Total_Devise",
            "Devise",
        ]

        # Add headers if this is the first row
        if ws.max_row == 1 or ws[f"A1"].value is None:
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="4472C4", end_color="4472C4", fill_type="solid"
                )
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Find next row
        next_row = ws.max_row + 1

        # Ensure headers for existing files where columns could be missing
        header_map = _ensure_headers(ws, headers)

        row_values = {
            "Date": quotation_data.get(
                "quote_date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ),
            "ID_Client": quotation_data.get("client_id", ""),
            "Numero_Dossier": quotation_data.get("numero_dossier", ""),
            "Nom_Client": quotation_data.get("client_name", ""),
            "Prénom_Client": quotation_data.get("client_first_name", ""),
            "Hôtel": quotation_data.get("hotel_name", ""),
            "Ville": quotation_data.get("city", ""),
            "Nuits": quotation_data.get("nights", 0),
            "Type_Chambre": quotation_data.get("room_type", ""),
            "Adultes": quotation_data.get("adults", 0),
            "Enfants": quotation_data.get("children", 0),
            "Plan_Repas": quotation_data.get("meal_plan", ""),
            "Période": quotation_data.get("period", ""),
            "Total_Devise": _parse_num(quotation_data.get("total_price", 0)),
            "Devise": quotation_data.get("currency", "Ariary"),
        }

        for header, value in row_values.items():
            col_idx = header_map.get(header)
            if col_idx:
                ws.cell(row=next_row, column=col_idx, value=value)

        # Adjust column widths
        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions["E"].width = 25
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 8
        ws.column_dimensions["H"].width = 14
        ws.column_dimensions["I"].width = 8
        ws.column_dimensions["J"].width = 8
        ws.column_dimensions["K"].width = 18
        ws.column_dimensions["L"].width = 12
        ws.column_dimensions["M"].width = 14
        ws.column_dimensions["N"].width = 10

        wb.save(CLIENT_EXCEL_PATH)
        logger.info(f"Quotation saved to row {next_row} in {COTATION_H_SHEET_NAME}")
        return next_row

    except Exception as e:
        logger.error(f"Failed to save quotation to Excel: {e}", exc_info=True)
        return -1


def load_all_hotel_quotations():
    """
    Load all hotel quotations from COTATION_H sheet

    Returns:
        list: List of quotation dictionaries
    """
    if not OPENPYXL_AVAILABLE:
        logger.warning("openpyxl not available. Cannot load from Excel.")
        return []

    if not os.path.exists(CLIENT_EXCEL_PATH):
        return []

    try:
        wb = load_workbook(CLIENT_EXCEL_PATH)
        if COTATION_H_SHEET_NAME not in wb.sheetnames:
            return []

        ws = wb[COTATION_H_SHEET_NAME]
        header_map = _get_header_map(ws, 1)

        def _cell(row_idx, *aliases, default=""):
            for alias in aliases:
                col = header_map.get(alias)
                if col:
                    value = ws.cell(row=row_idx, column=col).value
                    if value not in (None, ""):
                        return value
            return default

        quotations = []
        # Start from row 2 (skip headers)
        for row in range(2, ws.max_row + 1):
            if ws[f"A{row}"].value is None:
                continue

            quotation = {
                "row_number": row,
                "quote_date": _cell(row, "Date", default=ws[f"A{row}"].value or ""),
                "client_id": _cell(row, "ID_Client", default=ws[f"B{row}"].value or ""),
                "client_name": _cell(
                    row, "Nom_Client", default=ws[f"C{row}"].value or ""
                ),
                "hotel_name": _cell(
                    row, "Hôtel", "Hotel", default=ws[f"D{row}"].value or ""
                ),
                "city": _cell(row, "Ville", default=ws[f"E{row}"].value or ""),
                "nights": _parse_num(
                    _cell(row, "Nuits", default=ws[f"F{row}"].value or 0)
                ),
                "room_type": _cell(
                    row, "Type_Chambre", default=ws[f"G{row}"].value or ""
                ),
                "adults": _parse_num(
                    _cell(row, "Adultes", default=ws[f"H{row}"].value or 0)
                ),
                "children": _parse_num(
                    _cell(row, "Enfants", default=ws[f"I{row}"].value or 0)
                ),
                "meal_plan": _cell(
                    row, "Plan_Repas", default=ws[f"J{row}"].value or ""
                ),
                "period": _cell(row, "Période", default=ws[f"K{row}"].value or ""),
                "total_price": _parse_num(
                    _cell(row, "Total_Devise", default=ws[f"L{row}"].value or 0)
                ),
                "currency": _cell(
                    row, "Devise", default=ws[f"M{row}"].value or "Ariary"
                ),
            }
            quotations.append(quotation)

        return quotations

    except Exception as e:
        logger.error(f"Failed to load quotations from Excel: {e}", exc_info=True)
        return []


def get_quotations_grouped_by_client():
    """
    Get all hotel quotations grouped by client with subtotals

    Returns:
        dict: Dictionary with client_id as key and list of quotations as value,
              plus 'total' key with grand total per client
    """
    quotations = load_all_hotel_quotations()
    grouped = {}

    for quotation in quotations:
        client_id = quotation["client_id"]
        if client_id not in grouped:
            grouped[client_id] = {
                "client_name": quotation["client_name"],
                "quotations": [],
                "total": 0,
                "currency": quotation["currency"],
            }

        grouped[client_id]["quotations"].append(quotation)
        grouped[client_id]["total"] += quotation["total_price"]

    return grouped


def get_quotations_by_city():
    """
    Get all hotel quotations grouped by city with subtotals

    Returns:
        dict: Dictionary with city as key and list of quotations as value,
              plus 'total' key with grand total per city
    """
    quotations = load_all_hotel_quotations()
    grouped = {}

    for quotation in quotations:
        city = quotation["city"]
        if city not in grouped:
            grouped[city] = {
                "quotations": [],
                "total": 0,
                "currency": quotation["currency"],
            }

        grouped[city]["quotations"].append(quotation)
        grouped[city]["total"] += quotation["total_price"]

    return grouped


def load_collective_expenses_data():
    """
    Load all data from Frais collectifs sheet in data-hotel.xlsx
    
    Returns:
        list: List of dicts with FORFAIT, PRESTATAIRES, DESIGNATION, MONTANT, ID circuit
    """
    if not OPENPYXL_AVAILABLE:
        logger.warning("openpyxl not available. Cannot load collective expenses.")
        return []

    if not os.path.exists(HOTEL_EXCEL_PATH):
        return []

    wb = None
    try:
        wb = load_workbook(HOTEL_EXCEL_PATH)
        if FRAIS_COLLECTIFS_SHEET_NAME not in wb.sheetnames:
            return []

        ws = wb[FRAIS_COLLECTIFS_SHEET_NAME]
        rows = []
        
        for row_idx in range(2, ws.max_row + 1):
            forfait = ws.cell(row=row_idx, column=1).value
            prestataire = ws.cell(row=row_idx, column=2).value
            designation = ws.cell(row=row_idx, column=3).value
            montant = ws.cell(row=row_idx, column=4).value
            id_circuit = ws.cell(row=row_idx, column=5).value
            
            if not forfait and not prestataire and not designation:
                continue
                
            rows.append({
                "forfait": str(forfait or "").strip(),
                "prestataire": str(prestataire or "").strip(),
                "designation": str(designation or "").strip(),
                "montant": _parse_num(montant),
                "id_circuit": id_circuit,
            })
        
        return rows
    except Exception as e:
        logger.error(
            f"Failed to load collective expenses data: {e}", exc_info=True
        )
        return []
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def get_collective_expense_prestataires():
    """
    Get unique prestataires from Frais collectifs sheet
    
    Returns:
        list: Sorted list of unique prestataire names
    """
    data = load_collective_expenses_data()
    prestataires = set()
    for row in data:
        if row.get("prestataire"):
            prestataires.add(row["prestataire"])
    return sorted(prestataires)


def get_collective_expense_designations(prestataire=None):
    """
    Get designations, optionally filtered by prestataire
    
    Args:
        prestataire (str): Optional prestataire to filter by
    
    Returns:
        list: Sorted list of designations
    """
    data = load_collective_expenses_data()
    designations = set()
    for row in data:
        if prestataire is None or row.get("prestataire") == prestataire:
            if row.get("designation"):
                designations.add(row["designation"])
    return sorted(designations)


def get_collective_expense_montant(prestataire, designation):
    """
    Get montant for a given prestataire + designation combo
    
    Args:
        prestataire (str): Prestataire name
        designation (str): Designation
    
    Returns:
        float: Montant value, or 0 if not found
    """
    data = load_collective_expenses_data()
    for row in data:
        if (
            row.get("prestataire") == prestataire
            and row.get("designation") == designation
        ):
            return row.get("montant", 0)
    return 0


def get_collective_expense_forfait(prestataire, designation):
    """
    Get forfait for a given prestataire + designation combo
    
    Args:
        prestataire (str): Prestataire name
        designation (str): Designation
    
    Returns:
        str: Forfait value, or empty string if not found
    """
    data = load_collective_expenses_data()
    for row in data:
        if (
            row.get("prestataire") == prestataire
            and row.get("designation") == designation
        ):
            return str(row.get("forfait", "")).strip()
    return ""


def load_collective_expense_db_rows():
    """
    Load raw DB rows from data-hotel.xlsx / Frais collectifs sheet.

    Returns:
        list: row dictionaries with row_number and DB fields.
    """
    if not OPENPYXL_AVAILABLE:
        return []

    if not os.path.exists(HOTEL_EXCEL_PATH):
        return []

    wb = None
    try:
        wb = load_workbook(HOTEL_EXCEL_PATH)
        if FRAIS_COLLECTIFS_SHEET_NAME not in wb.sheetnames:
            return []

        ws = wb[FRAIS_COLLECTIFS_SHEET_NAME]
        rows = []
        for row_idx in range(2, ws.max_row + 1):
            forfait = ws.cell(row=row_idx, column=1).value
            prestataire = ws.cell(row=row_idx, column=2).value
            designation = ws.cell(row=row_idx, column=3).value
            montant = ws.cell(row=row_idx, column=4).value
            id_circuit = ws.cell(row=row_idx, column=5).value

            if all(v in (None, "") for v in [forfait, prestataire, designation, montant, id_circuit]):
                continue

            rows.append(
                {
                    "row_number": row_idx,
                    "forfait": str(forfait or "").strip(),
                    "prestataire": str(prestataire or "").strip(),
                    "designation": str(designation or "").strip(),
                    "montant": _parse_num(montant),
                    "id_circuit": "" if id_circuit is None else str(id_circuit).strip(),
                }
            )
        return rows
    except Exception as e:
        logger.error(f"Failed to load collective expense DB rows: {e}", exc_info=True)
        return []
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def save_collective_expense_db_row(row_data):
    """Insert one row into data-hotel.xlsx / Frais collectifs."""
    if not OPENPYXL_AVAILABLE:
        return -1

    wb = None
    try:
        if not os.path.exists(HOTEL_EXCEL_PATH):
            wb = Workbook()
            ws = wb.active
            ws.title = FRAIS_COLLECTIFS_SHEET_NAME
        else:
            wb = load_workbook(HOTEL_EXCEL_PATH)
            if FRAIS_COLLECTIFS_SHEET_NAME not in wb.sheetnames:
                ws = wb.create_sheet(FRAIS_COLLECTIFS_SHEET_NAME)
            else:
                ws = wb[FRAIS_COLLECTIFS_SHEET_NAME]

        if ws.max_row == 1 and ws.cell(row=1, column=1).value in (None, ""):
            headers = ["FORFAIT", "PRESTATAIRES", "DESIGNATION", "MONTANT", "ID circuit"]
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx, value=header)

        next_row = ws.max_row + 1
        ws.cell(row=next_row, column=1, value=row_data.get("forfait", ""))
        ws.cell(row=next_row, column=2, value=row_data.get("prestataire", ""))
        ws.cell(row=next_row, column=3, value=row_data.get("designation", ""))
        ws.cell(row=next_row, column=4, value=_parse_num(row_data.get("montant", 0)))
        ws.cell(row=next_row, column=5, value=row_data.get("id_circuit", ""))

        wb.save(HOTEL_EXCEL_PATH)
        return next_row
    except PermissionError:
        return -2
    except Exception as e:
        logger.error(f"Failed to save collective expense DB row: {e}", exc_info=True)
        return -1
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def update_collective_expense_db_row(row_number, row_data):
    """Update one row in data-hotel.xlsx / Frais collectifs."""
    if not OPENPYXL_AVAILABLE:
        return -1

    if not os.path.exists(HOTEL_EXCEL_PATH):
        return -1

    wb = None
    try:
        wb = load_workbook(HOTEL_EXCEL_PATH)
        if FRAIS_COLLECTIFS_SHEET_NAME not in wb.sheetnames:
            return -1

        ws = wb[FRAIS_COLLECTIFS_SHEET_NAME]
        ws.cell(row=row_number, column=1, value=row_data.get("forfait", ""))
        ws.cell(row=row_number, column=2, value=row_data.get("prestataire", ""))
        ws.cell(row=row_number, column=3, value=row_data.get("designation", ""))
        ws.cell(row=row_number, column=4, value=_parse_num(row_data.get("montant", 0)))
        if "id_circuit" in row_data:
            ws.cell(row=row_number, column=5, value=row_data.get("id_circuit", ""))

        wb.save(HOTEL_EXCEL_PATH)
        return 0
    except PermissionError:
        return -2
    except Exception as e:
        logger.error(f"Failed to update collective expense DB row {row_number}: {e}", exc_info=True)
        return -1
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def delete_collective_expense_db_row(row_number):
    """Delete one row from data-hotel.xlsx / Frais collectifs."""
    if not OPENPYXL_AVAILABLE:
        return False

    if not os.path.exists(HOTEL_EXCEL_PATH):
        return False

    wb = None
    try:
        wb = load_workbook(HOTEL_EXCEL_PATH)
        if FRAIS_COLLECTIFS_SHEET_NAME not in wb.sheetnames:
            return False

        ws = wb[FRAIS_COLLECTIFS_SHEET_NAME]
        ws.delete_rows(row_number)
        wb.save(HOTEL_EXCEL_PATH)
        return True
    except Exception as e:
        logger.error(f"Failed to delete collective expense DB row {row_number}: {e}", exc_info=True)
        return False
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def update_collective_expense_quotation_in_excel(row_number, form_data):
    """
    Update an existing collective expense quotation in Excel
    
    Args:
        row_number (int): Row number to update (1-indexed, excluding header)
        form_data (dict): Form data with headers as keys
    
    Returns:
        int: 0 success, -1 error, -2 PermissionError
    """
    if not OPENPYXL_AVAILABLE:
        logger.warning("openpyxl not available. Cannot update Excel.")
        return -1
    
    if not os.path.exists(CLIENT_EXCEL_PATH):
        logger.error(f"Excel file {CLIENT_EXCEL_PATH} not found")
        return -1
    
    wb = None
    try:
        wb = load_workbook(CLIENT_EXCEL_PATH)
        if COTATION_FRAIS_COL_SHEET_NAME not in wb.sheetnames:
            logger.error(f"Sheet {COTATION_FRAIS_COL_SHEET_NAME} not found")
            return -1
        
        ws = wb[COTATION_FRAIS_COL_SHEET_NAME]
        headers = get_collective_expense_headers()
        
        # Excel row is data row + 1 (for header)
        excel_row = row_number + 1
        
        for col_idx, header in enumerate(headers, start=1):
            value = form_data.get(header, "")
            ws.cell(row=excel_row, column=col_idx, value=value)
        
        wb.save(CLIENT_EXCEL_PATH)
        logger.info(f"Updated collective expense at row {row_number}")
        return 0
    except PermissionError:
        logger.error(f"Permission error updating collective expense at row {row_number}")
        return -2
    except Exception as e:
        logger.error(f"Error updating collective expense at row {row_number}: {e}", exc_info=True)
        return -1
    finally:
        if wb:
            wb.close()


def delete_collective_expense_from_excel(row_number):
    """
    Delete a collective expense quotation from Excel
    
    Args:
        row_number (int): Row number to delete (1-indexed in data, excludes header)
    
    Returns:
        bool: True if successful, False otherwise
    """
    if not OPENPYXL_AVAILABLE:
        logger.warning("openpyxl not available. Cannot delete from Excel.")
        return False
    
    if not os.path.exists(CLIENT_EXCEL_PATH):
        logger.error(f"Excel file {CLIENT_EXCEL_PATH} not found")
        return False
    
    wb = None
    try:
        wb = load_workbook(CLIENT_EXCEL_PATH)
        if COTATION_FRAIS_COL_SHEET_NAME not in wb.sheetnames:
            logger.error(f"Sheet {COTATION_FRAIS_COL_SHEET_NAME} not found")
            return False
        
        ws = wb[COTATION_FRAIS_COL_SHEET_NAME]
        # Excel row is data row + 1 (for header)
        excel_row = row_number + 1
        ws.delete_rows(excel_row)
        
        wb.save(CLIENT_EXCEL_PATH)
        logger.info(f"Deleted collective expense at row {row_number}")
        return True
    except Exception as e:
        logger.error(f"Error deleting collective expense at row {row_number}: {e}", exc_info=True)
        return False
    finally:
        if wb:
            wb.close()


def load_visite_excursion_data():
    """
    Load all data from Visite_excursion sheet in data-hotel.xlsx.

    Returns:
        list: List of dicts with prestation, designation, tarif_par_pax and raw fields
    """
    if not OPENPYXL_AVAILABLE:
        logger.warning("openpyxl not available. Cannot load visite & excursion data.")
        return []

    if not os.path.exists(HOTEL_EXCEL_PATH):
        return []

    wb = None
    try:
        wb = load_workbook(HOTEL_EXCEL_PATH)
        if VISITE_EXCURSION_SOURCE_SHEET_NAME not in wb.sheetnames:
            return []

        ws = wb[VISITE_EXCURSION_SOURCE_SHEET_NAME]

        header_index = {}
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=1, column=col).value
            if value is None:
                continue
            normalized = (
                str(value)
                .strip()
                .lower()
                .replace("_", " ")
                .replace("-", " ")
                .replace("é", "e")
                .replace("è", "e")
                .replace("ê", "e")
            )
            header_index[normalized] = col

        def _find_col(candidates):
            for candidate in candidates:
                normalized = (
                    str(candidate)
                    .strip()
                    .lower()
                    .replace("_", " ")
                    .replace("-", " ")
                    .replace("é", "e")
                    .replace("è", "e")
                    .replace("ê", "e")
                )
                if normalized in header_index:
                    return header_index[normalized]
            return None

        prestation_col = _find_col([
            "prestation",
            "prestations",
            "prestataire",
            "prestataires",
            "fournisseur",
            "provider",
        ])
        designation_col = _find_col(["designation", "désignation", "service", "libelle", "description"])
        tarif_col = _find_col(["tarif par pax", "tarif/pax", "tarif pax", "montant", "prix", "price"])

        if not prestation_col or not designation_col:
            return []

        rows = []
        for row_idx in range(2, ws.max_row + 1):
            prestation = ws.cell(row=row_idx, column=prestation_col).value
            designation = ws.cell(row=row_idx, column=designation_col).value
            tarif = ws.cell(row=row_idx, column=tarif_col).value if tarif_col else 0

            if not prestation and not designation:
                continue

            raw_fields = {}
            for col in range(1, ws.max_column + 1):
                h = ws.cell(row=1, column=col).value
                if h is None:
                    continue
                hk = (
                    str(h)
                    .strip()
                    .lower()
                    .replace("_", " ")
                    .replace("-", " ")
                    .replace("é", "e")
                    .replace("è", "e")
                    .replace("ê", "e")
                )
                raw_fields[hk] = str(ws.cell(row=row_idx, column=col).value or "").strip()

            rows.append(
                {
                    "prestation": str(prestation or "").strip(),
                    "prestataire": str(prestation or "").strip(),
                    "designation": str(designation or "").strip(),
                    "tarif_par_pax": _parse_num(tarif),
                    "fields": raw_fields,
                }
            )

        return rows
    except Exception as e:
        logger.error(f"Failed to load visite & excursion data: {e}", exc_info=True)
        return []
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def _normalize_visite_key(value):
    return (
        str(value or "")
        .strip()
        .lower()
        .replace("_", " ")
        .replace("-", " ")
        .replace("é", "e")
        .replace("è", "e")
        .replace("ê", "e")
    )


def _match_visite_filters(row, filters=None, ignore_keys=None):
    if not filters:
        return True

    ignore = {_normalize_visite_key(k) for k in (ignore_keys or [])}
    fields = row.get("fields", {})

    for key, expected in filters.items():
        if expected in (None, ""):
            continue

        nk = _normalize_visite_key(key)
        if nk in ignore:
            continue

        expected_text = str(expected).strip()
        if not expected_text:
            continue

        if nk in {"prestation", "prestations", "prestataire", "prestataires"}:
            if str(row.get("prestation") or "").strip() != expected_text:
                return False
            continue

        if nk in {"designation", "designations", "désignation"}:
            if str(row.get("designation") or "").strip() != expected_text:
                return False
            continue

        field_value = str(fields.get(nk, "")).strip()
        if field_value != expected_text:
            return False

    return True


def get_visite_excursion_prestataires(filters=None):
    data = load_visite_excursion_data()
    values = {
        row.get("prestation")
        for row in data
        if row.get("prestation") and _match_visite_filters(row, filters, ignore_keys=["prestation", "prestataire"])
    }
    return sorted(values)


def get_visite_excursion_designations(prestataire=None, filters=None):
    data = load_visite_excursion_data()
    effective_filters = dict(filters or {})
    if prestataire:
        effective_filters["prestations"] = prestataire

    values = set()
    for row in data:
        if _match_visite_filters(row, effective_filters, ignore_keys=["designation", "designations", "désignation"]):
            designation = row.get("designation")
            if designation:
                values.add(designation)
    return sorted(values)


def get_visite_excursion_montant(prestataire, designation, filters=None):
    data = load_visite_excursion_data()
    effective_filters = dict(filters or {})
    if prestataire:
        effective_filters["prestations"] = prestataire
    if designation:
        effective_filters["designation"] = designation

    for row in data:
        if _match_visite_filters(row, effective_filters):
            return row.get("tarif_par_pax", 0)
    return 0


def get_visite_excursion_db_headers():
    """Load header labels from data-hotel.xlsx / Visite_excursion sheet."""
    if not OPENPYXL_AVAILABLE:
        return []

    if not os.path.exists(HOTEL_EXCEL_PATH):
        return []

    wb = None
    try:
        wb = load_workbook(HOTEL_EXCEL_PATH)
        source_sheet = None
        if VISITE_EXCURSION_SOURCE_SHEET_NAME in wb.sheetnames:
            source_sheet = VISITE_EXCURSION_SOURCE_SHEET_NAME
        else:
            target = _normalize_header_key(VISITE_EXCURSION_SOURCE_SHEET_NAME)
            for sheet_name in wb.sheetnames:
                if _normalize_header_key(sheet_name) == target:
                    source_sheet = sheet_name
                    break

        if not source_sheet:
            return []

        ws = wb[source_sheet]
        headers = []
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=1, column=col).value
            if value is None:
                continue
            label = str(value).strip()
            if label:
                headers.append(label)
        return headers
    except Exception as e:
        logger.error(f"Failed to load visite excursion DB headers: {e}", exc_info=True)
        return []
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def load_visite_excursion_db_rows():
    """Load all DB rows from data-hotel.xlsx / Visite_excursion sheet."""
    if not OPENPYXL_AVAILABLE:
        return []

    if not os.path.exists(HOTEL_EXCEL_PATH):
        return []

    wb = None
    try:
        wb = load_workbook(HOTEL_EXCEL_PATH)
        source_sheet = None
        if VISITE_EXCURSION_SOURCE_SHEET_NAME in wb.sheetnames:
            source_sheet = VISITE_EXCURSION_SOURCE_SHEET_NAME
        else:
            target = _normalize_header_key(VISITE_EXCURSION_SOURCE_SHEET_NAME)
            for sheet_name in wb.sheetnames:
                if _normalize_header_key(sheet_name) == target:
                    source_sheet = sheet_name
                    break

        if not source_sheet:
            return []

        ws = wb[source_sheet]
        headers = get_visite_excursion_db_headers()
        if not headers:
            return []

        rows = []
        for row_idx in range(2, ws.max_row + 1):
            row_dict = {"row_number": row_idx}
            has_values = False
            for col_idx, header in enumerate(headers, start=1):
                value = ws.cell(row=row_idx, column=col_idx).value
                if value not in (None, ""):
                    has_values = True
                row_dict[header] = "" if value is None else value
            if has_values:
                rows.append(row_dict)
        return rows
    except Exception as e:
        logger.error(f"Failed to load visite excursion DB rows: {e}", exc_info=True)
        return []
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def save_visite_excursion_db_row(row_data):
    """Insert one DB row into data-hotel.xlsx / Visite_excursion sheet."""
    if not OPENPYXL_AVAILABLE:
        return -1

    wb = None
    try:
        if not os.path.exists(HOTEL_EXCEL_PATH):
            wb = Workbook()
            ws = wb.active
            ws.title = VISITE_EXCURSION_SOURCE_SHEET_NAME
        else:
            wb = load_workbook(HOTEL_EXCEL_PATH)
            if VISITE_EXCURSION_SOURCE_SHEET_NAME not in wb.sheetnames:
                ws = wb.create_sheet(VISITE_EXCURSION_SOURCE_SHEET_NAME)
            else:
                ws = wb[VISITE_EXCURSION_SOURCE_SHEET_NAME]

        headers = []
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=1, column=col).value
            if value is None:
                continue
            label = str(value).strip()
            if label:
                headers.append(label)

        if not headers:
            headers = [key for key in row_data.keys() if key != "row_number"]
            if not headers:
                headers = ["PRESTATIONS", "DESIGNATION", "Tarif par pax"]
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx, value=header)

        next_row = ws.max_row + 1
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=next_row, column=col_idx, value=row_data.get(header, ""))

        wb.save(HOTEL_EXCEL_PATH)
        return next_row
    except PermissionError:
        return -2
    except Exception as e:
        logger.error(f"Failed to save visite excursion DB row: {e}", exc_info=True)
        return -1
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def update_visite_excursion_db_row(row_number, row_data):
    """Update one DB row in data-hotel.xlsx / Visite_excursion sheet."""
    if not OPENPYXL_AVAILABLE:
        return -1

    if not os.path.exists(HOTEL_EXCEL_PATH):
        return -1

    wb = None
    try:
        wb = load_workbook(HOTEL_EXCEL_PATH)
        if VISITE_EXCURSION_SOURCE_SHEET_NAME not in wb.sheetnames:
            return -1

        ws = wb[VISITE_EXCURSION_SOURCE_SHEET_NAME]
        headers = get_visite_excursion_db_headers()
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_number, column=col_idx, value=row_data.get(header, ""))

        wb.save(HOTEL_EXCEL_PATH)
        return 0
    except PermissionError:
        return -2
    except Exception as e:
        logger.error(f"Failed to update visite excursion DB row {row_number}: {e}", exc_info=True)
        return -1
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def delete_visite_excursion_db_row(row_number):
    """Delete one DB row from data-hotel.xlsx / Visite_excursion sheet."""
    if not OPENPYXL_AVAILABLE:
        return False

    if not os.path.exists(HOTEL_EXCEL_PATH):
        return False

    wb = None
    try:
        wb = load_workbook(HOTEL_EXCEL_PATH)
        if VISITE_EXCURSION_SOURCE_SHEET_NAME not in wb.sheetnames:
            return False

        ws = wb[VISITE_EXCURSION_SOURCE_SHEET_NAME]
        ws.delete_rows(row_number)
        wb.save(HOTEL_EXCEL_PATH)
        return True
    except Exception as e:
        logger.error(f"Failed to delete visite excursion DB row {row_number}: {e}", exc_info=True)
        return False
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def get_avion_db_headers():
    """Load header labels from data-hotel.xlsx / avion sheet."""
    if not OPENPYXL_AVAILABLE:
        return []

    if not os.path.exists(HOTEL_EXCEL_PATH):
        return []

    wb = None
    try:
        wb = load_workbook(HOTEL_EXCEL_PATH)
        source_sheet = None
        if AVION_SOURCE_SHEET_NAME in wb.sheetnames:
            source_sheet = AVION_SOURCE_SHEET_NAME
        else:
            target = _normalize_header_key(AVION_SOURCE_SHEET_NAME)
            for sheet_name in wb.sheetnames:
                if _normalize_header_key(sheet_name) == target:
                    source_sheet = sheet_name
                    break

        if not source_sheet:
            return []

        ws = wb[source_sheet]
        headers = []
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=1, column=col).value
            if value is None:
                continue
            label = str(value).strip()
            if label:
                headers.append(label)
        return headers
    except Exception as e:
        logger.error(f"Failed to load avion DB headers: {e}", exc_info=True)
        return []
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def load_avion_db_rows():
    """Load all DB rows from data-hotel.xlsx / avion sheet."""
    if not OPENPYXL_AVAILABLE:
        return []

    if not os.path.exists(HOTEL_EXCEL_PATH):
        return []

    wb = None
    try:
        wb = load_workbook(HOTEL_EXCEL_PATH)
        source_sheet = None
        if AVION_SOURCE_SHEET_NAME in wb.sheetnames:
            source_sheet = AVION_SOURCE_SHEET_NAME
        else:
            target = _normalize_header_key(AVION_SOURCE_SHEET_NAME)
            for sheet_name in wb.sheetnames:
                if _normalize_header_key(sheet_name) == target:
                    source_sheet = sheet_name
                    break

        if not source_sheet:
            return []

        ws = wb[source_sheet]
        headers = get_avion_db_headers()
        if not headers:
            return []

        rows = []
        for row_idx in range(2, ws.max_row + 1):
            row_dict = {"row_number": row_idx}
            has_values = False
            for col_idx, header in enumerate(headers, start=1):
                value = ws.cell(row=row_idx, column=col_idx).value
                if value not in (None, ""):
                    has_values = True
                row_dict[header] = "" if value is None else value
            if has_values:
                rows.append(row_dict)
        return rows
    except Exception as e:
        logger.error(f"Failed to load avion DB rows: {e}", exc_info=True)
        return []
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def save_avion_db_row(row_data):
    """Insert one DB row into data-hotel.xlsx / avion sheet."""
    if not OPENPYXL_AVAILABLE:
        return -1

    wb = None
    try:
        if not os.path.exists(HOTEL_EXCEL_PATH):
            wb = Workbook()
            ws = wb.active
            ws.title = AVION_SOURCE_SHEET_NAME
        else:
            wb = load_workbook(HOTEL_EXCEL_PATH)
            if AVION_SOURCE_SHEET_NAME not in wb.sheetnames:
                ws = wb.create_sheet(AVION_SOURCE_SHEET_NAME)
            else:
                ws = wb[AVION_SOURCE_SHEET_NAME]

        headers = []
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=1, column=col).value
            if value is None:
                continue
            label = str(value).strip()
            if label:
                headers.append(label)

        if not headers:
            headers = [key for key in row_data.keys() if key != "row_number"]
            if not headers:
                headers = [
                    "Ville de départ",
                    "Ville d'arrivée",
                    "Tarif Adultes",
                    "Tarif Enfants",
                ]
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx, value=header)

        next_row = ws.max_row + 1
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=next_row, column=col_idx, value=row_data.get(header, ""))

        wb.save(HOTEL_EXCEL_PATH)
        return next_row
    except PermissionError:
        return -2
    except Exception as e:
        logger.error(f"Failed to save avion DB row: {e}", exc_info=True)
        return -1
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def update_avion_db_row(row_number, row_data):
    """Update one DB row in data-hotel.xlsx / avion sheet."""
    if not OPENPYXL_AVAILABLE:
        return -1

    if not os.path.exists(HOTEL_EXCEL_PATH):
        return -1

    wb = None
    try:
        wb = load_workbook(HOTEL_EXCEL_PATH)
        if AVION_SOURCE_SHEET_NAME not in wb.sheetnames:
            return -1

        ws = wb[AVION_SOURCE_SHEET_NAME]
        headers = get_avion_db_headers()
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_number, column=col_idx, value=row_data.get(header, ""))

        wb.save(HOTEL_EXCEL_PATH)
        return 0
    except PermissionError:
        return -2
    except Exception as e:
        logger.error(f"Failed to update avion DB row {row_number}: {e}", exc_info=True)
        return -1
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def delete_avion_db_row(row_number):
    """Delete one DB row from data-hotel.xlsx / avion sheet."""
    if not OPENPYXL_AVAILABLE:
        return False

    if not os.path.exists(HOTEL_EXCEL_PATH):
        return False

    wb = None
    try:
        wb = load_workbook(HOTEL_EXCEL_PATH)
        if AVION_SOURCE_SHEET_NAME not in wb.sheetnames:
            return False

        ws = wb[AVION_SOURCE_SHEET_NAME]
        ws.delete_rows(row_number)
        wb.save(HOTEL_EXCEL_PATH)
        return True
    except Exception as e:
        logger.error(f"Failed to delete avion DB row {row_number}: {e}", exc_info=True)
        return False
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def get_visite_excursion_headers():
    """
    Load header list from VISITE_EXCURSION sheet.
    """
    if not OPENPYXL_AVAILABLE:
        return []

    if not os.path.exists(CLIENT_EXCEL_PATH):
        return []

    wb = None
    try:
        wb = load_workbook(CLIENT_EXCEL_PATH)
        if VISITE_EXCURSION_SHEET_NAME not in wb.sheetnames:
            return []

        ws = wb[VISITE_EXCURSION_SHEET_NAME]
        headers = []
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=1, column=col).value
            if value is None:
                continue
            label = str(value).strip()
            if label:
                headers.append(label)
        return headers
    except Exception as e:
        logger.error(f"Failed to load visite & excursion headers: {e}", exc_info=True)
        return []
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def save_visite_excursion_quotation_to_excel(form_data):
    """
    Save a visite & excursion quotation row into VISITE_EXCURSION sheet.

    Returns:
        int: Saved row number, -1 on failure, -2 if file locked.
    """
    if not OPENPYXL_AVAILABLE:
        return -1

    wb = None
    try:
        if not os.path.exists(CLIENT_EXCEL_PATH):
            wb = Workbook()
            ws = wb.active
            ws.title = VISITE_EXCURSION_SHEET_NAME
        else:
            wb = load_workbook(CLIENT_EXCEL_PATH)
            if VISITE_EXCURSION_SHEET_NAME not in wb.sheetnames:
                ws = wb.create_sheet(VISITE_EXCURSION_SHEET_NAME)
            else:
                ws = wb[VISITE_EXCURSION_SHEET_NAME]

        headers = []
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=1, column=col).value
            if value is None:
                continue
            label = str(value).strip()
            if label:
                headers.append(label)

        if not headers:
            headers = list(form_data.keys())
            if "Date" not in headers:
                headers.insert(0, "Date")
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    start_color="4472C4", end_color="4472C4", fill_type="solid"
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")

        header_map = {header: index for index, header in enumerate(headers, start=1)}

        next_row = 2
        while ws.cell(row=next_row, column=1).value not in (None, ""):
            next_row += 1

        if next_row > 1048576:
            logger.error("VISITE_EXCURSION sheet is full (max Excel rows reached).")
            return -1
        if "Date" in header_map and not form_data.get("Date"):
            form_data = {
                "Date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                **form_data,
            }

        for header, col in header_map.items():
            value = form_data.get(header, "")
            ws.cell(row=next_row, column=col, value=value)

        wb.save(CLIENT_EXCEL_PATH)
        return next_row
    except PermissionError:
        return -2
    except Exception as e:
        logger.error(f"Failed to save visite & excursion quotation: {e}", exc_info=True)
        return -1
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def load_all_visite_excursion_quotations():
    """
    Load all visite & excursion quotations from VISITE_EXCURSION.
    """
    if not OPENPYXL_AVAILABLE:
        return []

    if not os.path.exists(CLIENT_EXCEL_PATH):
        return []

    wb = None
    try:
        wb = load_workbook(CLIENT_EXCEL_PATH)
        if VISITE_EXCURSION_SHEET_NAME not in wb.sheetnames:
            return []

        ws = wb[VISITE_EXCURSION_SHEET_NAME]
        headers = get_visite_excursion_headers()
        if not headers:
            return []

        rows = []
        for row_index in range(2, ws.max_row + 1):
            row_dict = {"row_number": row_index}
            has_values = False

            for col_index, header in enumerate(headers, start=1):
                value = ws.cell(row=row_index, column=col_index).value
                if value not in (None, ""):
                    has_values = True
                row_dict[header] = "" if value is None else value

            if has_values:
                rows.append(row_dict)

        return rows
    except Exception as e:
        logger.error(f"Failed to load visite & excursion quotations: {e}", exc_info=True)
        return []
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def update_visite_excursion_quotation_in_excel(row_number, form_data):
    if not OPENPYXL_AVAILABLE:
        return -1

    if not os.path.exists(CLIENT_EXCEL_PATH):
        return -1

    wb = None
    try:
        wb = load_workbook(CLIENT_EXCEL_PATH)
        if VISITE_EXCURSION_SHEET_NAME not in wb.sheetnames:
            return -1

        ws = wb[VISITE_EXCURSION_SHEET_NAME]
        headers = get_visite_excursion_headers()
        excel_row = row_number

        for col_idx, header in enumerate(headers, start=1):
            value = form_data.get(header, "")
            ws.cell(row=excel_row, column=col_idx, value=value)

        wb.save(CLIENT_EXCEL_PATH)
        return 0
    except PermissionError:
        return -2
    except Exception as e:
        logger.error(f"Error updating visite & excursion row {row_number}: {e}", exc_info=True)
        return -1
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def delete_visite_excursion_from_excel(row_number):
    if not OPENPYXL_AVAILABLE:
        return False

    if not os.path.exists(CLIENT_EXCEL_PATH):
        return False

    wb = None
    try:
        wb = load_workbook(CLIENT_EXCEL_PATH)
        if VISITE_EXCURSION_SHEET_NAME not in wb.sheetnames:
            return False

        ws = wb[VISITE_EXCURSION_SHEET_NAME]
        ws.delete_rows(row_number)

        wb.save(CLIENT_EXCEL_PATH)
        return True
    except Exception as e:
        logger.error(f"Error deleting visite & excursion row {row_number}: {e}", exc_info=True)
        return False
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def load_avion_source_data():
    """
    Load pricing rows from avion sheet in data-hotel.xlsx.

    Returns:
        list: Rows with tarifs adultes/enfants and normalized raw fields.
    """
    if not OPENPYXL_AVAILABLE:
        logger.warning("openpyxl not available. Cannot load avion data.")
        return []

    if not os.path.exists(HOTEL_EXCEL_PATH):
        return []

    wb = None
    try:
        wb = load_workbook(HOTEL_EXCEL_PATH)

        source_sheet = None
        if AVION_SOURCE_SHEET_NAME in wb.sheetnames:
            source_sheet = AVION_SOURCE_SHEET_NAME
        else:
            normalized_target = _normalize_header_key(AVION_SOURCE_SHEET_NAME)
            for sheet_name in wb.sheetnames:
                if _normalize_header_key(sheet_name) == normalized_target:
                    source_sheet = sheet_name
                    break

        if not source_sheet:
            return []

        ws = wb[source_sheet]

        header_index = {}
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=1, column=col).value
            if value is None:
                continue
            normalized = _normalize_header_key(value)
            if normalized:
                header_index[normalized] = col

        def _find_col(candidates):
            for candidate in candidates:
                normalized = _normalize_header_key(candidate)
                if normalized in header_index:
                    return header_index[normalized]
            return None

        tarif_adulte_col = _find_col(
            [
                "tarif adultes",
                "tarif adulte",
                "prix adultes",
                "prix adulte",
                "adulte",
                "adultes",
            ]
        )
        tarif_enfant_col = _find_col(
            [
                "tarifs enfants",
                "tarif enfants",
                "tarif enfant",
                "prix enfants",
                "prix enfant",
                "tarif bebe",
                "tarif bebes",
                "tarif bébé",
                "tarif bébés",
                "prix bebe",
                "prix bebes",
                "prix bébé",
                "prix bébés",
                "enfant",
                "enfants",
            ]
        )

        rows = []
        for row_idx in range(2, ws.max_row + 1):
            fields = {}
            has_value = False
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header is None:
                    continue
                key = _normalize_header_key(header)
                value = ws.cell(row=row_idx, column=col).value
                fields[key] = "" if value is None else str(value).strip()
                if value not in (None, ""):
                    has_value = True

            if not has_value:
                continue

            tarif_adulte = (
                ws.cell(row=row_idx, column=tarif_adulte_col).value
                if tarif_adulte_col
                else 0
            )
            tarif_enfant = (
                ws.cell(row=row_idx, column=tarif_enfant_col).value
                if tarif_enfant_col
                else 0
            )

            rows.append(
                {
                    "tarif_adulte": _parse_num(tarif_adulte),
                    "tarif_enfant": _parse_num(tarif_enfant),
                    "fields": fields,
                }
            )

        return rows
    except Exception as e:
        logger.error(f"Failed to load avion source data: {e}", exc_info=True)
        return []
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def get_avion_tarifs(filters=None):
    """
    Return first matching adult/child tariffs from avion source.

    Args:
        filters (dict): Optional filters by source columns.

    Returns:
        tuple: (tarif_adulte, tarif_enfant)
    """
    data = load_avion_source_data()
    if not data:
        return 0, 0

    normalized_filters = {}
    for key, value in (filters or {}).items():
        if value in (None, ""):
            continue
        normalized_filters[_normalize_header_key(key)] = str(value).strip()

    for row in data:
        if not normalized_filters:
            return row.get("tarif_adulte", 0), row.get("tarif_enfant", 0)

        fields = row.get("fields", {})
        matches = True
        for key, expected in normalized_filters.items():
            if key in {
                "date",
                "id client",
                "id_client",
                "ref client",
                "reference",
                "nom",
                "nom client",
                "nombre adulte",
                "nombre adultes",
                "nombre enfant",
                "nombre enfants",
                "tarif adulte",
                "tarif adultes",
                "tarif enfant",
                "tarif enfants",
                "montant adulte",
                "montant adultes",
                "montant enfant",
                "montant enfants",
                "total",
                "observation",
            }:
                continue
            if str(fields.get(key, "")).strip() != expected:
                matches = False
                break

        if matches:
            return row.get("tarif_adulte", 0), row.get("tarif_enfant", 0)

    first = data[0]
    return first.get("tarif_adulte", 0), first.get("tarif_enfant", 0)


def get_avion_departure_cities(filters=None):
    """Get unique departure cities from avion source data."""
    data = load_avion_source_data()
    values = set()

    normalized_filters = {
        _normalize_header_key(k): str(v).strip()
        for k, v in (filters or {}).items()
        if v not in (None, "")
    }

    for row in data:
        fields = row.get("fields", {})

        matches = True
        for key, expected in normalized_filters.items():
            if key in {
                "ville de depart",
                "ville depart",
                "ville d depart",
            }:
                continue
            if str(fields.get(key, "")).strip() != expected:
                matches = False
                break

        if not matches:
            continue

        city = (
            fields.get("ville de depart")
            or fields.get("ville depart")
            or fields.get("ville d depart")
            or ""
        )
        city = str(city).strip()
        if city:
            values.add(city)

    return sorted(values)


def get_avion_arrival_cities(filters=None):
    """Get unique arrival cities from avion source data."""
    data = load_avion_source_data()
    values = set()

    normalized_filters = {
        _normalize_header_key(k): str(v).strip()
        for k, v in (filters or {}).items()
        if v not in (None, "")
    }

    for row in data:
        fields = row.get("fields", {})

        matches = True
        for key, expected in normalized_filters.items():
            if key in {
                "ville d arrive",
                "ville de arrive",
                "ville arrive",
                "ville d arrivee",
                "ville de arrivee",
                "ville arrivee",
            }:
                continue
            if str(fields.get(key, "")).strip() != expected:
                matches = False
                break

        if not matches:
            continue

        city = (
            fields.get("ville d arrive")
            or fields.get("ville de arrive")
            or fields.get("ville arrive")
            or fields.get("ville d arrivee")
            or fields.get("ville de arrivee")
            or fields.get("ville arrivee")
            or ""
        )
        city = str(city).strip()
        if city:
            values.add(city)

    return sorted(values)


def get_avion_headers():
    """Load header list from AVION sheet in data.xlsx."""
    if not OPENPYXL_AVAILABLE:
        return []

    if not os.path.exists(CLIENT_EXCEL_PATH):
        return []

    wb = None
    try:
        wb = load_workbook(CLIENT_EXCEL_PATH)
        if AVION_SHEET_NAME not in wb.sheetnames:
            return []

        ws = wb[AVION_SHEET_NAME]
        headers = []
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=1, column=col).value
            if value is None:
                continue
            label = str(value).strip()
            if label:
                headers.append(label)
        return headers
    except Exception as e:
        logger.error(f"Failed to load AVION headers: {e}", exc_info=True)
        return []
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def save_air_ticket_quotation_to_excel(form_data):
    """
    Save an air ticket quotation row into AVION sheet.

    Returns:
        int: Saved row number, -1 on failure, -2 if file locked.
    """
    if not OPENPYXL_AVAILABLE:
        return -1

    wb = None
    try:
        if not os.path.exists(CLIENT_EXCEL_PATH):
            wb = Workbook()
            ws = wb.active
            ws.title = AVION_SHEET_NAME
        else:
            wb = load_workbook(CLIENT_EXCEL_PATH)
            if AVION_SHEET_NAME not in wb.sheetnames:
                ws = wb.create_sheet(AVION_SHEET_NAME)
            else:
                ws = wb[AVION_SHEET_NAME]

        headers = []
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=1, column=col).value
            if value is None:
                continue
            label = str(value).strip()
            if label:
                headers.append(label)

        if not headers:
            headers = list(form_data.keys())
            if "Date" not in headers:
                headers.insert(0, "Date")
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    start_color="4472C4", end_color="4472C4", fill_type="solid"
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")

        header_map = {header: index for index, header in enumerate(headers, start=1)}

        next_row = 2
        while ws.cell(row=next_row, column=1).value not in (None, ""):
            next_row += 1

        if next_row > 1048576:
            logger.error("AVION sheet is full (max Excel rows reached).")
            return -1

        if "Date" in header_map and not form_data.get("Date"):
            form_data = {
                "Date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                **form_data,
            }

        for header, col in header_map.items():
            value = form_data.get(header, "")
            ws.cell(row=next_row, column=col, value=value)

        wb.save(CLIENT_EXCEL_PATH)
        return next_row
    except PermissionError:
        return -2
    except Exception as e:
        logger.error(f"Failed to save air ticket quotation: {e}", exc_info=True)
        return -1
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def load_all_air_ticket_quotations():
    """Load all air ticket quotations from AVION sheet."""
    if not OPENPYXL_AVAILABLE:
        return []

    if not os.path.exists(CLIENT_EXCEL_PATH):
        return []

    wb = None
    try:
        wb = load_workbook(CLIENT_EXCEL_PATH)
        if AVION_SHEET_NAME not in wb.sheetnames:
            return []

        ws = wb[AVION_SHEET_NAME]
        headers = get_avion_headers()
        if not headers:
            return []

        rows = []
        for row_index in range(2, ws.max_row + 1):
            row_dict = {"row_number": row_index}
            has_values = False

            for col_index, header in enumerate(headers, start=1):
                value = ws.cell(row=row_index, column=col_index).value
                if value not in (None, ""):
                    has_values = True
                row_dict[header] = "" if value is None else value

            if has_values:
                rows.append(row_dict)

        return rows
    except Exception as e:
        logger.error(f"Failed to load air ticket quotations: {e}", exc_info=True)
        return []
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def update_air_ticket_quotation_in_excel(row_number, form_data):
    """Update one air ticket quotation row in AVION sheet."""
    if not OPENPYXL_AVAILABLE:
        return -1

    if not os.path.exists(CLIENT_EXCEL_PATH):
        return -1

    wb = None
    try:
        wb = load_workbook(CLIENT_EXCEL_PATH)
        if AVION_SHEET_NAME not in wb.sheetnames:
            return -1

        ws = wb[AVION_SHEET_NAME]
        headers = get_avion_headers()
        excel_row = row_number

        for col_idx, header in enumerate(headers, start=1):
            value = form_data.get(header, "")
            ws.cell(row=excel_row, column=col_idx, value=value)

        wb.save(CLIENT_EXCEL_PATH)
        return 0
    except PermissionError:
        return -2
    except Exception as e:
        logger.error(f"Error updating air ticket row {row_number}: {e}", exc_info=True)
        return -1
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def delete_air_ticket_from_excel(row_number):
    """Delete one air ticket quotation row from AVION sheet."""
    if not OPENPYXL_AVAILABLE:
        return False

    if not os.path.exists(CLIENT_EXCEL_PATH):
        return False

    wb = None
    try:
        wb = load_workbook(CLIENT_EXCEL_PATH)
        if AVION_SHEET_NAME not in wb.sheetnames:
            return False

        ws = wb[AVION_SHEET_NAME]
        ws.delete_rows(row_number)

        wb.save(CLIENT_EXCEL_PATH)
        return True
    except Exception as e:
        logger.error(f"Error deleting air ticket row {row_number}: {e}", exc_info=True)
        return False
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


PARAMETRAGE_DEFAULT_HEADERS = ["parametre", "valeur"]


def _ensure_parametrage_sheet(ws):
    header_style = {
        "font": Font(bold=True, color="FFFFFF"),
        "fill": PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid"),
        "alignment": Alignment(horizontal="center", vertical="center"),
    }
    return _ensure_headers(ws, PARAMETRAGE_DEFAULT_HEADERS, header_style)


def _normalize_param_name(value):
    return _normalize_header_key(value)


def _ensure_default_param_rows(ws, header_map):
    parameter_col = _find_header_column(
        header_map,
        "parametre",
        "paramètre",
        "PARAMETRE",
        "Parametre",
        "Paramètre",
    )
    value_col = _find_header_column(
        header_map,
        "valeur",
        "value",
        "VALEUR",
        "Valeur",
        "Value",
    )
    if not parameter_col or not value_col:
        return

    existing = set()
    for row in range(2, ws.max_row + 1):
        param_name = ws.cell(row=row, column=parameter_col).value
        if param_name not in (None, ""):
            existing.add(_normalize_param_name(param_name))

    defaults = ["Prix Essence", "Prix Gasoil"]
    for param in defaults:
        if _normalize_param_name(param) in existing:
            continue
        target_row = 2
        while ws.cell(row=target_row, column=parameter_col).value not in (None, ""):
            target_row += 1
        ws.cell(row=target_row, column=parameter_col, value=param)
        ws.cell(row=target_row, column=value_col, value="")


def get_parametrage_headers():
    if not OPENPYXL_AVAILABLE:
        return []

    wb = None
    try:
        changed = False
        if not os.path.exists(HOTEL_EXCEL_PATH):
            wb = Workbook()
            ws = wb.active
            ws.title = PARAMETRAGE_SHEET_NAME
            changed = True
        else:
            wb = load_workbook(HOTEL_EXCEL_PATH)
            if PARAMETRAGE_SHEET_NAME not in wb.sheetnames:
                ws = wb.create_sheet(PARAMETRAGE_SHEET_NAME)
                changed = True
            else:
                ws = wb[PARAMETRAGE_SHEET_NAME]

        before_headers = _get_header_map(ws)
        before_row_count = ws.max_row
        header_map = _ensure_parametrage_sheet(ws)
        _ensure_default_param_rows(ws, header_map)

        after_headers = _get_header_map(ws)
        if before_headers != after_headers or ws.max_row != before_row_count:
            changed = True

        if changed:
            wb.save(HOTEL_EXCEL_PATH)
        return list(PARAMETRAGE_DEFAULT_HEADERS)
    except PermissionError:
        try:
            wb = load_workbook(HOTEL_EXCEL_PATH, data_only=True)
            if PARAMETRAGE_SHEET_NAME not in wb.sheetnames:
                return list(PARAMETRAGE_DEFAULT_HEADERS)

            ws = wb[PARAMETRAGE_SHEET_NAME]
            header_map = _get_header_map(ws)
            parameter_col = _find_header_column(
                header_map,
                "parametre",
                "paramètre",
                "PARAMETRE",
                "Parametre",
                "Paramètre",
            )
            value_col = _find_header_column(
                header_map,
                "valeur",
                "value",
                "VALEUR",
                "Valeur",
                "Value",
            )
            if parameter_col and value_col:
                return list(PARAMETRAGE_DEFAULT_HEADERS)
            return []
        except Exception:
            return []
    except Exception as e:
        logger.error(f"Failed to load PARAMETRAGE headers: {e}", exc_info=True)
        return []
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def load_all_parametrages():
    if not OPENPYXL_AVAILABLE:
        return []

    wb = None
    try:
        if not os.path.exists(HOTEL_EXCEL_PATH):
            return []

        wb = load_workbook(HOTEL_EXCEL_PATH)
        if PARAMETRAGE_SHEET_NAME not in wb.sheetnames:
            return []

        ws = wb[PARAMETRAGE_SHEET_NAME]
        header_map = _get_header_map(ws)
        parameter_col = _find_header_column(
            header_map,
            "parametre",
            "paramètre",
            "PARAMETRE",
            "Parametre",
            "Paramètre",
        )
        value_col = _find_header_column(
            header_map,
            "valeur",
            "value",
            "VALEUR",
            "Valeur",
            "Value",
        )
        if not parameter_col or not value_col:
            return []

        rows = []
        for row in range(2, ws.max_row + 1):
            parameter = ws.cell(row=row, column=parameter_col).value
            value = ws.cell(row=row, column=value_col).value
            if parameter in (None, "") and value in (None, ""):
                continue
            rows.append(
                {
                    "row_number": row,
                    "PARAMETRE": "" if parameter is None else str(parameter).strip(),
                    "VALEUR": "" if value is None else value,
                }
            )

        return rows
    except Exception as e:
        logger.error(f"Failed to load PARAMETRAGE rows: {e}", exc_info=True)
        return []
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def save_parametrage_to_excel(form_data):
    if not OPENPYXL_AVAILABLE:
        return -1

    wb = None
    try:
        if not os.path.exists(HOTEL_EXCEL_PATH):
            wb = Workbook()
            ws = wb.active
            ws.title = PARAMETRAGE_SHEET_NAME
        else:
            wb = load_workbook(HOTEL_EXCEL_PATH)
            if PARAMETRAGE_SHEET_NAME not in wb.sheetnames:
                ws = wb.create_sheet(PARAMETRAGE_SHEET_NAME)
            else:
                ws = wb[PARAMETRAGE_SHEET_NAME]

        header_map = _ensure_parametrage_sheet(ws)
        _ensure_default_param_rows(ws, header_map)

        parameter_col = _find_header_column(
            header_map,
            "parametre",
            "paramètre",
            "PARAMETRE",
            "Parametre",
            "Paramètre",
        )
        value_col = _find_header_column(
            header_map,
            "valeur",
            "value",
            "VALEUR",
            "Valeur",
            "Value",
        )
        if not parameter_col or not value_col:
            return -1

        parameter = str(form_data.get("PARAMETRE", "")).strip()
        value = form_data.get("VALEUR", "")
        if not parameter:
            return -1

        normalized_parameter = _normalize_param_name(parameter)

        target_row = None
        for row in range(2, ws.max_row + 1):
            current = ws.cell(row=row, column=parameter_col).value
            if _normalize_param_name(current) == normalized_parameter:
                target_row = row
                break

        if target_row is None:
            target_row = 2
            while ws.cell(row=target_row, column=parameter_col).value not in (None, ""):
                target_row += 1

        ws.cell(row=target_row, column=parameter_col, value=parameter)
        ws.cell(row=target_row, column=value_col, value=value)

        wb.save(HOTEL_EXCEL_PATH)
        return target_row
    except PermissionError:
        return -2
    except Exception as e:
        logger.error(f"Failed to save PARAMETRAGE row: {e}", exc_info=True)
        return -1
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def update_parametrage_in_excel(row_number, form_data):
    if not OPENPYXL_AVAILABLE:
        return -1

    if not os.path.exists(HOTEL_EXCEL_PATH):
        return -1

    wb = None
    try:
        wb = load_workbook(HOTEL_EXCEL_PATH)
        if PARAMETRAGE_SHEET_NAME not in wb.sheetnames:
            return -1

        ws = wb[PARAMETRAGE_SHEET_NAME]
        header_map = _ensure_parametrage_sheet(ws)
        parameter_col = _find_header_column(
            header_map,
            "parametre",
            "paramètre",
            "PARAMETRE",
            "Parametre",
            "Paramètre",
        )
        value_col = _find_header_column(
            header_map,
            "valeur",
            "value",
            "VALEUR",
            "Valeur",
            "Value",
        )
        if not parameter_col or not value_col:
            return -1

        ws.cell(row=row_number, column=parameter_col, value=form_data.get("PARAMETRE", ""))
        ws.cell(row=row_number, column=value_col, value=form_data.get("VALEUR", ""))

        wb.save(HOTEL_EXCEL_PATH)
        return 0
    except PermissionError:
        return -2
    except Exception as e:
        logger.error(f"Failed to update PARAMETRAGE row {row_number}: {e}", exc_info=True)
        return -1
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def delete_parametrage_from_excel(row_number):
    if not OPENPYXL_AVAILABLE:
        return False

    if not os.path.exists(HOTEL_EXCEL_PATH):
        return False

    wb = None
    try:
        wb = load_workbook(HOTEL_EXCEL_PATH)
        if PARAMETRAGE_SHEET_NAME not in wb.sheetnames:
            return False

        ws = wb[PARAMETRAGE_SHEET_NAME]
        ws.delete_rows(row_number)
        wb.save(HOTEL_EXCEL_PATH)
        return True
    except Exception as e:
        logger.error(f"Failed to delete PARAMETRAGE row {row_number}: {e}", exc_info=True)
        return False
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def _resolve_transport_source_header_map(ws):
    header_map = _get_header_map(ws, 1)
    prestataire_col = _find_header_column(header_map, "Prestataire")
    type_col = _find_header_column(header_map, "Type de voiture", "Type voiture")
    if prestataire_col and type_col:
        return header_map, 2

    header_map = _get_header_map(ws, 2)
    prestataire_col = _find_header_column(header_map, "Prestataire")
    type_col = _find_header_column(header_map, "Type de voiture", "Type voiture")
    if prestataire_col and type_col:
        return header_map, 3

    return {}, 2


def _load_transport_source_rows():
    if not OPENPYXL_AVAILABLE:
        return []

    if not os.path.exists(HOTEL_EXCEL_PATH):
        return []

    wb = None
    try:
        wb = load_workbook(HOTEL_EXCEL_PATH, data_only=True)
        if TRANSPORT_SOURCE_SHEET_NAME not in wb.sheetnames:
            return []

        ws = wb[TRANSPORT_SOURCE_SHEET_NAME]
        header_map, data_start_row = _resolve_transport_source_header_map(ws)
        if not header_map:
            return []

        prestataire_col = _find_header_column(header_map, "Prestataire")
        type_col = _find_header_column(header_map, "Type de voiture", "Type voiture")
        place_col = _find_header_column(header_map, "Nombre de place", "Nombre places", "Places")
        location_col = _find_header_column(header_map, "Location par jour", "Location/jour")
        consommation_col = _find_header_column(
            header_map,
            "Consommation",
            "CONSOMATION",
            "Consomation",
            "CONSO",
            "Conso",
        )
        energie_col = _find_header_column(header_map, "ENERGIE", "Energie")

        rows = []
        for row in range(data_start_row, ws.max_row + 1):
            prestataire = ws.cell(row=row, column=prestataire_col).value if prestataire_col else None
            type_voiture = ws.cell(row=row, column=type_col).value if type_col else None
            if not prestataire and not type_voiture:
                continue

            rows.append(
                {
                    "prestataire": str(prestataire or "").strip(),
                    "type_voiture": str(type_voiture or "").strip(),
                    "nombre_place": _parse_num(ws.cell(row=row, column=place_col).value) if place_col else 0,
                    "location_par_jour": _parse_num(ws.cell(row=row, column=location_col).value) if location_col else 0,
                    "consommation": _parse_num(ws.cell(row=row, column=consommation_col).value) if consommation_col else 0,
                    "energie": str(ws.cell(row=row, column=energie_col).value or "").strip() if energie_col else "",
                }
            )

        return rows
    except Exception as e:
        logger.error(f"Failed to load transport source rows: {e}", exc_info=True)
        return []
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def get_transport_prestataires():
    values = {
        row.get("prestataire")
        for row in _load_transport_source_rows()
        if row.get("prestataire")
    }
    return sorted(values)


def get_transport_vehicle_types(prestataire=None):
    rows = _load_transport_source_rows()
    values = set()
    for row in rows:
        if prestataire and row.get("prestataire") != prestataire:
            continue
        vehicle = row.get("type_voiture")
        if vehicle:
            values.add(vehicle)
    return sorted(values)


def get_transport_vehicle_data(prestataire, type_voiture):
    for row in _load_transport_source_rows():
        if row.get("prestataire") == prestataire and row.get("type_voiture") == type_voiture:
            return row
    return {
        "prestataire": str(prestataire or "").strip(),
        "type_voiture": str(type_voiture or "").strip(),
        "nombre_place": 0,
        "location_par_jour": 0,
        "consommation": 0,
        "energie": "",
    }


def _load_km_mada_rows():
    if not OPENPYXL_AVAILABLE:
        return []

    if not os.path.exists(HOTEL_EXCEL_PATH):
        _invalidate_km_mada_cache()
        return []
    if not zipfile.is_zipfile(HOTEL_EXCEL_PATH):
        # Avoid repeated openpyxl exceptions when the workbook is temporarily invalid/corrupted.
        _invalidate_km_mada_cache()
        return []

    mtime = os.path.getmtime(HOTEL_EXCEL_PATH)
    now = monotonic()
    if (
        _KM_MADA_CACHE["path"] == HOTEL_EXCEL_PATH
        and _KM_MADA_CACHE["mtime"] == mtime
        and (now - _KM_MADA_CACHE["loaded_at"]) <= _KM_MADA_CACHE_TTL_SECONDS
    ):
        return list(_KM_MADA_CACHE["rows"])

    wb = None
    try:
        wb = load_workbook(HOTEL_EXCEL_PATH, data_only=True)
        if KM_MADA_SHEET_NAME not in wb.sheetnames:
            _invalidate_km_mada_cache()
            return []

        ws = wb[KM_MADA_SHEET_NAME]

        def _resolve_columns(header_row):
            header_map = _get_header_map(ws, header_row)
            repere = _find_header_column(
                header_map,
                "REPERES",
                "Reperes",
                "Repères",
                "REPERE",
                "Repere",
            )
            km = _find_header_column(
                header_map,
                "KM",
                "KM TOTAL",
                "KM PARTIEL",
                "KMS",
                "KILOMETRAGE",
                "Kilometrage",
                "Kilométrage",
                "KILOMETRES",
                "Kilometres",
                "Kilomètres",
                "Distance",
            )
            duree = _find_header_column(
                header_map,
                "Durée",
                "Duree",
                "Durée trajet",
                "Duree trajet",
                "Temps",
                "Temps trajet",
            )
            return repere, km, duree

        repere_col, km_col, duree_col = _resolve_columns(1)
        data_start_row = 2
        if not repere_col:
            repere_col, km_col, duree_col = _resolve_columns(2)
            data_start_row = 3

        if not repere_col:
            return []

        rows = []
        for row in range(data_start_row, ws.max_row + 1):
            repere = ws.cell(row=row, column=repere_col).value
            km = ws.cell(row=row, column=km_col).value if km_col else 0
            duree = ws.cell(row=row, column=duree_col).value if duree_col else 0
            if repere in (None, ""):
                continue
            rows.append(
                {
                    "repere": str(repere).strip(),
                    "km": _parse_num(km),
                    "duree": _parse_duration_hours(duree),
                }
            )

        lookup = {}
        for row in rows:
            repere = str(row.get("repere") or "").strip().lower()
            if repere:
                lookup[repere] = row
        _KM_MADA_CACHE["path"] = HOTEL_EXCEL_PATH
        _KM_MADA_CACHE["mtime"] = mtime
        _KM_MADA_CACHE["loaded_at"] = now
        _KM_MADA_CACHE["rows"] = rows
        _KM_MADA_CACHE["lookup"] = lookup
        return rows
    except zipfile.BadZipFile as e:
        _invalidate_km_mada_cache()
        _log_error_throttled(
            "km_mada_bad_zip",
            f"Failed to load KM_MADA rows (invalid workbook): {e}",
            exc_info=True,
        )
        return []
    except PermissionError as e:
        _log_error_throttled(
            "km_mada_permission",
            f"Failed to load KM_MADA rows (permission error): {e}",
            exc_info=True,
        )
        return []
    except Exception as e:
        _log_error_throttled("km_mada_unexpected", f"Failed to load KM_MADA rows: {e}")
        return []
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def get_km_mada_reperes():
    rows = _load_km_mada_rows()
    values = {row.get("repere") for row in rows if row.get("repere")}
    return sorted(values)


def get_km_mada_km_for_repere(repere):
    lookup = str(repere or "").strip().lower()
    if not lookup:
        return 0

    rows = _load_km_mada_rows()
    cached_row = _KM_MADA_CACHE["lookup"].get(lookup)
    if cached_row:
        return _parse_num(cached_row.get("km", 0))
    for row in rows:
        if str(row.get("repere") or "").strip().lower() == lookup:
            return _parse_num(row.get("km", 0))
    return 0


def get_km_mada_duration_for_repere(repere):
    lookup = str(repere or "").strip().lower()
    if not lookup:
        return 0.0

    rows = _load_km_mada_rows()
    cached_row = _KM_MADA_CACHE["lookup"].get(lookup)
    if cached_row:
        return _parse_duration_hours(cached_row.get("duree", 0))
    for row in rows:
        if str(row.get("repere") or "").strip().lower() == lookup:
            return _parse_duration_hours(row.get("duree", 0))
    return 0.0


def get_parametrage_value_by_name(parameter_name):
    target = _normalize_header_key(parameter_name)
    if not target:
        return 0

    for row in load_all_parametrages():
        name = _normalize_header_key(row.get("PARAMETRE"))
        if name == target:
            return _parse_num(row.get("VALEUR", 0))
    return 0


def get_transport_fuel_price(energie):
    norm = _normalize_header_key(energie)
    if not norm:
        return 0

    if "essence" in norm:
        return get_parametrage_value_by_name("Prix Essence")
    if "gasoil" in norm or "diesel" in norm:
        return get_parametrage_value_by_name("Prix Gasoil")

    return get_parametrage_value_by_name(energie)


def get_transport_headers():
    """Load header list from TRANSPORT sheet in data.xlsx (header row = 2)."""
    if not OPENPYXL_AVAILABLE:
        return []

    if not os.path.exists(CLIENT_EXCEL_PATH):
        return []

    wb = None
    try:
        wb = load_workbook(CLIENT_EXCEL_PATH)
        if TRANSPORT_SHEET_NAME not in wb.sheetnames:
            return []

        ws = wb[TRANSPORT_SHEET_NAME]
        headers = []
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=2, column=col).value
            if value is None:
                continue
            label = str(value).strip()
            if label:
                headers.append(label)
        return headers
    except Exception as e:
        logger.error(f"Failed to load TRANSPORT headers: {e}", exc_info=True)
        return []
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def save_transport_quotation_to_excel(form_data):
    """
    Save transport quotation row into data.xlsx/TRANSPORT.

    - Header row: 2
    - Data starts at row: 3
    """
    if not OPENPYXL_AVAILABLE:
        return -1

    wb = None
    try:
        if not os.path.exists(CLIENT_EXCEL_PATH):
            wb = Workbook()
            ws = wb.active
            ws.title = TRANSPORT_SHEET_NAME
        else:
            wb = load_workbook(CLIENT_EXCEL_PATH)
            if TRANSPORT_SHEET_NAME not in wb.sheetnames:
                ws = wb.create_sheet(TRANSPORT_SHEET_NAME)
            else:
                ws = wb[TRANSPORT_SHEET_NAME]

        headers = []
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=2, column=col).value
            if value is None:
                continue
            label = str(value).strip()
            if label:
                headers.append(label)

        if not headers:
            headers = list(form_data.keys())
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=2, column=col)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    start_color="4472C4", end_color="4472C4", fill_type="solid"
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")

        header_map = {header: index for index, header in enumerate(headers, start=1)}

        first_col = 1
        next_row = 3
        while ws.cell(row=next_row, column=first_col).value not in (None, ""):
            next_row += 1

        for header, col in header_map.items():
            ws.cell(row=next_row, column=col, value=form_data.get(header, ""))

        wb.save(CLIENT_EXCEL_PATH)
        return next_row
    except PermissionError:
        return -2
    except Exception as e:
        logger.error(f"Failed to save transport quotation: {e}", exc_info=True)
        return -1
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def load_all_transport_quotations():
    """Load all transport quotation rows from data.xlsx/TRANSPORT (data starts at row 3)."""
    if not OPENPYXL_AVAILABLE:
        return []

    if not os.path.exists(CLIENT_EXCEL_PATH):
        return []

    wb = None
    try:
        wb = load_workbook(CLIENT_EXCEL_PATH)
        if TRANSPORT_SHEET_NAME not in wb.sheetnames:
            return []

        ws = wb[TRANSPORT_SHEET_NAME]
        headers = get_transport_headers()
        if not headers:
            return []

        rows = []
        for row_index in range(3, ws.max_row + 1):
            row_dict = {"row_number": row_index}
            has_values = False
            for col_index, header in enumerate(headers, start=1):
                value = ws.cell(row=row_index, column=col_index).value
                if value not in (None, ""):
                    has_values = True
                row_dict[header] = "" if value is None else value
            if has_values:
                rows.append(row_dict)

        return rows
    except Exception as e:
        logger.error(f"Failed to load transport quotations: {e}", exc_info=True)
        return []
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def update_transport_quotation_in_excel(row_number, form_data):
    """Update one transport quotation row in data.xlsx/TRANSPORT."""
    if not OPENPYXL_AVAILABLE:
        return -1

    if not os.path.exists(CLIENT_EXCEL_PATH):
        return -1

    wb = None
    try:
        wb = load_workbook(CLIENT_EXCEL_PATH)
        if TRANSPORT_SHEET_NAME not in wb.sheetnames:
            return -1

        ws = wb[TRANSPORT_SHEET_NAME]
        headers = get_transport_headers()

        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_number, column=col_idx, value=form_data.get(header, ""))

        wb.save(CLIENT_EXCEL_PATH)
        return 0
    except PermissionError:
        return -2
    except Exception as e:
        logger.error(f"Failed to update transport row {row_number}: {e}", exc_info=True)
        return -1
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def delete_transport_from_excel(row_number):
    """Delete one transport quotation row from data.xlsx/TRANSPORT."""
    if not OPENPYXL_AVAILABLE:
        return False

    if not os.path.exists(CLIENT_EXCEL_PATH):
        return False

    wb = None
    try:
        wb = load_workbook(CLIENT_EXCEL_PATH)
        if TRANSPORT_SHEET_NAME not in wb.sheetnames:
            return False

        ws = wb[TRANSPORT_SHEET_NAME]
        ws.delete_rows(row_number)
        wb.save(CLIENT_EXCEL_PATH)
        return True
    except Exception as e:
        logger.error(f"Failed to delete transport row {row_number}: {e}", exc_info=True)
        return False
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def get_transport_db_headers():
    """Load header list from data-hotel.xlsx / TRANSPORT (source DB)."""
    if not OPENPYXL_AVAILABLE:
        return []

    if not os.path.exists(HOTEL_EXCEL_PATH):
        return []

    wb = None
    try:
        wb = load_workbook(HOTEL_EXCEL_PATH)
        if TRANSPORT_SOURCE_SHEET_NAME not in wb.sheetnames:
            return []

        ws = wb[TRANSPORT_SOURCE_SHEET_NAME]
        header_map, _data_start = _resolve_transport_source_header_map(ws)
        if not header_map:
            return []
        return list(header_map.keys())
    except Exception as e:
        logger.error(f"Failed to load transport DB headers: {e}", exc_info=True)
        return []
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def load_transport_db_rows():
    """Load raw DB rows from data-hotel.xlsx / TRANSPORT (source DB)."""
    if not OPENPYXL_AVAILABLE:
        return []

    if not os.path.exists(HOTEL_EXCEL_PATH):
        return []

    wb = None
    try:
        wb = load_workbook(HOTEL_EXCEL_PATH)
        if TRANSPORT_SOURCE_SHEET_NAME not in wb.sheetnames:
            return []

        ws = wb[TRANSPORT_SOURCE_SHEET_NAME]
        header_map, data_start_row = _resolve_transport_source_header_map(ws)
        if not header_map:
            return []

        headers = list(header_map.keys())
        rows = []
        for row_idx in range(data_start_row, ws.max_row + 1):
            row_dict = {"row_number": row_idx}
            has_values = False
            for header in headers:
                col = header_map[header]
                value = ws.cell(row=row_idx, column=col).value
                if value not in (None, ""):
                    has_values = True
                row_dict[header] = "" if value is None else value
            if has_values:
                rows.append(row_dict)
        return rows
    except Exception as e:
        logger.error(f"Failed to load transport DB rows: {e}", exc_info=True)
        return []
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def save_transport_db_row(row_data):
    """Save one row into data-hotel.xlsx / TRANSPORT (source DB)."""
    if not OPENPYXL_AVAILABLE:
        return -1

    wb = None
    try:
        if not os.path.exists(HOTEL_EXCEL_PATH):
            wb = Workbook()
            ws = wb.active
            ws.title = TRANSPORT_SOURCE_SHEET_NAME
        else:
            wb = load_workbook(HOTEL_EXCEL_PATH)
            if TRANSPORT_SOURCE_SHEET_NAME not in wb.sheetnames:
                ws = wb.create_sheet(TRANSPORT_SOURCE_SHEET_NAME)
            else:
                ws = wb[TRANSPORT_SOURCE_SHEET_NAME]

        header_map, data_start_row = _resolve_transport_source_header_map(ws)
        if not header_map:
            headers = list(row_data.keys()) if row_data else []
            if not headers:
                return -1
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx, value=header)
            header_map = _get_header_map(ws, 1)
            data_start_row = 2

        headers = list(header_map.keys())
        next_row = data_start_row
        while True:
            has_data = False
            for header in headers:
                col = header_map[header]
                if ws.cell(row=next_row, column=col).value not in (None, ""):
                    has_data = True
                    break
            if not has_data:
                break
            next_row += 1

        for header in headers:
            ws.cell(row=next_row, column=header_map[header], value=row_data.get(header, ""))

        wb.save(HOTEL_EXCEL_PATH)
        _invalidate_km_mada_cache()
        return next_row
    except PermissionError:
        return -2
    except Exception as e:
        logger.error(f"Failed to save transport DB row: {e}", exc_info=True)
        return -1
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def update_transport_db_row(row_number, row_data):
    """Update one row in data-hotel.xlsx / TRANSPORT (source DB)."""
    if not OPENPYXL_AVAILABLE:
        return -1

    if not os.path.exists(HOTEL_EXCEL_PATH):
        return -1

    wb = None
    try:
        wb = load_workbook(HOTEL_EXCEL_PATH)
        if TRANSPORT_SOURCE_SHEET_NAME not in wb.sheetnames:
            return -1

        ws = wb[TRANSPORT_SOURCE_SHEET_NAME]
        header_map, _data_start_row = _resolve_transport_source_header_map(ws)
        if not header_map:
            return -1

        for header, col in header_map.items():
            ws.cell(row=row_number, column=col, value=row_data.get(header, ""))

        wb.save(HOTEL_EXCEL_PATH)
        _invalidate_km_mada_cache()
        return 0
    except PermissionError:
        return -2
    except Exception as e:
        logger.error(f"Failed to update transport DB row {row_number}: {e}", exc_info=True)
        return -1
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def delete_transport_db_row(row_number):
    """Delete one row from data-hotel.xlsx / TRANSPORT (source DB)."""
    if not OPENPYXL_AVAILABLE:
        return False

    if not os.path.exists(HOTEL_EXCEL_PATH):
        return False

    wb = None
    try:
        wb = load_workbook(HOTEL_EXCEL_PATH)
        if TRANSPORT_SOURCE_SHEET_NAME not in wb.sheetnames:
            return False

        ws = wb[TRANSPORT_SOURCE_SHEET_NAME]
        ws.delete_rows(row_number)
        wb.save(HOTEL_EXCEL_PATH)
        return True
    except Exception as e:
        logger.error(f"Failed to delete transport DB row {row_number}: {e}", exc_info=True)
        return False
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def _resolve_km_mada_header_map(ws):
    header_map = _get_header_map(ws, 1)
    repere_col = _find_header_column(
        header_map,
        "REPERES",
        "Reperes",
        "Repères",
        "REPERE",
        "Repere",
    )
    if repere_col:
        return header_map, 2

    header_map = _get_header_map(ws, 2)
    repere_col = _find_header_column(
        header_map,
        "REPERES",
        "Reperes",
        "Repères",
        "REPERE",
        "Repere",
    )
    if repere_col:
        return header_map, 3

    return {}, 2


def get_km_mada_db_headers():
    """Load header list from data-hotel.xlsx / KM_MADA."""
    if not OPENPYXL_AVAILABLE:
        return []

    if not os.path.exists(HOTEL_EXCEL_PATH):
        return []

    wb = None
    try:
        wb = load_workbook(HOTEL_EXCEL_PATH)
        if KM_MADA_SHEET_NAME not in wb.sheetnames:
            return []

        ws = wb[KM_MADA_SHEET_NAME]
        header_map, _data_start = _resolve_km_mada_header_map(ws)
        if not header_map:
            return []
        return list(header_map.keys())
    except (PermissionError, OSError, ValueError, zipfile.BadZipFile) as e:
        logger.error(f"Failed to load KM_MADA DB headers: {e}", exc_info=True)
        return []
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def load_km_mada_db_rows():
    """Load raw DB rows from data-hotel.xlsx / KM_MADA."""
    if not OPENPYXL_AVAILABLE:
        return []

    if not os.path.exists(HOTEL_EXCEL_PATH):
        return []

    wb = None
    try:
        wb = load_workbook(HOTEL_EXCEL_PATH)
        if KM_MADA_SHEET_NAME not in wb.sheetnames:
            return []

        ws = wb[KM_MADA_SHEET_NAME]
        header_map, data_start_row = _resolve_km_mada_header_map(ws)
        if not header_map:
            return []

        headers = list(header_map.keys())
        rows = []
        for row_idx in range(data_start_row, ws.max_row + 1):
            row_dict = {"row_number": row_idx}
            has_values = False
            for header in headers:
                col = header_map[header]
                value = ws.cell(row=row_idx, column=col).value
                if value not in (None, ""):
                    has_values = True
                row_dict[header] = "" if value is None else value
            if has_values:
                rows.append(row_dict)
        return rows
    except (PermissionError, OSError, ValueError, zipfile.BadZipFile) as e:
        logger.error(f"Failed to load KM_MADA DB rows: {e}", exc_info=True)
        return []
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def save_km_mada_db_row(row_data):
    """Save one row into data-hotel.xlsx / KM_MADA."""
    if not OPENPYXL_AVAILABLE:
        return -1

    wb = None
    try:
        if not os.path.exists(HOTEL_EXCEL_PATH):
            wb = Workbook()
            ws = wb.active
            ws.title = KM_MADA_SHEET_NAME
        else:
            wb = load_workbook(HOTEL_EXCEL_PATH)
            if KM_MADA_SHEET_NAME not in wb.sheetnames:
                ws = wb.create_sheet(KM_MADA_SHEET_NAME)
            else:
                ws = wb[KM_MADA_SHEET_NAME]

        header_map, data_start_row = _resolve_km_mada_header_map(ws)
        if not header_map:
            headers = list(row_data.keys()) if row_data else []
            if not headers:
                return -1
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx, value=header)
            header_map = _get_header_map(ws, 1)
            data_start_row = 2

        headers = list(header_map.keys())
        next_row = data_start_row
        while True:
            has_data = False
            for header in headers:
                col = header_map[header]
                if ws.cell(row=next_row, column=col).value not in (None, ""):
                    has_data = True
                    break
            if not has_data:
                break
            next_row += 1

        for header in headers:
            ws.cell(row=next_row, column=header_map[header], value=row_data.get(header, ""))

        wb.save(HOTEL_EXCEL_PATH)
        return next_row
    except PermissionError:
        return -2
    except (OSError, ValueError, KeyError) as e:
        logger.error(f"Failed to save KM_MADA DB row: {e}", exc_info=True)
        return -1
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def update_km_mada_db_row(row_number, row_data):
    """Update one row in data-hotel.xlsx / KM_MADA."""
    if not OPENPYXL_AVAILABLE:
        return -1

    if not os.path.exists(HOTEL_EXCEL_PATH):
        return -1

    wb = None
    try:
        wb = load_workbook(HOTEL_EXCEL_PATH)
        if KM_MADA_SHEET_NAME not in wb.sheetnames:
            return -1

        ws = wb[KM_MADA_SHEET_NAME]
        header_map, _data_start_row = _resolve_km_mada_header_map(ws)
        if not header_map:
            return -1

        for header, col in header_map.items():
            ws.cell(row=row_number, column=col, value=row_data.get(header, ""))

        wb.save(HOTEL_EXCEL_PATH)
        return 0
    except PermissionError:
        return -2
    except (OSError, ValueError, KeyError) as e:
        logger.error(f"Failed to update KM_MADA DB row {row_number}: {e}", exc_info=True)
        return -1
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def delete_km_mada_db_row(row_number):
    """Delete one row from data-hotel.xlsx / KM_MADA."""
    if not OPENPYXL_AVAILABLE:
        return False

    if not os.path.exists(HOTEL_EXCEL_PATH):
        return False

    wb = None
    try:
        wb = load_workbook(HOTEL_EXCEL_PATH)
        if KM_MADA_SHEET_NAME not in wb.sheetnames:
            return False

        ws = wb[KM_MADA_SHEET_NAME]
        ws.delete_rows(row_number)
        wb.save(HOTEL_EXCEL_PATH)
        _invalidate_km_mada_cache()
        return True
    except (OSError, ValueError, KeyError) as e:
        logger.error(f"Failed to delete KM_MADA DB row {row_number}: {e}", exc_info=True)
        return False
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


INVOICE_STATUS_UNPAID = "non payée"
INVOICE_STATUS_PARTIAL = "payée avec acompte"
INVOICE_STATUS_PAID = "payée"
INVOICE_STATUSES = (
    INVOICE_STATUS_UNPAID,
    INVOICE_STATUS_PARTIAL,
    INVOICE_STATUS_PAID,
)


INVOICE_HEADERS = [
    "Date",
    "ID_Facture",
    "Source_Type",
    "Source_Ref",
    "Client_ID",
    "Client_Nom",
    "Devise",
    "Montant_HT",
    "Cout_HT",
    "Marge_%", 
    "Marge_Montant",
    "Base_Taxable_HT",
    "TVA_%",
    "TVA_Montant",
    "Total_TTC",
    "Acompte",
    "Reste_A_Payer",
    "Statut",
]


FINANCIAL_STATE_HEADERS = [
    "Date_MAJ",
    "Nb_Factures",
    "CA_HT",
    "Marge_Totale",
    "TVA_Totale",
    "CA_TTC",
    "Acomptes_Recus",
    "Encaissements_Estimes",
    "Restes_A_Encaisser",
    "Nb_Payees",
    "Nb_Payees_Avec_Acompte",
    "Nb_Non_Payees",
]


def _safe_float(value):
    return float(_parse_num(value))


def _normalize_invoice_status(status, total_ttc=0.0, acompte=0.0):
    normalized = _normalize_header_key(status)
    if normalized == _normalize_header_key(INVOICE_STATUS_PAID):
        return INVOICE_STATUS_PAID
    if normalized == _normalize_header_key(INVOICE_STATUS_PARTIAL):
        return INVOICE_STATUS_PARTIAL
    if normalized == _normalize_header_key(INVOICE_STATUS_UNPAID):
        return INVOICE_STATUS_UNPAID

    if total_ttc <= 0:
        return INVOICE_STATUS_UNPAID
    if acompte >= total_ttc:
        return INVOICE_STATUS_PAID
    if acompte > 0:
        return INVOICE_STATUS_PARTIAL
    return INVOICE_STATUS_UNPAID


def calculate_invoice_totals(
    montant_ht,
    cout_ht=0,
    marge_pct=0,
    tva_pct=0,
    acompte=0,
    statut="",
):
    """Compute invoice totals with margin and VAT."""
    montant_ht = max(0.0, _safe_float(montant_ht))
    cout_ht = max(0.0, _safe_float(cout_ht))
    marge_pct = max(0.0, _safe_float(marge_pct))
    tva_pct = max(0.0, _safe_float(tva_pct))
    acompte = max(0.0, _safe_float(acompte))

    marge_montant = montant_ht * (marge_pct / 100.0)
    if marge_montant == 0 and cout_ht > 0:
        marge_montant = max(0.0, montant_ht - cout_ht)

    base_taxable_ht = montant_ht + marge_montant
    tva_montant = base_taxable_ht * (tva_pct / 100.0)
    total_ttc = base_taxable_ht + tva_montant

    if acompte > total_ttc:
        acompte = total_ttc
    reste = max(0.0, total_ttc - acompte)
    normalized_status = _normalize_invoice_status(statut, total_ttc=total_ttc, acompte=acompte)

    if normalized_status == INVOICE_STATUS_PAID:
        acompte = total_ttc
        reste = 0.0
    elif normalized_status == INVOICE_STATUS_UNPAID:
        acompte = 0.0
        reste = total_ttc

    return {
        "Montant_HT": montant_ht,
        "Cout_HT": cout_ht,
        "Marge_%": marge_pct,
        "Marge_Montant": marge_montant,
        "Base_Taxable_HT": base_taxable_ht,
        "TVA_%": tva_pct,
        "TVA_Montant": tva_montant,
        "Total_TTC": total_ttc,
        "Acompte": acompte,
        "Reste_A_Payer": reste,
        "Statut": normalized_status,
    }


def _ensure_invoice_sheet(wb):
    if INVOICE_SHEET_NAME not in wb.sheetnames:
        ws = wb.create_sheet(INVOICE_SHEET_NAME)
    else:
        ws = wb[INVOICE_SHEET_NAME]

    if ws.max_row == 1 and ws["A1"].value is None:
        for col, header in enumerate(INVOICE_HEADERS, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
    else:
        _ensure_headers(ws, INVOICE_HEADERS)

    return ws


def _ensure_financial_state_sheet(wb):
    if FINANCIAL_STATE_SHEET_NAME not in wb.sheetnames:
        ws = wb.create_sheet(FINANCIAL_STATE_SHEET_NAME)
    else:
        ws = wb[FINANCIAL_STATE_SHEET_NAME]

    if ws.max_row == 1 and ws["A1"].value is None:
        for col, header in enumerate(FINANCIAL_STATE_HEADERS, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="1F4E78", end_color="1F4E78", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
    else:
        _ensure_headers(ws, FINANCIAL_STATE_HEADERS)

    return ws


def _next_invoice_id(ws):
    max_num = 0
    for row_idx in range(2, ws.max_row + 1):
        value = ws.cell(row=row_idx, column=2).value
        if not value:
            continue
        match = re.search(r"(\d+)$", str(value).strip())
        if match:
            max_num = max(max_num, int(match.group(1)))
    return f"FAC-{datetime.now().strftime('%Y%m')}-{max_num + 1:04d}"


def save_invoice_to_excel(invoice_data):
    """Create an invoice and automatically refresh the financial state."""
    if not OPENPYXL_AVAILABLE:
        return -1

    wb = None
    try:
        if not os.path.exists(FINANCIAL_EXCEL_PATH):
            wb = Workbook()
            ws_default = wb.active
            ws_default.title = INVOICE_SHEET_NAME
        else:
            wb = load_workbook(FINANCIAL_EXCEL_PATH)

        ws = _ensure_invoice_sheet(wb)
        calculations = calculate_invoice_totals(
            montant_ht=invoice_data.get("Montant_HT", invoice_data.get("montant_ht", 0)),
            cout_ht=invoice_data.get("Cout_HT", invoice_data.get("cout_ht", 0)),
            marge_pct=invoice_data.get("Marge_%", invoice_data.get("marge_pct", 0)),
            tva_pct=invoice_data.get("TVA_%", invoice_data.get("tva_pct", 0)),
            acompte=invoice_data.get("Acompte", invoice_data.get("acompte", 0)),
            statut=invoice_data.get("Statut", invoice_data.get("statut", "")),
        )

        row = ws.max_row + 1
        invoice_id = str(invoice_data.get("ID_Facture") or "").strip()
        if not invoice_id:
            invoice_id = _next_invoice_id(ws)

        values = {
            "Date": invoice_data.get("Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            "ID_Facture": invoice_id,
            "Source_Type": invoice_data.get("Source_Type", invoice_data.get("source_type", "")),
            "Source_Ref": invoice_data.get("Source_Ref", invoice_data.get("source_ref", "")),
            "Client_ID": invoice_data.get("Client_ID", invoice_data.get("client_id", "")),
            "Client_Nom": invoice_data.get("Client_Nom", invoice_data.get("client_nom", "")),
            "Devise": invoice_data.get("Devise", invoice_data.get("devise", "Ariary")),
            **calculations,
        }

        for col, header in enumerate(INVOICE_HEADERS, start=1):
            ws.cell(row=row, column=col, value=values.get(header, ""))

        _rebuild_financial_state_in_workbook(wb)
        wb.save(FINANCIAL_EXCEL_PATH)
        return row
    except PermissionError:
        return -2
    except Exception as e:
        logger.error(f"Failed to save invoice: {e}", exc_info=True)
        return -1
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def load_all_invoices():
    """Load all invoices."""
    if not OPENPYXL_AVAILABLE:
        return []
    if not os.path.exists(FINANCIAL_EXCEL_PATH):
        return []

    wb = None
    try:
        wb = load_workbook(FINANCIAL_EXCEL_PATH)
        if INVOICE_SHEET_NAME not in wb.sheetnames:
            return []

        ws = wb[INVOICE_SHEET_NAME]
        header_map = _get_header_map(ws, 1)
        if not header_map:
            return []

        rows = []
        for row_idx in range(2, ws.max_row + 1):
            row_data = {"row_number": row_idx}
            has_values = False
            for header in INVOICE_HEADERS:
                col = header_map.get(header)
                value = ws.cell(row=row_idx, column=col).value if col else ""
                if value not in (None, ""):
                    has_values = True
                row_data[header] = "" if value is None else value
            if has_values:
                rows.append(row_data)
        return rows
    except Exception as e:
        logger.error(f"Failed to load invoices: {e}", exc_info=True)
        return []
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def update_invoice_in_excel(row_number, invoice_data):
    """Update one invoice row and refresh financial state."""
    if not OPENPYXL_AVAILABLE:
        return -1
    if not os.path.exists(FINANCIAL_EXCEL_PATH):
        return -1

    wb = None
    try:
        wb = load_workbook(FINANCIAL_EXCEL_PATH)
        if INVOICE_SHEET_NAME not in wb.sheetnames:
            return -1

        ws = wb[INVOICE_SHEET_NAME]
        header_map = _get_header_map(ws, 1)
        if not header_map:
            _ensure_invoice_sheet(wb)
            header_map = _get_header_map(ws, 1)

        current = {}
        for header in INVOICE_HEADERS:
            col = header_map.get(header)
            current[header] = ws.cell(row=row_number, column=col).value if col else ""

        merged = {**current, **invoice_data}
        calculations = calculate_invoice_totals(
            montant_ht=merged.get("Montant_HT", 0),
            cout_ht=merged.get("Cout_HT", 0),
            marge_pct=merged.get("Marge_%", 0),
            tva_pct=merged.get("TVA_%", 0),
            acompte=merged.get("Acompte", 0),
            statut=merged.get("Statut", ""),
        )
        merged.update(calculations)

        for header in INVOICE_HEADERS:
            col = header_map.get(header)
            if not col:
                continue
            ws.cell(row=row_number, column=col, value=merged.get(header, ""))

        _rebuild_financial_state_in_workbook(wb)
        wb.save(FINANCIAL_EXCEL_PATH)
        return 0
    except PermissionError:
        return -2
    except Exception as e:
        logger.error(f"Failed to update invoice row {row_number}: {e}", exc_info=True)
        return -1
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def _rebuild_financial_state_in_workbook(wb):
    """Recompute the one-line financial state from invoices."""
    ws_invoice = _ensure_invoice_sheet(wb)
    ws_state = _ensure_financial_state_sheet(wb)
    header_map = _get_header_map(ws_invoice, 1)

    totals = {
        "Nb_Factures": 0,
        "CA_HT": 0.0,
        "Marge_Totale": 0.0,
        "TVA_Totale": 0.0,
        "CA_TTC": 0.0,
        "Acomptes_Recus": 0.0,
        "Encaissements_Estimes": 0.0,
        "Restes_A_Encaisser": 0.0,
        "Nb_Payees": 0,
        "Nb_Payees_Avec_Acompte": 0,
        "Nb_Non_Payees": 0,
    }

    def _invoice_cell(row_idx, header):
        col = header_map.get(header)
        if not col:
            return None
        return ws_invoice.cell(row=row_idx, column=col).value

    for row_idx in range(2, ws_invoice.max_row + 1):
        invoice_id_col = header_map.get("ID_Facture")
        if not invoice_id_col:
            continue
        invoice_id = ws_invoice.cell(row=row_idx, column=invoice_id_col).value
        if invoice_id in (None, ""):
            continue

        totals["Nb_Factures"] += 1
        totals["CA_HT"] += _safe_float(_invoice_cell(row_idx, "Montant_HT"))
        totals["Marge_Totale"] += _safe_float(_invoice_cell(row_idx, "Marge_Montant"))
        totals["TVA_Totale"] += _safe_float(_invoice_cell(row_idx, "TVA_Montant"))
        total_ttc = _safe_float(_invoice_cell(row_idx, "Total_TTC"))
        acompte = _safe_float(_invoice_cell(row_idx, "Acompte"))
        reste = _safe_float(_invoice_cell(row_idx, "Reste_A_Payer"))
        status = _normalize_invoice_status(
            _invoice_cell(row_idx, "Statut"),
            total_ttc=total_ttc,
            acompte=acompte,
        )

        totals["CA_TTC"] += total_ttc
        totals["Acomptes_Recus"] += acompte
        totals["Restes_A_Encaisser"] += reste
        totals["Encaissements_Estimes"] += max(acompte, total_ttc if status == INVOICE_STATUS_PAID else acompte)

        if status == INVOICE_STATUS_PAID:
            totals["Nb_Payees"] += 1
        elif status == INVOICE_STATUS_PARTIAL:
            totals["Nb_Payees_Avec_Acompte"] += 1
        else:
            totals["Nb_Non_Payees"] += 1

    header_map_state = _get_header_map(ws_state, 1)
    # Keep only most recent snapshot line.
    if ws_state.max_row > 1:
        ws_state.delete_rows(2, ws_state.max_row - 1)

    row = 2
    data = {
        "Date_MAJ": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        **totals,
    }
    for header in FINANCIAL_STATE_HEADERS:
        col = header_map_state.get(header)
        if col:
            ws_state.cell(row=row, column=col, value=data.get(header, ""))


def refresh_financial_state_from_invoices():
    """Public helper to force financial state rebuild from invoices."""
    if not OPENPYXL_AVAILABLE:
        return -1

    wb = None
    try:
        if not os.path.exists(FINANCIAL_EXCEL_PATH):
            wb = Workbook()
            wb.active.title = INVOICE_SHEET_NAME
        else:
            wb = load_workbook(FINANCIAL_EXCEL_PATH)

        _rebuild_financial_state_in_workbook(wb)
        wb.save(FINANCIAL_EXCEL_PATH)
        return 0
    except PermissionError:
        return -2
    except Exception as e:
        logger.error(f"Failed to refresh financial state: {e}", exc_info=True)
        return -1
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def load_financial_state_snapshot():
    """Load latest financial-state row."""
    if not OPENPYXL_AVAILABLE:
        return {}
    if not os.path.exists(FINANCIAL_EXCEL_PATH):
        return {}

    wb = None
    try:
        wb = load_workbook(FINANCIAL_EXCEL_PATH)
        if FINANCIAL_STATE_SHEET_NAME not in wb.sheetnames:
            return {}
        ws = wb[FINANCIAL_STATE_SHEET_NAME]
        header_map = _get_header_map(ws, 1)
        if not header_map or ws.max_row < 2:
            return {}
        result = {}
        for header in FINANCIAL_STATE_HEADERS:
            col = header_map.get(header)
            result[header] = ws.cell(row=2, column=col).value if col else ""
        return result
    except Exception as e:
        logger.error(f"Failed to load financial state snapshot: {e}", exc_info=True)
        return {}
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass
