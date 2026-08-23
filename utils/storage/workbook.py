"""Ouverture, sauvegarde et fermeture des classeurs Excel.

Avant ce module, 78 fonctions de utils.excel_handler repetaient le meme
prologue et le meme epilogue : verification d'openpyxl, verification du
fichier, verification de la feuille, `wb = None`, `try`, `wb.save(...)`,
`except PermissionError`, `except Exception` + log, puis un `finally` qui
fermait le classeur en avalant l'erreur. `open_workbook` porte tout cela une
fois.

Le module expose aussi une hierarchie d'exceptions. Les fonctions publiques
historiques, elles, doivent continuer a renvoyer leurs sentinelles (-1, -2,
False, []) parce que l'interface les teste directement : `sentinel_on_error`
fait cette traduction, au seul endroit ou elle a lieu.
"""

import os
import shutil
import socket
import tempfile
import threading
import time
from contextlib import contextmanager
from datetime import datetime
from functools import wraps

try:
    from openpyxl import Workbook, load_workbook

    OPENPYXL_AVAILABLE = True
except ImportError:  # pragma: no cover - depend de l'environnement
    OPENPYXL_AVAILABLE = False

from utils.logger import logger


class StorageError(Exception):
    """Base des erreurs d'acces au stockage Excel."""


class StorageUnavailable(StorageError):
    """openpyxl n'est pas installe."""


class WorkbookMissing(StorageError):
    """Le fichier classeur n'existe pas."""


class SheetMissing(StorageError):
    """La feuille demandee n'existe pas dans le classeur."""


class WorkbookLocked(StorageError):
    """Le fichier est verrouille, typiquement ouvert dans Excel."""


# Plafond de copies conservees par classeur. La sauvegarde etant desormais
# systematique avant chaque ecriture, sans plafond le dossier `backups`
# grossirait indefiniment.
MAX_BACKUPS_PER_FILE = 20


def _prune_backups(backup_dir, filename):
    """Ne garde que les MAX_BACKUPS_PER_FILE copies les plus recentes.

    Le tri porte sur le nom, qui contient un horodatage `%Y%m%d_%H%M%S` :
    l'ordre lexicographique est donc l'ordre chronologique. Seules les copies
    du classeur `filename` sont concernees, celles des autres classeurs du
    meme dossier sont laissees intactes.
    """
    prefixe = f"{filename}."
    try:
        copies = sorted(
            nom
            for nom in os.listdir(backup_dir)
            if nom.startswith(prefixe) and nom.endswith(".bak")
        )
    except OSError:
        return

    for obsolete in copies[:-MAX_BACKUPS_PER_FILE] if MAX_BACKUPS_PER_FILE else copies:
        try:
            os.remove(os.path.join(backup_dir, obsolete))
        except OSError as exc:  # pragma: no cover - depend du systeme de fichiers
            logger.warning(f"Backup obsolete non supprime ({obsolete}) : {exc}")


def create_backup(filepath, _horodatage=None):
    """Copie horodatee du classeur dans un sous-dossier `backups`.

    Args:
        filepath (str): chemin du classeur a sauvegarder.
        _horodatage (str | None): horodatage impose, pour les tests. Par
            defaut, l'heure courante au format `%Y%m%d_%H%M%S`.

    Returns:
        str | None: chemin de la copie, ou None en cas d'echec.
    """
    if not os.path.exists(filepath):
        logger.warning(f"File not found for backup: {filepath}")
        return None

    try:
        backup_dir = os.path.join(os.path.dirname(filepath), "backups")
        os.makedirs(backup_dir, exist_ok=True)

        # Microsecondes indispensables : a la seconde pres, deux ecritures
        # rapprochees -- un enregistrement qui touche deux feuilles, une saisie
        # en rafale -- produisaient le meme nom et la seconde copie ecrasait la
        # premiere. Le plafond de retention devenait alors illusoire.
        timestamp = _horodatage or datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        filename = os.path.basename(filepath)
        backup_path = os.path.join(backup_dir, f"{filename}.{timestamp}.bak")

        shutil.copy2(filepath, backup_path)
        logger.info(f"Backup created: {backup_path}")

        _prune_backups(backup_dir, filename)

        return backup_path
    except Exception as e:
        logger.error(f"Failed to create backup for {filepath}: {e}", exc_info=True)
        return None


def _save_atomic(wb, path):
    """Ecrit le classeur sans jamais laisser `path` dans un etat partiel.

    openpyxl reserialise l'archive .xlsx entiere a chaque `save`. Ecrire
    directement sur la destination signifie qu'une coupure en cours de route
    (disque plein, arret brutal, reseau coupe sur un partage) laisse un
    fichier tronque et illisible. On ecrit donc dans un temporaire du meme
    dossier -- donc du meme systeme de fichiers -- puis on bascule par
    `os.replace`, qui est atomique.
    """
    dossier = os.path.dirname(path) or "."
    os.makedirs(dossier, exist_ok=True)

    fd, temporaire = tempfile.mkstemp(
        dir=dossier, prefix=f".{os.path.basename(path)}.", suffix=".tmp"
    )
    os.close(fd)
    try:
        wb.save(temporaire)
        os.replace(temporaire, path)
    except BaseException:
        try:
            os.remove(temporaire)
        except OSError:  # pragma: no cover - depend du systeme de fichiers
            pass
        raise


# ── Verrou d'ecriture entre processus ─────────────────────────────────────────

# Attente maximale avant d'abandonner. Une ecriture prend ~1,1 s sur le plus
# gros classeur ; 15 s laissent passer une dizaine d'ecritures en file sans
# faire patienter l'utilisateur indefiniment.
DELAI_VERROU_S = 15.0

# Au-dela, un verrou est considere abandonne. Une application tuee ou un poste
# eteint laisse son fichier derriere lui ; sans peremption, le classeur serait
# bloque jusqu'a une suppression manuelle.
PEREMPTION_VERROU_S = 120.0

_INTERVALLE_TENTATIVE_S = 0.05

# Verrous deja detenus par ce fil, par chemin absolu. Le verrou doit etre
# reentrant : `save_client_to_excel` appelle `_save_client_infos_to_excel`, et
# les deux sont protegees. Sans cela l'appel interne attendrait le verrou que
# l'appel externe detient deja, jusqu'a expiration du delai -- l'ecriture
# imbriquee etait alors abandonnee en silence.
_detenus = threading.local()


def _deja_detenu(chemin_verrou):
    return chemin_verrou in getattr(_detenus, "chemins", ())


def _chemin_du_verrou(path):
    return path + ".lock"


def _verrou_perime(chemin_verrou, peremption_s):
    try:
        age = time.time() - os.path.getmtime(chemin_verrou)
    except OSError:
        return False  # disparu entre-temps : la prochaine tentative tranchera
    return age > peremption_s


@contextmanager
def verrou_de_classeur(path, delai_max_s=None, peremption_s=None):
    """Verrou exclusif entre processus sur un classeur, le temps du bloc.

    Un fichier voisin `<classeur>.lock` cree en O_CREAT|O_EXCL, et non un
    `fcntl.flock` : flock n'est pas fiable sur les partages reseau SMB et NFS,
    qui sont precisement le cas d'usage multi-poste vise. La creation exclusive
    d'un fichier, elle, est atomique sur ces systemes.

    Le verrou couvre tout le cycle lecture-modification-ecriture, pas seulement
    la sauvegarde : c'est pendant la modification en memoire que le second
    poste lisait une version qu'il allait ensuite ecraser.

    Args:
        path (str): chemin du classeur a proteger.
        delai_max_s (float | None): attente avant d'abandonner.
        peremption_s (float | None): age au-dela duquel un verrou est repris.

    Raises:
        WorkbookLocked: le verrou n'a pas pu etre pris dans le delai.
    """
    delai_max_s = DELAI_VERROU_S if delai_max_s is None else delai_max_s
    peremption_s = PEREMPTION_VERROU_S if peremption_s is None else peremption_s

    chemin_verrou = os.path.abspath(_chemin_du_verrou(path))

    if _deja_detenu(chemin_verrou):
        # Prise imbriquee dans le meme fil : le verrou est deja a nous. On cede
        # sans rien creer, et surtout sans liberer a la sortie de ce bloc-ci.
        yield True
        return

    signature = (
        f"machine={socket.gethostname()} pid={os.getpid()} "
        f"depuis={datetime.now().isoformat(timespec='seconds')}"
    )

    echeance = time.monotonic() + delai_max_s
    descripteur = None
    while True:
        try:
            descripteur = os.open(chemin_verrou, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
            break
        except FileExistsError:
            if _verrou_perime(chemin_verrou, peremption_s):
                logger.warning(
                    f"Verrou perime repris sur {os.path.basename(path)} "
                    "(application interrompue ?)"
                )
                try:
                    os.unlink(chemin_verrou)
                except OSError:
                    pass
                continue
            if time.monotonic() >= echeance:
                detenteur = ""
                try:
                    with open(chemin_verrou, encoding="utf-8") as f:
                        detenteur = f.read().strip()
                except OSError:
                    pass
                raise WorkbookLocked(
                    f"{path} est en cours de modification ailleurs. {detenteur}"
                )
            time.sleep(_INTERVALLE_TENTATIVE_S)
        except OSError as exc:
            # Dossier en lecture seule, disque plein : ne pas bloquer
            # l'enregistrement pour autant. Sans verrou, on retombe sur le
            # comportement d'avant, qui reste preferable a un refus d'ecrire.
            logger.warning(f"Verrou impossible sur {path} ({exc}) : ecriture sans.")
            yield False
            return

    try:
        try:
            os.write(descripteur, signature.encode("utf-8"))
        except OSError:
            pass
        os.close(descripteur)
        descripteur = None
        if not hasattr(_detenus, "chemins"):
            _detenus.chemins = set()
        _detenus.chemins.add(chemin_verrou)
        yield True
    finally:
        getattr(_detenus, "chemins", set()).discard(chemin_verrou)
        if descripteur is not None:
            try:
                os.close(descripteur)
            except OSError:
                pass
        try:
            os.unlink(chemin_verrou)
        except OSError:
            pass


@contextmanager
def open_workbook(
    path,
    sheet=None,
    *,
    create=False,
    write=False,
    backup=None,
    data_only=False,
    delai_max_s=None,
):
    """Ouvre un classeur et cede la feuille demandee.

    Args:
        path (str): chemin du classeur.
        sheet (str | None): feuille a ceder. Si None, cede le classeur entier.
        create (bool): cree le fichier et/ou la feuille s'ils manquent.
        write (bool): sauvegarde le classeur a la sortie du bloc, de facon
            atomique (temporaire puis bascule).
        backup (bool | None): copie de securite avant sauvegarde. Par defaut
            (None), active des que `write` est vrai : la protection ne doit pas
            dependre du site d'appel. Sans effet si le fichier n'existe pas
            encore, ou en lecture seule.
        data_only (bool): lit les valeurs calculees plutot que les formules.
        delai_max_s (float | None): attente maximale du verrou d'ecriture.
            Sans effet en lecture, qui ne verrouille pas.

    Yields:
        La feuille demandee, ou le classeur si `sheet` vaut None.

    Raises:
        StorageUnavailable: openpyxl absent.
        WorkbookMissing: fichier absent et `create` faux.
        SheetMissing: feuille absente et `create` faux.
        WorkbookLocked: fichier verrouille par une autre application, ou
            modifie au meme moment depuis un autre poste.

    Note:
        En ecriture, un verrou entre processus couvre tout le bloc : le cycle
        lecture-modification-ecriture doit etre indivisible, sans quoi deux
        postes travaillant sur un dossier partage s'ecrasent mutuellement. Les
        lectures ne verrouillent pas et restent donc paralleles.
    """
    if not OPENPYXL_AVAILABLE:
        raise StorageUnavailable(f"openpyxl indisponible, acces impossible a {path}")

    if backup is None:
        backup = write

    exists = os.path.exists(path)
    if not exists and not create:
        raise WorkbookMissing(path)

    if write:
        with verrou_de_classeur(path, delai_max_s=delai_max_s):
            yield from _ouvrir(path, sheet, create, write, backup, data_only, exists)
    else:
        yield from _ouvrir(path, sheet, create, write, backup, data_only, exists)


def _ouvrir(path, sheet, create, write, backup, data_only, exists):
    """Cycle d'ouverture proprement dit, verrou deja pris s'il y a lieu."""
    wb = None
    try:
        if exists:
            try:
                wb = load_workbook(path, data_only=data_only)
            except PermissionError as exc:
                raise WorkbookLocked(path) from exc
        else:
            wb = Workbook()
            if sheet is not None:
                wb.active.title = sheet

        if sheet is None:
            target = wb
        elif sheet in wb.sheetnames:
            target = wb[sheet]
        elif create:
            target = wb.create_sheet(sheet)
        else:
            raise SheetMissing(f"{sheet} absente de {path}")

        yield target

        if write:
            if backup and exists:
                create_backup(path)
            try:
                _save_atomic(wb, path)
            except PermissionError as exc:
                raise WorkbookLocked(path) from exc
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def sentinel_on_error(missing, locked=None, *, label=""):
    """Traduit les exceptions du stockage vers les sentinelles historiques.

    L'interface teste directement les valeurs de retour (-1, -2, False, []),
    donc la signature des fonctions publiques ne peut pas changer. Ce decorateur
    concentre la conversion au lieu de la repeter dans chaque fonction.

    Reproduit fidelement le comportement d'origine : une cause attendue
    (openpyxl absent, fichier ou feuille manquants) est silencieuse, alors
    qu'une erreur inattendue est journalisee avec sa trace.

    Args:
        missing: valeur renvoyee pour une cause attendue ou une erreur.
        locked: valeur renvoyee si le fichier est verrouille. Par defaut,
            identique a `missing`.
        label (str): libelle utilise dans le message de log.
    """
    locked_value = missing if locked is None else locked

    def decorate(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            try:
                return func(*args, **kwargs)
            except WorkbookLocked:
                return locked_value
            except (StorageUnavailable, WorkbookMissing, SheetMissing):
                return missing
            except Exception as exc:
                logger.error(
                    f"{label or func.__name__} a echoue : {exc}", exc_info=True
                )
                return missing

        return wrapper

    return decorate
