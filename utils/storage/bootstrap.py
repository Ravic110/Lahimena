"""Amorcage des classeurs de donnees.

Les fichiers `data.xlsx` et `data-hotel.xlsx` ne sont pas versionnes : ils
contiennent les donnees vivantes de l'agence. Sur un poste neuf, ils sont donc
absents -- et l'application demarre sans rien dire, chaque lecture renvoyant
une liste vide. Rien ne distingue alors "aucun client enregistre" de "le
fichier de donnees n'existe pas".

Ce module fournit les deux moities de la reponse :

* `export_templates()` fige la *structure* des classeurs de production --
  noms de feuilles, lignes d'en-tete, largeurs de colonnes -- dans des
  gabarits versionnables, sans aucune donnee metier ;
* `ensure_workbooks()` installe ces gabarits quand les classeurs manquent.

Pourquoi un gabarit plutot qu'une liste d'en-tetes en dur : les feuilles de
cotation (`COTATION_H`, `COTATION_TRANSPORT`, `COTATION_REST`,
`COTATION_FRAIS_COL`, `AVION`) n'ont aucun en-tete declare dans le code. Il est
lu depuis le classeur existant, quel qu'il soit. Les reecrire de memoire
reviendrait a inventer une structure qui divergerait de celle en production.

Ce que l'amorcage ne peut pas faire : recreer les donnees de reference
(catalogue hotels, distances KM_MADA, tarifs transport et avion, circuits).
Ce sont des donnees metier, pas de la structure. `describe_state()` les
signale explicitement plutot que de les laisser passer pour du vide normal.
"""

import os
import shutil
from copy import copy

import config
from utils.logger import logger
from utils.storage.workbook import OPENPYXL_AVAILABLE, open_workbook

if OPENPYXL_AVAILABLE:  # pragma: no cover - depend de l'environnement
    from openpyxl import Workbook, load_workbook

TEMPLATES_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
    "templates",
)

# Feuilles dont la ligne 1 porte un bandeau de regroupement et non les
# en-tetes : ceux-ci sont en ligne 2. Le reste des feuilles n'a qu'une ligne
# d'en-tete. Se tromper ici ferait fuiter une ligne de donnees dans le gabarit.
LIGNES_D_ENTETE = {
    ("data.xlsx", "TRANSPORT"): 2,
    ("data-hotel.xlsx", "BDD_HOTEL"): 2,
}

DEFAUT_LIGNES_D_ENTETE = 1

# Feuilles porteuses de donnees de reference. Vides, l'application se lance
# mais ne sait plus chiffrer : ce n'est pas un etat de depart normal.
FEUILLES_DE_REFERENCE = {
    "data-hotel.xlsx": (
        "BDD_HOTEL",
        "Circuits",
        "KM_MADA",
        "TRANSPORT",
        "PARAMETRE",
    ),
}


def _classeurs():
    """(chemin vivant, chemin du gabarit) pour les deux classeurs."""
    return [
        (config.CLIENT_EXCEL_PATH, _chemin_gabarit(config.CLIENT_EXCEL_PATH)),
        (config.HOTEL_EXCEL_PATH, _chemin_gabarit(config.HOTEL_EXCEL_PATH)),
    ]


def _chemin_gabarit(chemin_vivant):
    nom = os.path.basename(chemin_vivant)
    base, _ = os.path.splitext(nom)
    return os.path.join(TEMPLATES_DIR, f"{base}.template.xlsx")


def _lignes_d_entete(nom_fichier, nom_feuille):
    return LIGNES_D_ENTETE.get((nom_fichier, nom_feuille), DEFAUT_LIGNES_D_ENTETE)


def export_templates(chemins=None):
    """Fige la structure des classeurs vivants dans `templates/`.

    Ne copie que les lignes d'en-tete, les fusions qui les concernent et les
    largeurs de colonnes. Aucune ligne de donnees ne traverse.

    Returns:
        list[str]: gabarits ecrits.
    """
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl est requis pour exporter les gabarits")

    os.makedirs(TEMPLATES_DIR, exist_ok=True)
    ecrits = []

    for vivant, gabarit in chemins or _classeurs():
        if not os.path.exists(vivant):
            logger.warning(f"Classeur absent, gabarit ignore : {vivant}")
            continue

        nom_fichier = os.path.basename(vivant)
        source = load_workbook(vivant)
        cible = Workbook()
        cible.remove(cible.active)
        try:
            for nom_feuille in source.sheetnames:
                origine = source[nom_feuille]
                copie = cible.create_sheet(nom_feuille)
                n = _lignes_d_entete(nom_fichier, nom_feuille)

                for ligne in range(1, n + 1):
                    for col in range(1, origine.max_column + 1):
                        cellule = origine.cell(row=ligne, column=col)
                        neuve = copie.cell(row=ligne, column=col, value=cellule.value)
                        if cellule.has_style:
                            # `_style` est un index dans la table du classeur
                            # source : il ne veut rien dire dans la cible. On
                            # recopie donc les objets de style un a un.
                            neuve.font = copy(cellule.font)
                            neuve.fill = copy(cellule.fill)
                            neuve.border = copy(cellule.border)
                            neuve.alignment = copy(cellule.alignment)
                            neuve.number_format = cellule.number_format

                for plage in origine.merged_cells.ranges:
                    if plage.max_row <= n:
                        copie.merge_cells(str(plage))

                for lettre, dim in origine.column_dimensions.items():
                    if dim.width:
                        copie.column_dimensions[lettre].width = dim.width

            cible.save(gabarit)
            ecrits.append(gabarit)
            logger.info(f"Gabarit ecrit : {gabarit}")
        finally:
            source.close()
            cible.close()

    return ecrits


def missing_workbooks():
    """Classeurs attendus mais absents du disque."""
    return [vivant for vivant, _ in _classeurs() if not os.path.exists(vivant)]


def ensure_workbooks():
    """Installe les classeurs manquants depuis les gabarits versionnes.

    Ne touche jamais a un classeur existant : l'amorcage cree, il ne repare
    pas et n'ecrase pas.

    Returns:
        list[str]: classeurs installes lors de cet appel.
    """
    installes = []
    for vivant, gabarit in _classeurs():
        if os.path.exists(vivant):
            continue
        if not os.path.exists(gabarit):
            logger.error(
                f"Classeur {vivant} absent et aucun gabarit dans {gabarit}. "
                "Regenerer avec : python scripts/init_data.py export-template"
            )
            continue
        os.makedirs(os.path.dirname(vivant) or ".", exist_ok=True)
        shutil.copy2(gabarit, vivant)
        installes.append(vivant)
        logger.info(f"Classeur initialise depuis le gabarit : {vivant}")
    return installes


def describe_state():
    """Etat des classeurs, feuille par feuille.

    Returns:
        dict: {nom_fichier: {"present": bool, "chemin": str,
                             "feuilles": {nom: nb_lignes_de_donnees},
                             "references_vides": [noms]}}
    """
    etat = {}
    for vivant, _ in _classeurs():
        nom_fichier = os.path.basename(vivant)
        entree = {"chemin": vivant, "present": os.path.exists(vivant)}
        entree["feuilles"] = {}
        entree["references_vides"] = []

        if entree["present"] and OPENPYXL_AVAILABLE:
            with open_workbook(vivant) as wb:
                for nom_feuille in wb.sheetnames:
                    ws = wb[nom_feuille]
                    n = _lignes_d_entete(nom_fichier, nom_feuille)
                    lignes = max(0, ws.max_row - n)
                    entree["feuilles"][nom_feuille] = lignes
                    attendues = FEUILLES_DE_REFERENCE.get(nom_fichier, ())
                    if nom_feuille in attendues and lignes == 0:
                        entree["references_vides"].append(nom_feuille)

        etat[nom_fichier] = entree
    return etat
