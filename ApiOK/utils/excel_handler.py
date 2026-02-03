"""
Excel file handling utilities
"""

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

import os
from config import DATA_EXCEL_PATH, CLIENT_SHEET_NAME


def save_client_to_excel(client_data):
    """
    Save client data to Excel file

    Args:
        client_data (dict): Client data dictionary

    Returns:
        int: Row number where data was saved, or -1 if failed
    """
    if not OPENPYXL_AVAILABLE:
        print("Warning: openpyxl not available. Cannot save to Excel.")
        return -1

    # Create file if it doesn't exist
    if not os.path.exists(DATA_EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = CLIENT_SHEET_NAME

        # Headers (A1:M1)
        headers = ['Date', 'Réf. Client', 'Nom', 'Téléphone', 'Email', 'Période',
                  'Restauration', 'Hébergement', 'Chambre', 'Enfant', 'Âge Enfant',
                  'Forfait', 'Circuit']

        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=i, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        wb.save(DATA_EXCEL_PATH)

    # Open existing file
    wb = load_workbook(DATA_EXCEL_PATH)
    if CLIENT_SHEET_NAME not in wb.sheetnames:
        ws = wb.create_sheet(CLIENT_SHEET_NAME)
    else:
        ws = wb[CLIENT_SHEET_NAME]

    # Find last empty row (column A)
    last_row = 2
    while ws[f'A{last_row}'].value is not None:
        last_row += 1

    # Write data
    ws[f'A{last_row}'] = client_data.get('Timestamp', '')
    ws[f'B{last_row}'] = client_data.get('Ref_Client', '')
    ws[f'C{last_row}'] = client_data.get('Nom', '')
    ws[f'D{last_row}'] = client_data.get('Téléphone', '')
    ws[f'E{last_row}'] = client_data.get('Email', '')
    ws[f'F{last_row}'] = client_data.get('Période', '')
    ws[f'G{last_row}'] = client_data.get('Restauration', '')
    ws[f'H{last_row}'] = client_data.get('Hébergement', '')
    ws[f'I{last_row}'] = client_data.get('Chambre', '')
    ws[f'J{last_row}'] = client_data.get('Enfant', '')
    ws[f'K{last_row}'] = client_data.get('Âge_Enfant', '')
    ws[f'L{last_row}'] = client_data.get('Forfait', '')
    ws[f'M{last_row}'] = client_data.get('Circuit', '')

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 25)

    wb.save(DATA_EXCEL_PATH)
    return last_row


def load_all_clients():
    """
    Load all client data from Excel file

    Returns:
        list: List of client dictionaries
    """
    if not OPENPYXL_AVAILABLE:
        print("Warning: openpyxl not available. Cannot load from Excel.")
        return []

    if not os.path.exists(DATA_EXCEL_PATH):
        return []

    wb = load_workbook(DATA_EXCEL_PATH)
    if CLIENT_SHEET_NAME not in wb.sheetnames:
        return []

    ws = wb[CLIENT_SHEET_NAME]

    clients = []
    # Start from row 2 (skip headers)
    for row in range(2, ws.max_row + 1):
        if ws[f'A{row}'].value is None:
            continue

        client = {
            'row_number': row,
            'timestamp': ws[f'A{row}'].value or '',
            'ref_client': ws[f'B{row}'].value or '',
            'nom': ws[f'C{row}'].value or '',
            'telephone': ws[f'D{row}'].value or '',
            'email': ws[f'E{row}'].value or '',
            'periode': ws[f'F{row}'].value or '',
            'restauration': ws[f'G{row}'].value or '',
            'hebergement': ws[f'H{row}'].value or '',
            'chambre': ws[f'I{row}'].value or '',
            'enfant': ws[f'J{row}'].value or '',
            'age_enfant': ws[f'K{row}'].value or '',
            'forfait': ws[f'L{row}'].value or '',
            'circuit': ws[f'M{row}'].value or ''
        }
        clients.append(client)

    return clients


def update_client_in_excel(row_number, client_data):
    """
    Update client data in Excel file

    Args:
        row_number (int): Row number to update
        client_data (dict): Updated client data dictionary

    Returns:
        bool: True if successful, False otherwise
    """
    if not OPENPYXL_AVAILABLE:
        print("Warning: openpyxl not available. Cannot update Excel.")
        return False

    if not os.path.exists(DATA_EXCEL_PATH):
        return False

    wb = load_workbook(DATA_EXCEL_PATH)
    if CLIENT_SHEET_NAME not in wb.sheetnames:
        return False

    ws = wb[CLIENT_SHEET_NAME]

    # Update data
    ws[f'A{row_number}'] = client_data.get('Timestamp', '')
    ws[f'B{row_number}'] = client_data.get('Ref_Client', '')
    ws[f'C{row_number}'] = client_data.get('Nom', '')
    ws[f'D{row_number}'] = client_data.get('Téléphone', '')
    ws[f'E{row_number}'] = client_data.get('Email', '')
    ws[f'F{row_number}'] = client_data.get('Période', '')
    ws[f'G{row_number}'] = client_data.get('Restauration', '')
    ws[f'H{row_number}'] = client_data.get('Hébergement', '')
    ws[f'I{row_number}'] = client_data.get('Chambre', '')
    ws[f'J{row_number}'] = client_data.get('Enfant', '')
    ws[f'K{row_number}'] = client_data.get('Âge_Enfant', '')
    ws[f'L{row_number}'] = client_data.get('Forfait', '')
    ws[f'M{row_number}'] = client_data.get('Circuit', '')

    wb.save(DATA_EXCEL_PATH)
    return True


def delete_client_from_excel(row_number):
    """
    Delete client data from Excel file

    Args:
        row_number (int): Row number to delete

    Returns:
        bool: True if successful, False otherwise
    """
    if not OPENPYXL_AVAILABLE:
        print("Warning: openpyxl not available. Cannot delete from Excel.")
        return False

    if not os.path.exists(DATA_EXCEL_PATH):
        return False

    wb = load_workbook(DATA_EXCEL_PATH)
    if CLIENT_SHEET_NAME not in wb.sheetnames:
        return False

    ws = wb[CLIENT_SHEET_NAME]

    # Delete row
    ws.delete_rows(row_number)

    wb.save(DATA_EXCEL_PATH)
    return True