"""Tests des primitives de feuille (utils.storage.sheet).

Ces fonctions etaient auparavant noyees dans utils.excel_handler et n'etaient
atteignables qu'a travers un acces disque. Isolees, elles se testent seules.
"""

import pytest
from openpyxl import Workbook

from utils.storage.sheet import (
    index_header_map,
    next_empty_row,
    read_data_rows,
    resolve_sheet_name,
    write_row,
)


@pytest.fixture
def feuille():
    wb = Workbook()
    ws = wb.active
    ws.title = "Donnees"
    return ws


class TestResolveSheetName:
    def test_correspondance_exacte(self):
        wb = Workbook()
        wb.active.title = "Visite_excursion"
        assert resolve_sheet_name(wb, "Visite_excursion") == "Visite_excursion"

    def test_casse_differente(self):
        wb = Workbook()
        wb.active.title = "visite_excursion"
        assert resolve_sheet_name(wb, "Visite_excursion") == "visite_excursion"

    def test_underscore_contre_espace(self):
        wb = Workbook()
        wb.active.title = "Visite excursion"
        assert resolve_sheet_name(wb, "Visite_excursion") == "Visite excursion"

    def test_accents_ignores(self):
        wb = Workbook()
        wb.active.title = "Hotels"
        assert resolve_sheet_name(wb, "Hôtels") == "Hotels"

    def test_feuille_absente(self):
        wb = Workbook()
        wb.active.title = "Autre"
        assert resolve_sheet_name(wb, "Introuvable") is None


class TestIndexHeaderMap:
    def test_numerote_a_partir_de_un(self):
        assert index_header_map(["A", "B", "C"]) == {"A": 1, "B": 2, "C": 3}

    def test_liste_vide(self):
        assert index_header_map([]) == {}


class TestReadDataRows:
    def test_lit_les_lignes_avec_leur_numero(self, feuille):
        feuille.cell(row=2, column=1, value="a1")
        feuille.cell(row=3, column=1, value="a2")

        rows = read_data_rows(feuille, {"A": 1}, 2)

        assert [r["row_number"] for r in rows] == [2, 3]
        assert [r["A"] for r in rows] == ["a1", "a2"]

    def test_cellule_vide_devient_chaine_vide(self, feuille):
        feuille.cell(row=2, column=1, value="a1")

        rows = read_data_rows(feuille, {"A": 1, "B": 2}, 2)

        assert rows[0]["B"] == ""

    def test_ligne_entierement_vide_ignoree(self, feuille):
        feuille.cell(row=2, column=1, value="a1")
        feuille.cell(row=4, column=1, value="a3")

        rows = read_data_rows(feuille, {"A": 1}, 2)

        assert [r["row_number"] for r in rows] == [2, 4]

    def test_respecte_la_ligne_de_depart(self, feuille):
        feuille.cell(row=2, column=1, value="entete_secondaire")
        feuille.cell(row=3, column=1, value="a1")

        rows = read_data_rows(feuille, {"A": 1}, 3)

        assert [r["A"] for r in rows] == ["a1"]

    def test_zero_conserve_sa_valeur(self, feuille):
        """0 n'est pas vide : la ligne doit etre retenue."""
        feuille.cell(row=2, column=1, value=0)

        rows = read_data_rows(feuille, {"A": 1}, 2)

        assert rows[0]["A"] == 0


class TestNextEmptyRow:
    def test_feuille_vide_renvoie_la_ligne_de_depart(self, feuille):
        assert next_empty_row(feuille, {"A": 1}) == 2

    def test_saute_les_lignes_occupees(self, feuille):
        feuille.cell(row=2, column=1, value="a1")
        feuille.cell(row=3, column=1, value="a2")

        assert next_empty_row(feuille, {"A": 1}) == 4

    def test_reutilise_un_trou(self, feuille):
        """Difference avec max_row + 1 : le balayage comble les trous."""
        feuille.cell(row=2, column=1, value="a1")
        feuille.cell(row=4, column=1, value="a3")

        assert next_empty_row(feuille, {"A": 1}) == 3

    def test_ne_regarde_que_les_colonnes_connues(self, feuille):
        feuille.cell(row=2, column=5, value="hors perimetre")

        assert next_empty_row(feuille, {"A": 1}) == 2


class TestWriteRow:
    def test_ecrit_les_valeurs(self, feuille):
        write_row(feuille, {"A": 1, "B": 2}, 2, {"A": "a1", "B": "b1"})

        assert feuille.cell(row=2, column=1).value == "a1"
        assert feuille.cell(row=2, column=2).value == "b1"

    def test_colonne_absente_est_vidée(self, feuille):
        feuille.cell(row=2, column=2, value="ancien")

        write_row(feuille, {"A": 1, "B": 2}, 2, {"A": "a1"})

        assert feuille.cell(row=2, column=2).value == ""
