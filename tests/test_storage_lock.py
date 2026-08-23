"""Verrou d'ecriture entre processus sur les classeurs.

`open_workbook(write=True)` fait un cycle lecture-modification-ecriture : il
charge le classeur entier, cede la feuille a l'appelant, puis reecrit le tout.
Deux postes travaillant sur un dossier partage entrent en collision sur toute
la duree de ce cycle, et non seulement pendant la sauvegarde. Le second a
enregistrer ecrase les modifications du premier, sans erreur ni message.

Le verrou est un fichier voisin cree en O_CREAT|O_EXCL, et non un flock :
`fcntl.flock` n'est pas fiable sur les partages reseau SMB et NFS, qui sont
precisement le cas d'usage multi-poste vise ici.
"""

import multiprocessing
import os
import threading
import time

import pytest
from openpyxl import Workbook, load_workbook

from utils.storage.workbook import (
    WorkbookLocked,
    open_workbook,
    verrou_de_classeur,
)


def _prendre_depuis_un_autre_fil(chemin, tenu, relacher):
    """Detient le verrou dans un fil distinct.

    La reentrance est par fil : une prise imbriquee dans le meme fil passe,
    ce qui est voulu. Pour eprouver la contention il faut donc un concurrent
    reel -- ici un fil, dans les tests un processus.
    """
    with verrou_de_classeur(chemin):
        tenu.set()
        relacher.wait(timeout=10)


@pytest.fixture
def classeur(tmp_path):
    chemin = str(tmp_path / "data.xlsx")
    wb = Workbook()
    wb.active.title = "F"
    wb.active.cell(row=1, column=1, value="entete")
    wb.save(chemin)
    wb.close()
    return chemin


def _lignes(chemin):
    wb = load_workbook(chemin)
    ws = wb["F"]
    valeurs = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    wb.close()
    return [v for v in valeurs if v is not None]


# ── Le verrou lui-meme ────────────────────────────────────────────────────────


class TestVerrou:
    def test_acquiert_et_libere(self, classeur):
        with verrou_de_classeur(classeur):
            assert os.path.exists(classeur + ".lock")
        assert not os.path.exists(classeur + ".lock")

    def test_libere_meme_en_cas_d_exception(self, classeur):
        with pytest.raises(ValueError):
            with verrou_de_classeur(classeur):
                raise ValueError("boum")
        assert not os.path.exists(classeur + ".lock")

    def test_un_second_preneur_echoue_apres_le_delai(self, classeur):
        tenu, relacher = threading.Event(), threading.Event()
        fil = threading.Thread(
            target=_prendre_depuis_un_autre_fil, args=(classeur, tenu, relacher)
        )
        fil.start()
        tenu.wait(timeout=5)
        try:
            debut = time.perf_counter()
            with pytest.raises(WorkbookLocked):
                with verrou_de_classeur(classeur, delai_max_s=0.3):
                    pass
            assert time.perf_counter() - debut >= 0.3
        finally:
            relacher.set()
            fil.join(timeout=5)

    def test_le_verrou_nomme_son_detenteur(self, classeur):
        """Pour diagnostiquer un verrou coince sans avoir a deviner."""
        with verrou_de_classeur(classeur):
            contenu = open(classeur + ".lock", encoding="utf-8").read()
        assert str(os.getpid()) in contenu

    def test_un_verrou_perime_est_repris(self, classeur):
        """Une application tuee laisse son verrou : il ne doit pas bloquer a vie."""
        chemin_verrou = classeur + ".lock"
        with open(chemin_verrou, "w", encoding="utf-8") as f:
            f.write("machine=fantome pid=999999")
        vieux = time.time() - 3600
        os.utime(chemin_verrou, (vieux, vieux))

        with verrou_de_classeur(classeur, delai_max_s=0.5, peremption_s=1.0):
            pass  # doit reussir sans attendre

    def test_un_verrou_recent_n_est_pas_repris(self, classeur):
        chemin_verrou = classeur + ".lock"
        with open(chemin_verrou, "w", encoding="utf-8") as f:
            f.write("machine=autre pid=1")
        with pytest.raises(WorkbookLocked):
            with verrou_de_classeur(classeur, delai_max_s=0.2, peremption_s=3600):
                pass
        os.unlink(chemin_verrou)


# ── Integration avec open_workbook ────────────────────────────────────────────


class TestOpenWorkbook:
    def test_l_ecriture_prend_le_verrou(self, classeur):
        with open_workbook(classeur, "F", write=True) as ws:
            assert os.path.exists(classeur + ".lock")
            ws.cell(row=2, column=1, value="x")
        assert not os.path.exists(classeur + ".lock")

    def test_la_lecture_ne_prend_pas_le_verrou(self, classeur):
        """Les lectures doivent rester paralleles : elles ne modifient rien."""
        with open_workbook(classeur, "F") as ws:
            assert ws is not None
            assert not os.path.exists(classeur + ".lock")

    def test_plusieurs_lectures_simultanees(self, classeur):
        with open_workbook(classeur, "F"):
            with open_workbook(classeur, "F"):
                pass  # ne doit pas bloquer


# ── Le scenario reel ──────────────────────────────────────────────────────────


def _ajouter_ligne(chemin, valeur, barriere):
    """Ajoute une ligne. Lance depuis un autre processus."""
    import sys

    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    from utils.storage.workbook import open_workbook

    barriere.wait()  # les deux processus demarrent ensemble
    with open_workbook(chemin, "F", write=True, backup=False, delai_max_s=30) as ws:
        time.sleep(0.15)  # elargit la fenetre lecture-modification-ecriture
        ws.cell(row=ws.max_row + 1, column=1, value=valeur)


class TestDeuxPostes:
    """Le scenario qui perdait des donnees : deux postes, un dossier partage."""

    def test_aucune_ecriture_n_est_perdue(self, classeur):
        barriere = multiprocessing.Barrier(2)
        processus = [
            multiprocessing.Process(
                target=_ajouter_ligne, args=(classeur, f"poste-{i}", barriere)
            )
            for i in range(2)
        ]
        for p in processus:
            p.start()
        for p in processus:
            p.join(timeout=60)

        assert all(p.exitcode == 0 for p in processus), "un processus a echoue"
        assert sorted(_lignes(classeur)) == ["poste-0", "poste-1"]


# ── Le decorateur applique aux fonctions d'ecriture d'excel_handler ───────────


class TestSousVerrou:
    """Les 36 fonctions d'ecriture d'excel_handler font toutes un cycle
    lecture-modification-ecriture a la main. Le decorateur prend le verrou sur
    la totalite de l'appel, la ou `_save_workbook` ne protegeait que la
    sauvegarde : entre le chargement et l'enregistrement, un autre poste
    pouvait lire une version qu'il allait ensuite ecraser.
    """

    def test_le_verrou_est_tenu_pendant_tout_l_appel(self, classeur):
        import utils.excel_handler as eh

        vu = {}

        @eh._sous_verrou_ecriture(lambda: classeur, sentinelle=-2)
        def ecrire():
            vu["verrou_present"] = os.path.exists(classeur + ".lock")
            return 42

        assert ecrire() == 42
        assert vu["verrou_present"] is True
        assert not os.path.exists(classeur + ".lock")

    def test_verrou_indisponible_renvoie_la_sentinelle(self, classeur):
        import utils.excel_handler as eh

        @eh._sous_verrou_ecriture(lambda: classeur, sentinelle=-2, delai_max_s=0.2)
        def ecrire():
            return "ne doit pas etre atteint"

        tenu, relacher = threading.Event(), threading.Event()
        fil = threading.Thread(
            target=_prendre_depuis_un_autre_fil, args=(classeur, tenu, relacher)
        )
        fil.start()
        tenu.wait(timeout=5)
        try:
            assert ecrire() == -2
        finally:
            relacher.set()
            fil.join(timeout=5)

    def test_la_sentinelle_des_suppressions_est_False(self, classeur):
        import utils.excel_handler as eh

        @eh._sous_verrou_ecriture(lambda: classeur, sentinelle=False, delai_max_s=0.2)
        def supprimer():
            return True

        tenu, relacher = threading.Event(), threading.Event()
        fil = threading.Thread(
            target=_prendre_depuis_un_autre_fil, args=(classeur, tenu, relacher)
        )
        fil.start()
        tenu.wait(timeout=5)
        try:
            assert supprimer() is False
        finally:
            relacher.set()
            fil.join(timeout=5)

    def test_les_exceptions_metier_remontent(self, classeur):
        import utils.excel_handler as eh

        @eh._sous_verrou_ecriture(lambda: classeur, sentinelle=-2)
        def ecrire():
            raise ValueError("erreur metier")

        with pytest.raises(ValueError):
            ecrire()
        assert not os.path.exists(classeur + ".lock")


class TestToutesLesEcrituresSontProtegees:
    """Garde-fou : une nouvelle fonction d'ecriture ne doit pas passer entre
    les mailles. Le test relit le module et verifie que chaque fonction
    appelant `_save_workbook` porte bien le decorateur."""

    def test_aucune_ecriture_sans_verrou(self):
        import ast
        import pathlib

        src = pathlib.Path("utils/excel_handler.py").read_text(encoding="utf-8")
        arbre = ast.parse(src)
        manquantes = []
        for node in arbre.body:
            if not isinstance(node, ast.FunctionDef):
                continue
            if node.name == "_save_workbook":
                continue
            corps = ast.unparse(node)
            if "_save_workbook(" not in corps:
                continue
            decorateurs = [ast.unparse(d) for d in node.decorator_list]
            if not any("_sous_verrou_ecriture" in d for d in decorateurs):
                manquantes.append(node.name)
        assert manquantes == [], f"ecritures sans verrou : {manquantes}"


class TestReentrance:
    """Le verrou doit etre reentrant dans un meme fil.

    `save_client_to_excel` appelle `_save_client_infos_to_excel` : deux
    fonctions decorees, un seul classeur. Sans reentrance, l'appel interne
    attend le verrou que l'appel externe detient deja, jusqu'a l'expiration du
    delai -- et la feuille INFOS_CLIENTS n'etait jamais ecrite, silencieusement.
    """

    def test_deux_prises_imbriquees_dans_le_meme_fil(self, classeur):
        with verrou_de_classeur(classeur, delai_max_s=0.5):
            with verrou_de_classeur(classeur, delai_max_s=0.5):
                assert os.path.exists(classeur + ".lock")
        assert not os.path.exists(classeur + ".lock")

    def test_le_verrou_n_est_liberé_qu_a_la_sortie_de_la_prise_externe(self, classeur):
        with verrou_de_classeur(classeur):
            with verrou_de_classeur(classeur):
                pass
            assert os.path.exists(classeur + ".lock"), "libere trop tot"
        assert not os.path.exists(classeur + ".lock")

    def test_deux_classeurs_differents_restent_independants(self, tmp_path, classeur):
        autre = str(tmp_path / "autre.xlsx")
        with verrou_de_classeur(classeur):
            with verrou_de_classeur(autre, delai_max_s=0.5):
                assert os.path.exists(autre + ".lock")

    def test_l_ecriture_imbriquee_aboutit(self, classeur):
        """Le scenario reel : une fonction decoree en appelle une autre."""
        import utils.excel_handler as eh

        @eh._sous_verrou_ecriture(lambda: classeur, sentinelle=-2, delai_max_s=1.0)
        def interne():
            return "ecrit"

        @eh._sous_verrou_ecriture(lambda: classeur, sentinelle=-2, delai_max_s=1.0)
        def externe():
            return interne()

        assert externe() == "ecrit"
