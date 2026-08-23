"""Tests du socle d'ouverture des classeurs (utils.storage.workbook)."""

import os

import pytest
from openpyxl import Workbook, load_workbook

from utils.storage import workbook as wbmod
from utils.storage.workbook import (
    SheetMissing,
    StorageUnavailable,
    WorkbookLocked,
    WorkbookMissing,
    create_backup,
    open_workbook,
    sentinel_on_error,
)


@pytest.fixture
def classeur(tmp_path):
    """Classeur existant avec une feuille "Donnees" et une valeur."""
    path = str(tmp_path / "test.xlsx")
    wb = Workbook()
    wb.active.title = "Donnees"
    wb.active.cell(row=1, column=1, value="entete")
    wb.save(path)
    wb.close()
    return path


@pytest.fixture
def chemin_absent(tmp_path):
    return str(tmp_path / "absent.xlsx")


class TestLecture:
    def test_cede_la_feuille_demandee(self, classeur):
        with open_workbook(classeur, "Donnees") as ws:
            assert ws.cell(row=1, column=1).value == "entete"

    def test_sans_feuille_cede_le_classeur(self, classeur):
        with open_workbook(classeur) as wb:
            assert "Donnees" in wb.sheetnames

    def test_fichier_absent_leve_workbook_missing(self, chemin_absent):
        with pytest.raises(WorkbookMissing):
            with open_workbook(chemin_absent, "Donnees"):
                pass

    def test_feuille_absente_leve_sheet_missing(self, classeur):
        with pytest.raises(SheetMissing):
            with open_workbook(classeur, "Inconnue"):
                pass

    def test_sans_openpyxl_leve_storage_unavailable(self, classeur, monkeypatch):
        monkeypatch.setattr(wbmod, "OPENPYXL_AVAILABLE", False)
        with pytest.raises(StorageUnavailable):
            with open_workbook(classeur, "Donnees"):
                pass

    def test_lecture_seule_ne_modifie_pas_le_fichier(self, classeur):
        avant = os.path.getmtime(classeur)
        with open_workbook(classeur, "Donnees") as ws:
            ws.cell(row=2, column=1, value="ignore")
        assert os.path.getmtime(classeur) == avant


class TestEcriture:
    def test_write_sauvegarde_les_modifications(self, classeur):
        with open_workbook(classeur, "Donnees", write=True) as ws:
            ws.cell(row=2, column=1, value="ajout")

        wb = load_workbook(classeur)
        assert wb["Donnees"].cell(row=2, column=1).value == "ajout"
        wb.close()

    def test_create_cree_le_fichier_et_la_feuille(self, chemin_absent):
        with open_workbook(chemin_absent, "Neuve", create=True, write=True) as ws:
            ws.cell(row=1, column=1, value="x")

        assert os.path.exists(chemin_absent)
        wb = load_workbook(chemin_absent)
        assert wb.sheetnames == ["Neuve"]
        wb.close()

    def test_create_ajoute_une_feuille_a_un_classeur_existant(self, classeur):
        with open_workbook(classeur, "Seconde", create=True, write=True) as ws:
            ws.cell(row=1, column=1, value="y")

        wb = load_workbook(classeur)
        assert set(wb.sheetnames) == {"Donnees", "Seconde"}
        wb.close()

    def test_backup_cree_une_copie_avant_ecriture(self, classeur):
        with open_workbook(classeur, "Donnees", write=True, backup=True) as ws:
            ws.cell(row=2, column=1, value="ajout")

        backups = os.listdir(os.path.join(os.path.dirname(classeur), "backups"))
        assert len(backups) == 1

    def test_backup_sans_effet_si_le_fichier_est_neuf(self, chemin_absent):
        with open_workbook(
            chemin_absent, "Neuve", create=True, write=True, backup=True
        ) as ws:
            ws.cell(row=1, column=1, value="x")

        assert not os.path.exists(
            os.path.join(os.path.dirname(chemin_absent), "backups")
        )

    def test_backup_est_actif_par_defaut_en_ecriture(self, classeur):
        """Une ecriture sauvegarde sans qu'on ait a le demander.

        Le defaut inverse a laisse 59 ecritures pour 3 backups dans le code
        historique : la protection ne doit pas dependre du site d'appel.
        """
        with open_workbook(classeur, "Donnees", write=True) as ws:
            ws.cell(row=2, column=1, value="ajout")

        backups = os.listdir(os.path.join(os.path.dirname(classeur), "backups"))
        assert len(backups) == 1

    def test_backup_desactivable_explicitement(self, classeur):
        with open_workbook(classeur, "Donnees", write=True, backup=False) as ws:
            ws.cell(row=2, column=1, value="ajout")

        assert not os.path.exists(os.path.join(os.path.dirname(classeur), "backups"))

    def test_lecture_seule_ne_sauvegarde_pas(self, classeur):
        with open_workbook(classeur, "Donnees") as ws:
            _ = ws.cell(row=1, column=1).value

        assert not os.path.exists(os.path.join(os.path.dirname(classeur), "backups"))

    def test_echec_de_sauvegarde_laisse_le_fichier_intact(self, classeur, monkeypatch):
        """Une ecriture interrompue ne doit pas tronquer le classeur.

        wb.save() ecrivait directement sur le fichier de destination : une
        coupure en cours de route laissait un .xlsx illisible.
        """
        original = open(classeur, "rb").read()

        def saborder(self, chemin):
            with open(chemin, "wb") as f:
                f.write(b"PK\x03\x04 tronque")
            raise OSError("disque plein")

        monkeypatch.setattr(Workbook, "save", saborder)

        with pytest.raises(OSError):
            with open_workbook(classeur, "Donnees", write=True) as ws:
                ws.cell(row=2, column=1, value="ajout")

        assert open(classeur, "rb").read() == original
        assert (
            load_workbook(classeur)["Donnees"].cell(row=1, column=1).value == "entete"
        )

    def test_aucun_fichier_temporaire_ne_survit(self, classeur):
        with open_workbook(classeur, "Donnees", write=True) as ws:
            ws.cell(row=2, column=1, value="ajout")

        restes = [
            n
            for n in os.listdir(os.path.dirname(classeur))
            if n.endswith(".tmp") or n.startswith(".")
        ]
        assert restes == []

    def test_exception_dans_le_bloc_annule_la_sauvegarde(self, classeur):
        avant = os.path.getmtime(classeur)
        with pytest.raises(ValueError):
            with open_workbook(classeur, "Donnees", write=True) as ws:
                ws.cell(row=2, column=1, value="ajout")
                raise ValueError("echec metier")
        assert os.path.getmtime(classeur) == avant

    def test_fichier_verrouille_leve_workbook_locked(self, classeur, monkeypatch):
        def refuser(*args, **kwargs):
            raise PermissionError(classeur)

        monkeypatch.setattr(wbmod, "load_workbook", refuser)
        with pytest.raises(WorkbookLocked):
            with open_workbook(classeur, "Donnees"):
                pass


class TestSentinelOnError:
    """Traduction des exceptions vers les codes de retour historiques."""

    def test_valeur_normale_traverse(self):
        @sentinel_on_error(-1)
        def operation():
            return 42

        assert operation() == 42

    @pytest.mark.parametrize(
        "exception",
        [StorageUnavailable, WorkbookMissing, SheetMissing],
    )
    def test_causes_attendues_renvoient_la_sentinelle(self, exception):
        @sentinel_on_error(-1)
        def operation():
            raise exception("x")

        assert operation() == -1

    def test_verrouillage_a_sa_propre_sentinelle(self):
        @sentinel_on_error(-1, -2)
        def operation():
            raise WorkbookLocked("x")

        assert operation() == -2

    def test_verrouillage_retombe_sur_missing_par_defaut(self):
        @sentinel_on_error(False)
        def operation():
            raise WorkbookLocked("x")

        assert operation() is False

    def test_erreur_inattendue_est_journalisee(self, monkeypatch):
        appels = []
        monkeypatch.setattr(wbmod.logger, "error", lambda msg, **kw: appels.append(msg))

        @sentinel_on_error([], label="lecture")
        def operation():
            raise RuntimeError("panne")

        assert operation() == []
        assert len(appels) == 1
        assert "lecture" in appels[0]

    def test_causes_attendues_ne_sont_pas_journalisees(self, monkeypatch):
        """Comportement d'origine : pas de bruit pour un fichier absent."""
        appels = []
        monkeypatch.setattr(wbmod.logger, "error", lambda msg, **kw: appels.append(msg))

        @sentinel_on_error([])
        def operation():
            raise WorkbookMissing("x")

        assert operation() == []
        assert appels == []

    def test_conserve_nom_et_docstring(self):
        @sentinel_on_error(-1)
        def operation():
            """Documentation."""

        assert operation.__name__ == "operation"
        assert operation.__doc__ == "Documentation."


class TestCreateBackup:
    def test_cree_une_copie_horodatee(self, classeur):
        chemin = create_backup(classeur)
        assert chemin is not None
        assert os.path.exists(chemin)
        assert chemin.endswith(".bak")

    def test_purge_les_copies_les_plus_anciennes(self, classeur):
        """Le backup systematique ne doit pas remplir le disque.

        Une agence qui enregistre 60 fois par jour un classeur de 150 Ko
        accumulerait 3 Go par an sans plafond.
        """
        for i in range(wbmod.MAX_BACKUPS_PER_FILE + 5):
            create_backup(classeur, _horodatage=f"20260101_0000{i:02d}")

        backup_dir = os.path.join(os.path.dirname(classeur), "backups")
        restants = sorted(os.listdir(backup_dir))
        assert len(restants) == wbmod.MAX_BACKUPS_PER_FILE
        # Les plus recentes sont conservees, les plus anciennes supprimees.
        assert restants[-1].endswith(f"0000{wbmod.MAX_BACKUPS_PER_FILE + 4:02d}.bak")

    def test_purge_ne_touche_pas_aux_autres_classeurs(self, classeur, tmp_path):
        autre = str(tmp_path / "autre.xlsx")
        wb = Workbook()
        wb.save(autre)
        wb.close()

        create_backup(autre, _horodatage="20260101_000000")
        for i in range(wbmod.MAX_BACKUPS_PER_FILE + 3):
            create_backup(classeur, _horodatage=f"20260101_0000{i:02d}")

        backup_dir = os.path.join(os.path.dirname(classeur), "backups")
        noms = os.listdir(backup_dir)
        assert sum(1 for n in noms if n.startswith("autre.xlsx")) == 1

    def test_fichier_absent_renvoie_none(self, chemin_absent):
        assert create_backup(chemin_absent) is None
