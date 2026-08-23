"""Resolution du nom de feuille, tolerante aux variantes d'ecriture.

`config.py` nomme les feuilles telles qu'elles etaient censees s'appeler ;
les classeurs reels s'en ecartent. Constate sur data-hotel.xlsx :

    config.AVION_SOURCE_SHEET_NAME             = "avion"
    feuille reelle                             = "Avion"
    config.VISITE_EXCURSION_SOURCE_SHEET_NAME  = "Visite_excursion"
    feuille reelle                             = "Visite&excursion"

Le premier ecart etait rattrape par un repli maison dans load_avion_source_data.
Le second ne l'etait pas : `if NOM not in wb.sheetnames: return []` renvoyait une
liste vide, et le catalogue visites/excursions restait invisible dans
l'application sans le moindre message.
"""

import pytest
from openpyxl import Workbook

from utils.storage.sheet import find_sheet, normalize_sheet_key


class TestNormalizeSheetKey:
    @pytest.mark.parametrize(
        "a, b",
        [
            ("avion", "Avion"),
            ("Visite_excursion", "Visite&excursion"),
            ("Visite_excursion", "VISITE & EXCURSION"),
            ("Frais collectifs", "frais_collectifs"),
            ("KM_MADA", "km mada"),
            ("Parametre", "PARAMÈTRE"),
        ],
    )
    def test_variantes_equivalentes(self, a, b):
        assert normalize_sheet_key(a) == normalize_sheet_key(b)

    @pytest.mark.parametrize(
        "a, b",
        [
            ("TRANSPORT", "TRANSPORTS"),
            ("Circuits", "Circuit"),
            ("AVION", "COTATION_AVION"),
        ],
    )
    def test_noms_distincts_le_restent(self, a, b):
        assert normalize_sheet_key(a) != normalize_sheet_key(b)

    def test_valeur_vide(self):
        assert normalize_sheet_key(None) == ""
        assert normalize_sheet_key("   ") == ""


class TestFindSheet:
    @pytest.fixture
    def classeur(self):
        wb = Workbook()
        wb.active.title = "BDD_HOTEL"
        wb.create_sheet("Visite&excursion")
        wb.create_sheet("Avion")
        yield wb
        wb.close()

    def test_correspondance_exacte(self, classeur):
        assert find_sheet(classeur, "BDD_HOTEL") == "BDD_HOTEL"

    def test_casse_differente(self, classeur):
        assert find_sheet(classeur, "avion") == "Avion"

    def test_separateur_different(self, classeur):
        assert find_sheet(classeur, "Visite_excursion") == "Visite&excursion"

    def test_absente_renvoie_none(self, classeur):
        assert find_sheet(classeur, "INEXISTANTE") is None

    def test_le_nom_exact_prime_sur_une_variante(self):
        """Deux orthographes coexistantes : chacune resout vers elle-meme.

        (Excel interdit deux feuilles ne differant que par la casse, mais
        `Visite_excursion` et `Visite&excursion` peuvent bien coexister.)
        """
        wb = Workbook()
        wb.active.title = "Visite_excursion"
        wb.create_sheet("Visite&excursion")
        try:
            assert find_sheet(wb, "Visite_excursion") == "Visite_excursion"
            assert find_sheet(wb, "Visite&excursion") == "Visite&excursion"
        finally:
            wb.close()
