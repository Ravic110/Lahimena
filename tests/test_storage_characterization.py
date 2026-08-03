"""Tests de caracterisation de la couche d'acces aux donnees Excel.

Ces tests figent le comportement ACTUEL de utils.excel_handler avant sa
refactorisation en utils/storage/. Ils ne decrivent pas un comportement
souhaitable : ils capturent l'existant, bizarreries comprises, pour que le
deplacement du code soit verifiable.

Methode : instantane. Chaque entite joue un scenario CRUD identique contre un
classeur temporaire ; on enregistre les valeurs de retour ET le contenu brut de
la feuille apres chaque etape. Le resultat est compare a une reference
versionnee (`tests/fixtures/storage_characterization.json`).

L'instantane est necessaire parce que les six entites, en apparence jumelles,
divergent en realite sur quatre axes : decouverte des en-tetes, vocabulaire des
cles, point d'insertion, et coercition des valeurs. Ecrire des assertions a la
main entite par entite serait devine ; l'instantane, lui, constate.

Regeneration (uniquement pour acter un changement de comportement voulu) :
    LAHIMENA_UPDATE_SNAPSHOT=1 pytest tests/test_storage_characterization.py
"""

import json
import os
from pathlib import Path

import pytest
from openpyxl import load_workbook

import config
from config import (
    AVION_SOURCE_SHEET_NAME,
    FRAIS_COLLECTIFS_SHEET_NAME,
    KM_MADA_SHEET_NAME,
    TRANSPORT_SOURCE_SHEET_NAME,
    VISITE_EXCURSION_SOURCE_SHEET_NAME,
)
from utils import excel_handler
from utils.storage import workbook as wbmod

SNAPSHOT_PATH = Path(__file__).parent / "fixtures" / "storage_characterization.json"
UPDATE_SNAPSHOT = os.getenv("LAHIMENA_UPDATE_SNAPSHOT") == "1"

# Cles metier propres a collective_expense : c'est la seule entite dont le
# vocabulaire d'entree ne correspond pas aux libelles de colonnes.
COLLECTIVE_ROWS = (
    {
        "forfait": "F1",
        "prestataire": "P1",
        "designation": "D1",
        "montant": "1 200",
        "id_circuit": "C1",
    },
    {
        "forfait": "F2",
        "prestataire": "P2",
        "designation": "D2",
        "montant": 800,
        "id_circuit": "C2",
    },
    {
        "forfait": "F3",
        "prestataire": "P3",
        "designation": "D3",
        "montant": 950,
        "id_circuit": "C3",
    },
)

# Pour les autres, les cles du dict deviennent les en-tetes de colonnes.
GENERIC_ROWS = (
    {"Colonne A": "a1", "Colonne B": "b1"},
    {"Colonne A": "a2", "Colonne B": "b2"},
    {"Colonne A": "a3", "Colonne B": "b3"},
)

# transport et km_mada resolvent leurs en-tetes en cherchant des colonnes
# sentinelles ("Prestataire" + "Type de voiture", "REPERES"). Sans elles, le
# resolveur renvoie un mapping vide et load/update tombent dans leur chemin
# degenere. On fournit donc des libelles reconnus pour exercer le cas nominal.
TRANSPORT_ROWS = (
    {"Prestataire": "P1", "Type de voiture": "4x4", "Nombre de place": 4},
    {"Prestataire": "P2", "Type de voiture": "Minibus", "Nombre de place": 12},
    {"Prestataire": "P3", "Type de voiture": "Berline", "Nombre de place": 3},
)

KM_MADA_ROWS = (
    {"REPERES": "Tana-Tamatave", "KM": 350},
    {"REPERES": "Tana-Majunga", "KM": 570},
    {"REPERES": "Tana-Tulear", "KM": 940},
)

ENTITIES = {
    "circuit": ("Circuits", "circuit_db_row", "circuit_db_rows", GENERIC_ROWS),
    "collective_expense": (
        FRAIS_COLLECTIFS_SHEET_NAME,
        "collective_expense_db_row",
        "collective_expense_db_rows",
        COLLECTIVE_ROWS,
    ),
    "transport": (
        TRANSPORT_SOURCE_SHEET_NAME,
        "transport_db_row",
        "transport_db_rows",
        TRANSPORT_ROWS,
    ),
    "avion": (AVION_SOURCE_SHEET_NAME, "avion_db_row", "avion_db_rows", GENERIC_ROWS),
    "visite_excursion": (
        VISITE_EXCURSION_SOURCE_SHEET_NAME,
        "visite_excursion_db_row",
        "visite_excursion_db_rows",
        GENERIC_ROWS,
    ),
    "km_mada": (KM_MADA_SHEET_NAME, "km_mada_db_row", "km_mada_db_rows", KM_MADA_ROWS),
}


def _rediriger_classeur(monkeypatch, path):
    """Redirige data-hotel.xlsx vers un chemin temporaire.

    Deux points de substitution pendant la migration : les fonctions restees
    dans excel_handler lisent la constante importee dans leur module, celles
    deplacees vers utils.storage la relisent depuis config a chaque appel.
    """
    monkeypatch.setattr(excel_handler, "HOTEL_EXCEL_PATH", path)
    monkeypatch.setattr(config, "HOTEL_EXCEL_PATH", path)


def _jsonable(value):
    """Rend une valeur comparable et serialisable, sans masquer son type."""
    if isinstance(value, bool) or value is None or isinstance(value, (int, float, str)):
        return value
    if isinstance(value, dict):
        return {str(k): _jsonable(v) for k, v in value.items()}
    if isinstance(value, (list, tuple)):
        return [_jsonable(v) for v in value]
    return f"<{type(value).__name__}:{value}>"


def _dump_sheet(path, sheet):
    """Contenu brut de la feuille, ligne par ligne. None si absente."""
    if not os.path.exists(path):
        return None
    wb = load_workbook(path)
    try:
        if sheet not in wb.sheetnames:
            return None
        ws = wb[sheet]
        return [
            [_jsonable(ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1)]
            for r in range(1, ws.max_row + 1)
        ]
    finally:
        wb.close()


def _run_scenario(tmp_path, monkeypatch, entity):
    """Joue le meme scenario CRUD pour une entite et enregistre tout."""
    sheet, row_suffix, rows_suffix, payloads = ENTITIES[entity]
    path = str(tmp_path / "data-hotel.xlsx")
    _rediriger_classeur(monkeypatch, path)

    save = getattr(excel_handler, f"save_{row_suffix}")
    update = getattr(excel_handler, f"update_{row_suffix}")
    delete = getattr(excel_handler, f"delete_{row_suffix}")
    load = getattr(excel_handler, f"load_{rows_suffix}")

    steps = {}

    # Sentinelles sur fichier absent, avant toute ecriture.
    steps["load_sur_fichier_absent"] = _jsonable(load())
    steps["update_sur_fichier_absent"] = _jsonable(update(2, payloads[0]))
    steps["delete_sur_fichier_absent"] = _jsonable(delete(2))

    # Le premier save doit creer le classeur et la feuille.
    steps["save_1"] = _jsonable(save(payloads[0]))
    steps["grille_apres_save_1"] = _dump_sheet(path, sheet)

    steps["save_2"] = _jsonable(save(payloads[1]))
    steps["grille_apres_save_2"] = _dump_sheet(path, sheet)
    steps["load_apres_saves"] = _jsonable(load())

    steps["update_ligne_2"] = _jsonable(update(2, payloads[2]))
    steps["grille_apres_update"] = _dump_sheet(path, sheet)
    steps["load_apres_update"] = _jsonable(load())

    steps["delete_ligne_2"] = _jsonable(delete(2))
    steps["grille_apres_delete"] = _dump_sheet(path, sheet)
    steps["load_apres_delete"] = _jsonable(load())

    return steps


def _load_reference():
    if not SNAPSHOT_PATH.exists():
        return {}
    return json.loads(SNAPSHOT_PATH.read_text(encoding="utf-8"))


@pytest.mark.parametrize("entity", sorted(ENTITIES))
def test_comportement_crud_inchange(entity, tmp_path, monkeypatch):
    """Le scenario CRUD doit reproduire l'instantane de reference."""
    observed = _run_scenario(tmp_path, monkeypatch, entity)

    if UPDATE_SNAPSHOT:
        reference = _load_reference()
        reference[entity] = observed
        SNAPSHOT_PATH.parent.mkdir(parents=True, exist_ok=True)
        SNAPSHOT_PATH.write_text(
            json.dumps(reference, indent=2, ensure_ascii=False, sort_keys=True) + "\n",
            encoding="utf-8",
        )
        pytest.skip(f"instantane regenere pour {entity}")

    reference = _load_reference()
    assert entity in reference, (
        f"Aucun instantane de reference pour {entity}. "
        "Generer avec LAHIMENA_UPDATE_SNAPSHOT=1."
    )

    expected = reference[entity]
    divergences = [
        f"  {step}:\n    attendu  = {expected.get(step)!r}\n    obtenu   = {value!r}"
        for step, value in observed.items()
        if expected.get(step) != value
    ]
    assert not divergences, (
        f"Comportement modifie pour {entity} :\n" + "\n".join(divergences)
    )


class TestOpenpyxlIndisponible:
    """Sentinelles renvoyees quand openpyxl manque, pour les six entites."""

    @pytest.fixture(autouse=True)
    def _sans_openpyxl(self, monkeypatch):
        # Deux points de substitution pendant la migration, comme pour le
        # chemin du classeur : l'ancien module et le nouveau socle.
        monkeypatch.setattr(excel_handler, "OPENPYXL_AVAILABLE", False)
        monkeypatch.setattr(wbmod, "OPENPYXL_AVAILABLE", False)

    @pytest.mark.parametrize("entity", sorted(ENTITIES))
    def test_sentinelles(self, entity, tmp_path, monkeypatch):
        sheet, row_suffix, rows_suffix, payloads = ENTITIES[entity]
        _rediriger_classeur(monkeypatch, str(tmp_path / "data-hotel.xlsx"))
        assert getattr(excel_handler, f"save_{row_suffix}")(payloads[0]) == -1
        assert getattr(excel_handler, f"update_{row_suffix}")(2, payloads[0]) == -1
        assert getattr(excel_handler, f"delete_{row_suffix}")(2) is False
        assert getattr(excel_handler, f"load_{rows_suffix}")() == []


class TestEnTetesEnDeuxiemeLigne:
    """transport et km_mada acceptent des en-tetes en ligne 1 OU en ligne 2.

    La feuille reelle utilise des colonnes groupees : une ligne de groupes
    surmonte la ligne d'en-tetes. Les resolveurs essaient donc la ligne 1, puis
    la ligne 2, et decalent le debut des donnees en consequence. C'est la
    branche la plus fragile au deplacement de code.
    """

    @pytest.fixture
    def classeur(self, tmp_path, monkeypatch):
        from openpyxl import Workbook

        path = str(tmp_path / "data-hotel.xlsx")
        _rediriger_classeur(monkeypatch, path)
        return path, Workbook()

    def test_transport_lit_les_donnees_a_partir_de_la_ligne_3(self, classeur):
        path, wb = classeur
        ws = wb.active
        ws.title = TRANSPORT_SOURCE_SHEET_NAME
        ws.cell(row=1, column=1, value="IDENTIFICATION")  # ligne de groupes
        ws.cell(row=2, column=1, value="Prestataire")
        ws.cell(row=2, column=2, value="Type de voiture")
        ws.cell(row=3, column=1, value="P1")
        ws.cell(row=3, column=2, value="4x4")
        wb.save(path)
        wb.close()

        rows = excel_handler.load_transport_db_rows()

        assert [r["row_number"] for r in rows] == [3]
        assert rows[0]["Prestataire"] == "P1"

    def test_km_mada_lit_les_donnees_a_partir_de_la_ligne_3(self, classeur):
        path, wb = classeur
        ws = wb.active
        ws.title = KM_MADA_SHEET_NAME
        ws.cell(row=1, column=1, value="DISTANCES")  # ligne de groupes
        ws.cell(row=2, column=1, value="REPERES")
        ws.cell(row=2, column=2, value="KM")
        ws.cell(row=3, column=1, value="Tana-Tamatave")
        ws.cell(row=3, column=2, value=350)
        wb.save(path)
        wb.close()

        rows = excel_handler.load_km_mada_db_rows()

        assert [r["row_number"] for r in rows] == [3]
        assert rows[0]["REPERES"] == "Tana-Tamatave"

    def test_sans_colonne_sentinelle_le_resolveur_renonce(self, classeur):
        """Comportement actuel : mapping vide -> liste vide, sans erreur."""
        path, wb = classeur
        ws = wb.active
        ws.title = KM_MADA_SHEET_NAME
        ws.cell(row=1, column=1, value="Colonne inconnue")
        ws.cell(row=2, column=1, value="valeur")
        wb.save(path)
        wb.close()

        assert excel_handler.load_km_mada_db_rows() == []


class TestInvalidationCacheKmMada:
    """km_mada est la seule entite qui invalide un cache sur ecriture.

    Divergence reelle entre les six : a preserver telle quelle.
    """

    def test_delete_invalide_le_cache(self, tmp_path, monkeypatch):
        path = str(tmp_path / "data-hotel.xlsx")
        _rediriger_classeur(monkeypatch, path)
        excel_handler.save_km_mada_db_row({"Colonne A": "a1"})

        excel_handler._KM_MADA_CACHE["loaded_at"] = 12345.0
        excel_handler.delete_km_mada_db_row(2)

        assert excel_handler._KM_MADA_CACHE["loaded_at"] == 0.0

    def test_les_autres_entites_ne_touchent_pas_ce_cache(self, tmp_path, monkeypatch):
        path = str(tmp_path / "data-hotel.xlsx")
        _rediriger_classeur(monkeypatch, path)
        excel_handler.save_circuit_db_row({"Colonne A": "a1"})

        excel_handler._KM_MADA_CACHE["loaded_at"] = 12345.0
        excel_handler.delete_circuit_db_row(2)

        assert excel_handler._KM_MADA_CACHE["loaded_at"] == 12345.0


class TestSurfaceApiPublique:
    """L'API publique doit rester stable pendant toute la refactorisation.

    39 fichiers importent des noms depuis utils.excel_handler. Ce test garantit
    qu'aucun ne disparait derriere la facade.
    """

    def test_toutes_les_fonctions_publiques_restent_appelables(self):
        manquantes = [
            nom for nom in API_PUBLIQUE if not callable(getattr(excel_handler, nom, None))
        ]
        assert manquantes == []

    def test_la_liste_de_reference_couvre_le_module(self):
        """Detecte une fonction publique ajoutee sans etre listee ici."""
        exposees = {
            nom
            for nom in dir(excel_handler)
            if not nom.startswith("_") and callable(getattr(excel_handler, nom))
        }
        # On ignore ce qui est importe depuis openpyxl / stdlib.
        definies_ici = {
            nom
            for nom in exposees
            if getattr(getattr(excel_handler, nom), "__module__", "")
            in ("utils.excel_handler", "utils.storage")
            or getattr(getattr(excel_handler, nom), "__module__", "").startswith(
                "utils.storage."
            )
        }
        assert definies_ici - set(API_PUBLIQUE) == set()


API_PUBLIQUE = [
    "calculate_invoice_totals",
    "create_backup",
    "delete_air_ticket_from_excel",
    "delete_avion_db_row",
    "delete_circuit_db_row",
    "delete_client_from_excel",
    "delete_collective_expense_db_row",
    "delete_collective_expense_from_excel",
    "delete_hotel_from_excel",
    "delete_km_mada_db_row",
    "delete_parametrage_from_excel",
    "delete_transport_db_row",
    "delete_transport_from_excel",
    "delete_visite_excursion_db_row",
    "delete_visite_excursion_from_excel",
    "get_avion_arrival_cities",
    "get_avion_compagnies",
    "get_avion_db_headers",
    "get_avion_departure_cities",
    "get_avion_headers",
    "get_avion_tarifs",
    "get_circuit_db_headers",
    "get_collective_expense_designations",
    "get_collective_expense_forfait",
    "get_collective_expense_headers",
    "get_collective_expense_montant",
    "get_collective_expense_prestataires",
    "get_km_mada_db_headers",
    "get_km_mada_duration_for_repere",
    "get_km_mada_km_for_repere",
    "get_km_mada_reperes",
    "get_parametrage_headers",
    "get_parametrage_value_by_name",
    "get_quotations_by_city",
    "get_quotations_grouped_by_client",
    "get_segment_distance",
    "get_transport_db_headers",
    "get_transport_fuel_price",
    "get_transport_headers",
    "get_transport_prestataires",
    "get_transport_vehicle_data",
    "get_transport_vehicle_types",
    "get_visite_excursion_db_headers",
    "get_visite_excursion_designations",
    "get_visite_excursion_headers",
    "get_visite_excursion_montant",
    "get_visite_excursion_prestataires",
    "load_active_client_invoice_from_excel",
    "load_active_client_quote_from_excel",
    "load_all_air_ticket_quotations",
    "load_all_circuits",
    "load_all_clients",
    "load_all_collective_expense_quotations",
    "load_all_hotel_quotations",
    "load_all_hotels",
    "load_all_invoices",
    "load_all_parametrages",
    "load_all_transport_quotations",
    "load_all_visite_excursion_quotations",
    "load_avion_db_rows",
    "load_avion_source_data",
    "load_circuit_catalog",
    "load_circuit_db_rows",
    "load_client_air_ticket_cotation",
    "load_client_collective_cotation",
    "load_client_hotel_cotation",
    "load_client_restauration_cotation",
    "load_client_transport_cotation",
    "load_collective_expense_db_rows",
    "load_collective_expenses_data",
    "load_financial_state_snapshot",
    "load_km_mada_db_rows",
    "load_transport_db_rows",
    "load_visite_excursion_data",
    "load_visite_excursion_db_rows",
    "migrate_normalize_infos_clients",
    "normalize_city_name",
    "refresh_financial_state_from_invoices",
    "save_active_client_invoice_to_excel",
    "save_active_client_quote_to_excel",
    "save_air_ticket_quotation_to_excel",
    "save_avion_db_row",
    "save_circuit_db_row",
    "save_client_air_ticket_cotation_to_excel",
    "save_client_collective_cotation_to_excel",
    "save_client_hotel_cotation_to_excel",
    "save_client_restauration_cotation_to_excel",
    "save_client_to_excel",
    "save_client_transport_cotation_to_excel",
    "save_collective_expense_db_row",
    "save_collective_expense_quotation_to_excel",
    "save_hotel_quotation_to_excel",
    "save_hotel_to_excel",
    "save_invoice_to_excel",
    "save_km_mada_db_row",
    "save_parametrage_to_excel",
    "save_transport_db_row",
    "save_transport_quotation_to_excel",
    "save_visite_excursion_db_row",
    "save_visite_excursion_quotation_to_excel",
    "update_air_ticket_quotation_in_excel",
    "update_avion_db_row",
    "update_circuit_db_row",
    "update_client_in_excel",
    "update_client_statut",
    "update_collective_expense_db_row",
    "update_collective_expense_quotation_in_excel",
    "update_hotel_in_excel",
    "update_invoice_in_excel",
    "update_km_mada_db_row",
    "update_parametrage_in_excel",
    "update_transport_db_row",
    "update_transport_quotation_in_excel",
    "update_visite_excursion_db_row",
    "update_visite_excursion_quotation_in_excel",
]
