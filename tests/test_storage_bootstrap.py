"""Amorcage des classeurs sur un poste neuf.

Le scenario couvert est celui d'un clone frais : `data.xlsx` et
`data-hotel.xlsx` sont absents (ils ne sont pas versionnes), et l'application
doit demarrer sur une structure coherente plutot que sur des lectures vides
indistinguables d'une base sans donnees.
"""

import os
import shutil

import pytest
from openpyxl import Workbook, load_workbook

import config
from utils.storage import bootstrap


@pytest.fixture
def atelier(tmp_path, monkeypatch):
    """Deux classeurs vivants et un dossier de gabarits, tous temporaires."""
    client = str(tmp_path / "data.xlsx")
    hotel = str(tmp_path / "data-hotel.xlsx")
    gabarits = str(tmp_path / "templates")

    monkeypatch.setattr(config, "CLIENT_EXCEL_PATH", client)
    monkeypatch.setattr(config, "HOTEL_EXCEL_PATH", hotel)
    monkeypatch.setattr(bootstrap, "TEMPLATES_DIR", gabarits)

    return {"client": client, "hotel": hotel, "gabarits": gabarits}


def _fabriquer_classeur(chemin, feuilles):
    """feuilles : {nom: (entetes, lignes_de_donnees)}"""
    wb = Workbook()
    wb.remove(wb.active)
    for nom, (entetes, donnees) in feuilles.items():
        ws = wb.create_sheet(nom)
        ws.append(entetes)
        for ligne in donnees:
            ws.append(ligne)
    wb.save(chemin)
    wb.close()


class TestExportTemplates:
    def test_le_gabarit_reprend_les_feuilles_et_les_entetes(self, atelier):
        _fabriquer_classeur(
            atelier["client"],
            {"DEMANDE_CLIENT": (["Date", "Nom"], [["2026-01-01", "Rakoto"]])},
        )
        _fabriquer_classeur(
            atelier["hotel"], {"PARAMETRE": (["parametre", "valeur"], [])}
        )

        ecrits = bootstrap.export_templates()
        assert len(ecrits) == 2

        wb = load_workbook(ecrits[0])
        try:
            assert wb.sheetnames == ["DEMANDE_CLIENT"]
            ws = wb["DEMANDE_CLIENT"]
            assert [c.value for c in ws[1]] == ["Date", "Nom"]
        finally:
            wb.close()

    def test_aucune_donnee_metier_ne_traverse(self, atelier):
        """Le gabarit est versionne : il ne doit porter aucune donnee client."""
        _fabriquer_classeur(
            atelier["client"],
            {
                "DEMANDE_CLIENT": (
                    ["Date", "Nom", "Email"],
                    [
                        ["2026-01-01", "Rakoto", "rakoto@example.mg"],
                        ["2026-02-02", "Rasoa", "rasoa@example.mg"],
                    ],
                )
            },
        )
        _fabriquer_classeur(
            atelier["hotel"], {"PARAMETRE": (["parametre", "valeur"], [])}
        )

        gabarit = bootstrap.export_templates()[0]
        wb = load_workbook(gabarit)
        try:
            ws = wb["DEMANDE_CLIENT"]
            assert ws.max_row == 1
            contenu = [
                ws.cell(row=r, column=c).value
                for r in range(1, ws.max_row + 1)
                for c in range(1, ws.max_column + 1)
            ]
            assert "Rakoto" not in contenu
            assert "rakoto@example.mg" not in contenu
        finally:
            wb.close()

    def test_feuille_a_bandeau_garde_ses_deux_lignes(self, atelier, monkeypatch):
        """BDD_HOTEL a un bandeau en L1 et ses en-tetes en L2."""
        wb = Workbook()
        ws = wb.active
        ws.title = "BDD_HOTEL"
        ws.append(["Hotel"])
        ws.append(["Ville", "HTL"])
        ws.append(["Tana", "Le Louvre"])
        wb.save(atelier["hotel"])
        wb.close()
        _fabriquer_classeur(atelier["client"], {"DEMANDE_CLIENT": (["Date"], [])})

        bootstrap.export_templates()
        gabarit = os.path.join(atelier["gabarits"], "data-hotel.template.xlsx")
        wb = load_workbook(gabarit)
        try:
            ws = wb["BDD_HOTEL"]
            assert ws.cell(row=1, column=1).value == "Hotel"
            assert ws.cell(row=2, column=1).value == "Ville"
            assert ws.max_row == 2  # la ligne "Tana" n'a pas suivi
        finally:
            wb.close()

    def test_classeur_source_absent_est_ignore(self, atelier):
        _fabriquer_classeur(atelier["client"], {"DEMANDE_CLIENT": (["Date"], [])})
        ecrits = bootstrap.export_templates()
        assert len(ecrits) == 1


class TestEnsureWorkbooks:
    @pytest.fixture
    def gabarits_prets(self, atelier):
        _fabriquer_classeur(
            atelier["client"], {"DEMANDE_CLIENT": (["Date", "Nom"], [])}
        )
        _fabriquer_classeur(
            atelier["hotel"], {"PARAMETRE": (["parametre", "valeur"], [])}
        )
        bootstrap.export_templates()
        os.remove(atelier["client"])
        os.remove(atelier["hotel"])
        return atelier

    def test_installe_les_classeurs_manquants(self, gabarits_prets):
        assert bootstrap.missing_workbooks() == [
            gabarits_prets["client"],
            gabarits_prets["hotel"],
        ]

        installes = bootstrap.ensure_workbooks()

        assert sorted(installes) == sorted(
            [gabarits_prets["client"], gabarits_prets["hotel"]]
        )
        assert bootstrap.missing_workbooks() == []

    def test_le_classeur_installe_a_la_bonne_structure(self, gabarits_prets):
        bootstrap.ensure_workbooks()
        wb = load_workbook(gabarits_prets["client"])
        try:
            assert wb.sheetnames == ["DEMANDE_CLIENT"]
            assert [c.value for c in wb["DEMANDE_CLIENT"][1]] == ["Date", "Nom"]
        finally:
            wb.close()

    def test_ne_touche_jamais_a_un_classeur_existant(self, gabarits_prets):
        """L'amorcage cree ; il ne repare pas et n'ecrase pas."""
        _fabriquer_classeur(
            gabarits_prets["client"],
            {"DEMANDE_CLIENT": (["Date", "Nom"], [["2026-01-01", "Rakoto"]])},
        )
        avant = open(gabarits_prets["client"], "rb").read()

        installes = bootstrap.ensure_workbooks()

        assert gabarits_prets["client"] not in installes
        assert open(gabarits_prets["client"], "rb").read() == avant

    def test_sans_gabarit_ne_leve_pas(self, atelier):
        """Un gabarit manquant est journalise, pas fatal au demarrage."""
        assert bootstrap.ensure_workbooks() == []
        assert len(bootstrap.missing_workbooks()) == 2


class TestDescribeState:
    def test_signale_un_classeur_absent(self, atelier):
        etat = bootstrap.describe_state()
        assert etat["data.xlsx"]["present"] is False
        assert etat["data.xlsx"]["feuilles"] == {}

    def test_compte_les_lignes_de_donnees_hors_entete(self, atelier):
        _fabriquer_classeur(
            atelier["client"],
            {"DEMANDE_CLIENT": (["Date"], [["a"], ["b"], ["c"]])},
        )
        etat = bootstrap.describe_state()
        assert etat["data.xlsx"]["feuilles"]["DEMANDE_CLIENT"] == 3

    def test_signale_les_references_vides(self, atelier):
        """Une feuille de reference vide n'est pas un etat de depart normal."""
        _fabriquer_classeur(
            atelier["hotel"],
            {
                "BDD_HOTEL": (["Hotel"], [["Ville", "HTL"]]),
                "KM_MADA": (["REPERES", "KM"], []),
                "Circuits": (["ID circuit"], [["C1"]]),
            },
        )
        etat = bootstrap.describe_state()["data-hotel.xlsx"]
        assert "KM_MADA" in etat["references_vides"]
        assert "Circuits" not in etat["references_vides"]

    def test_une_feuille_non_referentielle_vide_ne_derange_pas(self, atelier):
        _fabriquer_classeur(atelier["client"], {"DEMANDE_CLIENT": (["Date"], [])})
        assert bootstrap.describe_state()["data.xlsx"]["references_vides"] == []


class TestAllerRetourComplet:
    def test_gabarit_puis_installation_reproduit_la_structure(self, atelier):
        """Le cycle export -> suppression -> init doit etre neutre."""
        feuilles = {
            "DEMANDE_CLIENT": (["Date", "Nom"], [["2026-01-01", "Rakoto"]]),
            "COTATION_H": (["Date", "Hôtel", "Nuits"], [["2026-01-01", "Louvre", 3]]),
        }
        _fabriquer_classeur(atelier["client"], feuilles)
        _fabriquer_classeur(
            atelier["hotel"], {"PARAMETRE": (["parametre", "valeur"], [])}
        )

        bootstrap.export_templates()
        os.remove(atelier["client"])
        bootstrap.ensure_workbooks()

        wb = load_workbook(atelier["client"])
        try:
            assert wb.sheetnames == list(feuilles)
            for nom, (entetes, _) in feuilles.items():
                assert [c.value for c in wb[nom][1]] == entetes
                assert wb[nom].max_row == 1
        finally:
            wb.close()
