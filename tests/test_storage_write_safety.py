"""Garanties d'ecriture de la couche de stockage, vues des fonctions publiques.

Le socle (`utils.storage.workbook`) est teste separement. Ici on verifie que
les fonctions publiques en beneficient reellement : le document de conception
annoncait un backup systematique, mais aucun site d'appel ne le demandait --
`backup=True` n'apparaissait nulle part dans le depot.

Deux garanties sont couvertes :
  1. toute ecriture reussie laisse une copie de securite de l'etat precedent ;
  2. toute ecriture interrompue laisse le fichier d'origine intact.
"""

import os
import shutil

import pytest
from openpyxl import Workbook, load_workbook

import config
from utils import excel_handler
from utils.storage import workbook as wbmod


@pytest.fixture
def classeur_hotel(tmp_path, monkeypatch):
    """Redirige data-hotel.xlsx vers un temporaire, deja peuple d'une ligne."""
    path = str(tmp_path / "data-hotel.xlsx")
    monkeypatch.setattr(excel_handler, "HOTEL_EXCEL_PATH", path)
    monkeypatch.setattr(config, "HOTEL_EXCEL_PATH", path)

    excel_handler.save_circuit_db_row({"Nom": "Circuit A", "Duree": 3})
    assert os.path.exists(path)
    return path


def tmp_path_de(chemin_bak):
    """Chemin .xlsx equivalent, openpyxl refusant de lire une extension .bak."""
    return chemin_bak + ".xlsx"


def _backups(path):
    dossier = os.path.join(os.path.dirname(path), "backups")
    if not os.path.isdir(dossier):
        return []
    return sorted(n for n in os.listdir(dossier) if n.endswith(".bak"))


class TestCopieDeSecurite:
    def test_une_ecriture_laisse_une_copie(self, classeur_hotel):
        avant = _backups(classeur_hotel)
        excel_handler.save_circuit_db_row({"Nom": "Circuit B", "Duree": 5})
        assert len(_backups(classeur_hotel)) == len(avant) + 1

    def test_la_copie_contient_l_etat_precedent(self, classeur_hotel):
        excel_handler.save_circuit_db_row({"Nom": "Circuit B", "Duree": 5})
        assert len(excel_handler.load_circuit_db_rows()) == 2

        copie = _backups(classeur_hotel)[-1]
        chemin = os.path.join(os.path.dirname(classeur_hotel), "backups", copie)
        # openpyxl refuse l'extension .bak : on relit via une copie .xlsx.
        lisible = str(tmp_path_de(chemin))
        shutil.copy2(chemin, lisible)
        wb = load_workbook(lisible)
        try:
            # La copie date d'avant l'ajout : elle ne contient qu'un circuit.
            feuille = wb["Circuits"]
            lignes = [
                r for r in feuille.iter_rows(min_row=2, values_only=True) if any(r)
            ]
            assert len(lignes) == 1
        finally:
            wb.close()

    def test_update_et_delete_sauvegardent_aussi(self, classeur_hotel):
        excel_handler.save_circuit_db_row({"Nom": "Circuit B", "Duree": 5})
        apres_save = len(_backups(classeur_hotel))

        excel_handler.update_circuit_db_row(2, {"Nom": "Circuit A bis", "Duree": 4})
        assert len(_backups(classeur_hotel)) == apres_save + 1

        excel_handler.delete_circuit_db_row(2)
        assert len(_backups(classeur_hotel)) == apres_save + 2

    def test_une_lecture_ne_sauvegarde_pas(self, classeur_hotel):
        avant = _backups(classeur_hotel)
        excel_handler.load_circuit_db_rows()
        assert _backups(classeur_hotel) == avant

    def test_le_nombre_de_copies_est_plafonne(self, classeur_hotel, monkeypatch):
        monkeypatch.setattr(wbmod, "MAX_BACKUPS_PER_FILE", 3)
        for i in range(8):
            excel_handler.save_circuit_db_row({"Nom": f"Circuit {i}", "Duree": i})
        assert len(_backups(classeur_hotel)) == 3


class TestEcritureInterrompue:
    def test_le_classeur_reste_lisible_apres_un_echec(
        self, classeur_hotel, monkeypatch
    ):
        """Une coupure pendant la sauvegarde ne doit pas tronquer le fichier."""
        octets_avant = open(classeur_hotel, "rb").read()

        def saborder(self, chemin):
            with open(chemin, "wb") as f:
                f.write(b"PK\x03\x04 archive tronquee")
            raise OSError("disque plein")

        monkeypatch.setattr(Workbook, "save", saborder)
        resultat = excel_handler.save_circuit_db_row({"Nom": "Circuit B"})

        # La sentinelle d'erreur historique est preservee.
        assert resultat == -1
        assert open(classeur_hotel, "rb").read() == octets_avant
        assert len(excel_handler.load_circuit_db_rows()) == 1

    def test_aucun_temporaire_ne_traine(self, classeur_hotel, monkeypatch):
        def saborder(self, chemin):
            raise OSError("disque plein")

        monkeypatch.setattr(Workbook, "save", saborder)
        excel_handler.save_circuit_db_row({"Nom": "Circuit B"})

        dossier = os.path.dirname(classeur_hotel)
        assert [n for n in os.listdir(dossier) if n.endswith(".tmp")] == []
